from __future__ import annotations

import math
import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook


BASE_XLSX = Path(r"C:\Users\tyboy\OneDrive\Documents\Nasa project\outputs\europa_ice_subsurface_simulation\v8.xlsx")
OUT_XLSX = Path(r"C:\Users\tyboy\OneDrive\Documents\Nasa project\outputs\europa_ice_subsurface_simulation\v9.xlsx")

C_NS = "http://schemas.openxmlformats.org/drawingml/2006/chart"
A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

ET.register_namespace("", C_NS)
ET.register_namespace("a", A_NS)
ET.register_namespace("r", R_NS)

NS = {"c": C_NS}


def clean_title(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def chart_title(root: ET.Element) -> str:
    title = root.find(".//c:chart/c:title//{" + A_NS + "}t", NS)
    return clean_title(title.text if title is not None else "")


def cache_values_from_num_ref(num_ref: ET.Element) -> list[float]:
    cache = num_ref.find("c:numCache", NS)
    if cache is None:
        return []
    values: list[float] = []
    for point in cache.findall("c:pt", NS):
        value = point.find("c:v", NS)
        if value is None or value.text is None:
            continue
        try:
            values.append(float(value.text))
        except ValueError:
            continue
    return values


def chart1_surface_from_cache(xlsx_path: Path) -> tuple[list[float], list[float]]:
    with zipfile.ZipFile(xlsx_path) as archive:
        root = ET.fromstring(archive.read("xl/charts/chart1.xml"))
    series = root.findall(".//c:scatterChart/c:ser", NS)
    if len(series) < 2:
        raise RuntimeError("Could not read the cached topography from the existing Dashboard chart.")

    # The workbook's subsurface live data uses the off-nadir target terrain as the icy surface.
    off_nadir_target = series[1]
    x_ref = off_nadir_target.find("c:xVal/c:numRef", NS)
    y_ref = off_nadir_target.find("c:yVal/c:numRef", NS)
    if x_ref is None or y_ref is None:
        raise RuntimeError("The Dashboard topography chart is missing cached x/y values.")

    x_values = cache_values_from_num_ref(x_ref)
    surface_values = cache_values_from_num_ref(y_ref)
    if not x_values or len(x_values) != len(surface_values):
        raise RuntimeError("The cached topography values are incomplete.")
    return x_values, surface_values


def value_at(ws, cell: str, fallback=None):
    value = ws[cell].value
    if isinstance(value, str) and value.startswith("="):
        return fallback
    return value


def input_values(wb) -> dict[int, float]:
    sub = wb["Subsurface_Inputs"]
    main = wb["Inputs"]
    values: dict[int, float] = {}
    for row in range(4, 58):
        value = sub.cell(row, 3).value
        if isinstance(value, str) and value.startswith("="):
            if row == 34:
                value = main["C12"].value
            else:
                value = 0
        if isinstance(value, str):
            if value.upper() == "TRUE":
                value = 1
            elif value.upper() == "FALSE":
                value = 0
        if value is not None:
            values[row] = float(value)
    return values


def compute_subsurface_series(wb, x: list[float], surface: list[float]) -> dict[str, list[tuple[list[float], list[float]]]]:
    inp = input_values(wb)
    avg_surface = sum(surface) / len(surface)

    upper_depth = []
    upper_elevation = []
    lens_strength = []
    lens_depth = []
    lens_elevation = []
    ocean_depth = []
    ocean_elevation = []

    for xi, surf in zip(x, surface):
        upper = (
            inp[5]
            + inp[6] * math.sin(2 * math.pi * (xi + inp[7]) / inp[8])
            + inp[9] * math.cos(2 * math.pi * xi / inp[10])
            + inp[11] * (surf - avg_surface)
        )
        strength = min(
            1,
            max(
                0,
                inp[13] * math.exp(-0.5 * ((xi - inp[14]) / inp[15]) ** 2)
                + inp[16] * math.exp(-0.5 * ((xi - inp[17]) / inp[18]) ** 2),
            ),
        )
        lens = inp[19] + inp[20] * math.sin(2 * math.pi * (xi - inp[21]) / inp[22]) - inp[23] * strength
        ocean = (
            inp[25]
            + inp[26] * math.sin(2 * math.pi * (xi + inp[27]) / inp[28])
            + inp[29] * math.cos(2 * math.pi * (xi - inp[30]) / inp[31])
            - inp[32] * (surf - avg_surface)
        )
        upper_depth.append(upper)
        upper_elevation.append(surf - upper)
        lens_strength.append(strength)
        lens_depth.append(lens)
        lens_elevation.append(surf - lens)
        ocean_depth.append(ocean)
        ocean_elevation.append(surf - ocean)

    avg_ocean_elevation = sum(ocean_elevation) / len(ocean_elevation)
    n_ice = inp[34]
    speed_of_light = inp[35]
    attenuation = inp[36]
    threshold = inp[43]

    upper_delay = [2 * n_ice * depth / speed_of_light * 1_000_000 for depth in upper_depth]
    lens_delay = [2 * n_ice * depth / speed_of_light * 1_000_000 for depth in lens_depth]
    ocean_delay = [2 * n_ice * depth / speed_of_light * 1_000_000 for depth in ocean_depth]

    lens_echo = [
        inp[38] + inp[39] * strength - 2 * attenuation * (depth / 1000)
        for strength, depth in zip(lens_strength, lens_depth)
    ]
    ocean_echo = [
        inp[40] - 2 * attenuation * (depth / 1000) - inp[41] * abs(elev - avg_ocean_elevation)
        for depth, elev in zip(ocean_depth, ocean_elevation)
    ]
    lens_margin = [value - threshold for value in lens_echo]
    ocean_margin = [value - threshold for value in ocean_echo]

    boundary_low = [max(0, depth - inp[44]) for depth in ocean_depth]
    boundary_high = [depth + inp[44] for depth in ocean_depth]
    thin_shell = [depth * inp[45] for depth in ocean_depth]
    thick_shell = [depth * inp[46] for depth in ocean_depth]
    no_ocean_margin = [inp[47] - threshold + 2 * math.sin(2 * math.pi * xi / 42) for xi in x]
    zero = [0 for _ in x]

    clutter_upper = [inp[57] + 1.5 * math.sin(2 * math.pi * xi / 35) for xi in x]
    shallow_return = [delay + inp[56] * math.sin(2 * math.pi * xi / 31) for xi, delay in zip(x, upper_delay)]
    lens_return = [delay + inp[56] * math.sin(2 * math.pi * xi / 27) for xi, delay in zip(x, lens_delay)]
    ocean_return = [delay + inp[56] * math.sin(2 * math.pi * xi / 47) for xi, delay in zip(x, ocean_delay)]

    material_strength = [-18, -14, -10, -6, -2]
    evidence_scores = [inp[48], inp[49], inp[50], inp[51]]

    return {
        "Subsurface Truth Model: Icy Layers": [
            (x, surface),
            (x, upper_elevation),
            (x, lens_elevation),
            (x, ocean_elevation),
        ],
        "Scenario Comparison: Thin / Medium / Thick Ice": [
            (x, thin_shell),
            (x, ocean_depth),
            (x, thick_shell),
        ],
        "Boundary Uncertainty Band": [
            (x, boundary_low),
            (x, ocean_depth),
            (x, boundary_high),
        ],
        "Ocean Model vs No-Ocean Control": [
            (x, ocean_margin),
            (x, no_ocean_margin),
            (x, zero),
        ],
        "Radargram-Style Return Timing With Clutter": [
            (x, clutter_upper),
            (x, shallow_return),
            (x, lens_return),
            (x, ocean_return),
        ],
        "Detectability Margin vs Threshold": [
            (x, lens_margin),
            (x, ocean_margin),
            (x, zero),
        ],
        "Reflection Strength by Material / Interface": [
            ([1, 2, 3, 4, 5], material_strength),
        ],
        "Cross-Instrument Evidence Score": [
            ([1, 2, 3, 4], evidence_scores),
        ],
    }


def format_number(value: float) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and not math.isfinite(value):
        return ""
    if abs(value) < 1e-12:
        value = 0
    return f"{value:.12g}"


def replace_cache(num_ref: ET.Element, values: list[float]) -> None:
    for old_cache in list(num_ref.findall("c:numCache", NS)):
        num_ref.remove(old_cache)

    cache = ET.Element(f"{{{C_NS}}}numCache")
    format_code = ET.SubElement(cache, f"{{{C_NS}}}formatCode")
    format_code.text = "General"
    count = ET.SubElement(cache, f"{{{C_NS}}}ptCount")
    count.set("val", str(len(values)))
    for idx, value in enumerate(values):
        point = ET.SubElement(cache, f"{{{C_NS}}}pt")
        point.set("idx", str(idx))
        point_value = ET.SubElement(point, f"{{{C_NS}}}v")
        point_value.text = format_number(value)
    num_ref.append(cache)


def patch_chart_xml(xml_bytes: bytes, cache_map: dict[str, list[tuple[list[float], list[float]]]]) -> tuple[bytes, bool]:
    root = ET.fromstring(xml_bytes)
    title = chart_title(root)
    if title not in cache_map:
        return xml_bytes, False

    series_nodes = root.findall(".//c:scatterChart/c:ser", NS)
    series_cache = cache_map[title]
    if len(series_nodes) != len(series_cache):
        raise RuntimeError(f"Series count mismatch for {title}: chart has {len(series_nodes)}, cache has {len(series_cache)}.")

    for series_node, (x_values, y_values) in zip(series_nodes, series_cache):
        x_ref = series_node.find("c:xVal/c:numRef", NS)
        y_ref = series_node.find("c:yVal/c:numRef", NS)
        if x_ref is None or y_ref is None:
            raise RuntimeError(f"Missing x/y reference in chart: {title}")
        replace_cache(x_ref, x_values)
        replace_cache(y_ref, y_values)

    return ET.tostring(root, encoding="utf-8", xml_declaration=False), True


def copy_with_patched_chart_caches(cache_map: dict[str, list[tuple[list[float], list[float]]]]) -> int:
    patched = 0
    with zipfile.ZipFile(BASE_XLSX, "r") as src, zipfile.ZipFile(OUT_XLSX, "w", zipfile.ZIP_DEFLATED) as dst:
        for item in src.infolist():
            data = src.read(item.filename)
            if item.filename.startswith("xl/charts/chart") and item.filename.endswith(".xml"):
                data, did_patch = patch_chart_xml(data, cache_map)
                if did_patch:
                    patched += 1
            dst.writestr(item, data)
    return patched


def main() -> None:
    wb = load_workbook(BASE_XLSX, data_only=False)
    x_values, surface_values = chart1_surface_from_cache(BASE_XLSX)
    cache_map = compute_subsurface_series(wb, x_values, surface_values)
    patched = copy_with_patched_chart_caches(cache_map)
    if patched != 8:
        raise RuntimeError(f"Expected to patch 8 subsurface charts, patched {patched}.")
    print(OUT_XLSX)


if __name__ == "__main__":
    main()
