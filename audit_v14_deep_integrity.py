from __future__ import annotations

import re
import zipfile
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


XLSX = Path(r"C:\Users\tyboy\OneDrive\Documents\Nasa project\outputs\europa_ice_subsurface_simulation\v14.xlsx")


ERROR_TOKENS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NULL!", "#NUM!")
CHART_REF_RE = re.compile(r"^'?(?P<sheet>[^']+?)'?!\$(?P<c1>[A-Z]+)\$(?P<r1>\d+):\$(?P<c2>[A-Z]+)\$(?P<r2>\d+)$")


def chart_title(chart) -> str:
    try:
        return chart.title.tx.rich.p[0].r[0].t or ""
    except Exception:
        return ""


def series_name(series) -> str:
    try:
        return series.tx.v or ""
    except Exception:
        return ""


def numeric_ref(series, axis: str) -> str | None:
    holder = getattr(series, axis, None)
    num_ref = getattr(holder, "numRef", None)
    return getattr(num_ref, "f", None)


def range_values(ws, ref: str):
    match = CHART_REF_RE.match(ref)
    if not match:
        return None
    sheet_name = match.group("sheet")
    min_col, min_row, max_col, max_row = range_boundaries(
        f"{match.group('c1')}{match.group('r1')}:{match.group('c2')}{match.group('r2')}"
    )
    return sheet_name, min_row, max_row, min_col, max_col


def main():
    wb_formula = load_workbook(XLSX, data_only=False, read_only=False)
    wb_values = load_workbook(XLSX, data_only=True, read_only=True)

    issues: list[str] = []
    caveats: list[str] = []
    facts: list[str] = []

    # General workbook error scan.
    visible_errors = []
    formula_error_tokens = []
    external_refs = []
    na_formulas = []
    volatile_formulas = []

    for ws in wb_formula.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    upper = value.upper()
                    if any(token in upper for token in ERROR_TOKENS):
                        formula_error_tokens.append((ws.title, cell.coordinate, value[:160]))
                    if "[" in value or "HTTP" in upper:
                        external_refs.append((ws.title, cell.coordinate, value[:160]))
                    if "NA()" in upper:
                        na_formulas.append((ws.title, cell.coordinate, value[:160]))
                    if any(fn in upper for fn in ("NOW()", "RAND()", "RANDBETWEEN(", "TODAY()")):
                        volatile_formulas.append((ws.title, cell.coordinate, value[:160]))

    for ws in wb_values.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("#"):
                    visible_errors.append((ws.title, cell.coordinate, value))

    if visible_errors:
        issues.append(f"Visible cached spreadsheet errors remain: {len(visible_errors)} examples={visible_errors[:8]}")
    if formula_error_tokens:
        issues.append(f"Formula text contains hard error tokens: {len(formula_error_tokens)} examples={formula_error_tokens[:5]}")
    if external_refs:
        issues.append(f"External workbook/web references found: {len(external_refs)} examples={external_refs[:5]}")
    if volatile_formulas:
        caveats.append(f"Volatile formulas found: {len(volatile_formulas)} examples={volatile_formulas[:5]}")

    na_by_sheet = Counter(sheet for sheet, _, _ in na_formulas)
    if na_by_sheet:
        caveats.append(f"NA() formulas still exist outside visible outputs: {dict(na_by_sheet)}")

    # Chart object and series link checks.
    for sheet_name in ("Dashboard", "Subsurface_Dashboard"):
        ws = wb_formula[sheet_name]
        facts.append(f"{sheet_name}: {len(ws._charts)} native chart objects")
        for idx, chart in enumerate(ws._charts, start=1):
            title = chart_title(chart)
            if not chart.series:
                issues.append(f"{sheet_name} chart {idx} ({title}) has no series.")
                continue
            for sidx, series in enumerate(chart.series, start=1):
                for axis in ("xVal", "yVal"):
                    ref = numeric_ref(series, axis)
                    if not ref:
                        issues.append(f"{sheet_name} chart {idx} ({title}) series {sidx} missing {axis} numeric reference.")
                        continue
                    parsed = range_values(wb_formula, ref)
                    if not parsed:
                        issues.append(f"{sheet_name} chart {idx} ({title}) series {sidx} has unparsed {axis} ref: {ref}")
                        continue
                    ref_sheet, min_row, max_row, min_col, max_col = parsed
                    if ref_sheet not in wb_formula.sheetnames:
                        issues.append(f"{sheet_name} chart {idx} ({title}) points to missing sheet: {ref}")
                        continue
                    f_ws = wb_formula[ref_sheet]
                    v_ws = wb_values[ref_sheet]
                    if max_row > f_ws.max_row or max_col > f_ws.max_column:
                        issues.append(f"{sheet_name} chart {idx} ({title}) points outside used range: {ref}")
                    formula_cells = 0
                    error_cells = 0
                    blank_y_cells = 0
                    for row in range(min_row, max_row + 1):
                        for col in range(min_col, max_col + 1):
                            f_value = f_ws.cell(row, col).value
                            v_value = v_ws.cell(row, col).value
                            if isinstance(f_value, str) and f_value.startswith("="):
                                formula_cells += 1
                            if isinstance(v_value, str) and v_value.startswith("#"):
                                error_cells += 1
                            if axis == "yVal" and v_value in (None, ""):
                                blank_y_cells += 1
                    if sheet_name == "Subsurface_Dashboard" and ref_sheet != "Subsurface_Chart_Data":
                        issues.append(f"Subsurface chart {idx} ({title}) is not linked through Subsurface_Chart_Data: {ref}")
                    if error_cells:
                        issues.append(f"{sheet_name} chart {idx} ({title}) has {error_cells} error values in {axis} range {ref}.")
                    if sheet_name == "Subsurface_Dashboard" and formula_cells == 0:
                        issues.append(f"Subsurface chart {idx} ({title}) {axis} range is not formula-driven: {ref}")
                    if blank_y_cells and "Radargram" not in title:
                        issues.append(f"{sheet_name} chart {idx} ({title}) has unexpected blank y-values in {ref}.")

    # Subsurface dashboard KPI formulas should point to live model sheets.
    subdash = wb_formula["Subsurface_Dashboard"]
    dashboard_formulas = []
    for row in subdash.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                dashboard_formulas.append((cell.coordinate, cell.value))
    bad_dashboard_formulas = [
        (coord, formula)
        for coord, formula in dashboard_formulas
        if not any(sheet in formula for sheet in ("Subsurface_Live_Data", "Subsurface_Inputs", "Subsurface_Materials_Evidence"))
    ]
    if bad_dashboard_formulas:
        issues.append(f"Subsurface dashboard has formulas not tied to subsurface live/input sheets: {bad_dashboard_formulas}")
    facts.append(f"Subsurface_Dashboard formula cells: {len(dashboard_formulas)}")

    # Audit results.
    audit = wb_values["Subsurface_Model_Audit"]
    audit_results = [audit.cell(row, 4).value for row in range(4, audit.max_row + 1)]
    if any(result != "PASS" for result in audit_results):
        issues.append(f"Subsurface_Model_Audit has non-PASS results: {audit_results}")
    else:
        facts.append(f"Subsurface_Model_Audit: {len(audit_results)}/{len(audit_results)} PASS")

    # Integration checks for blank weak-lens returns.
    radar = wb_values["Subsurface_Radargram_Data"]
    live = wb_values["Subsurface_Live_Data"]
    blank_lens_returns = sum(1 for row in range(2, 243) if radar.cell(row, 5).value in (None, ""))
    weak_lens_flags = sum(1 for row in range(2, 243) if live.cell(row, 21).value == "Weak/no lens")
    if blank_lens_returns != weak_lens_flags:
        issues.append(f"Blank lens returns ({blank_lens_returns}) do not match weak-lens flags ({weak_lens_flags}).")
    else:
        facts.append(f"Blank weak-lens returns match weak-lens flags: {blank_lens_returns}")

    # Sheet visibility/presentation caveat.
    helper_sheets = [
        "Subsurface_Chart_Data",
        "Native_Chart_Data",
        "Dashboard_Live_Data",
        "Scenario_Data",
        "Chart_Data",
    ]
    visible_helpers = [name for name in helper_sheets if name in wb_formula.sheetnames and wb_formula[name].sheet_state == "visible"]
    if visible_helpers:
        caveats.append(f"Helper/data sheets are visible: {visible_helpers}. This is fine for auditing, but not presentation-clean.")

    print("FACTS")
    for fact in facts:
        print("-", fact)
    print("ISSUES")
    if issues:
        for issue in issues:
            print("-", issue)
    else:
        print("- none")
    print("CAVEATS")
    if caveats:
        for caveat in caveats:
            print("-", caveat)
    else:
        print("- none")


if __name__ == "__main__":
    main()
