from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


BASE_XLSX = Path(r"C:\Users\tyboy\OneDrive\Documents\Nasa project\outputs\europa_ice_subsurface_simulation\v4.xlsx")
OUT_XLSX = Path(r"C:\Users\tyboy\OneDrive\Documents\Nasa project\outputs\europa_ice_subsurface_simulation\v5.xlsx")


COLORS = {
    "navy": "1F4E79",
    "dark": "1F2937",
    "muted": "526174",
    "blue": "2F6DB3",
    "green": "219653",
    "orange": "E76F24",
    "gold": "D49A22",
    "purple": "7048B8",
    "red": "C43A3A",
    "gray": "808080",
    "light_blue": "D9EAF7",
    "yellow": "FFF2CC",
    "grid": "D4DAE3",
}


QUICK_VALUES = {
    "C5": 1150,
    "C13": 0.95,
    "C14": -24,
    "C16": 0.70,
    "C17": 30,
    "C19": 5100,
    "C25": 15000,
    "C26": 760,
    "C34": "=Inputs!$C$12",
    "C36": 0.9,
    "C42": 0.18,
}


def chart_title(chart):
    try:
        rich = chart.title.tx.rich
        if rich and rich.p and rich.p[0].r:
            return rich.p[0].r[0].t or ""
    except Exception:
        return ""
    return ""


def keep_existing_chart_fix(wb):
    for chart in wb["Dashboard"]._charts:
        if chart_title(chart) == "Terrain Baseline: Total Radar Elevation Error":
            chart.legend = None


def set_calc_mode(wb):
    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass


def delete_sheet(wb, name):
    if name in wb.sheetnames:
        del wb[name]


def style_header(ws, row, min_col, max_col):
    for col in range(min_col, max_col + 1):
        cell = ws.cell(row, col)
        cell.fill = PatternFill("solid", fgColor=COLORS["navy"])
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def add_borders(ws, min_row, max_row, min_col, max_col):
    thin = Side(style="thin", color=COLORS["grid"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def reset_dashboard_input_links(wb):
    inputs = wb["Subsurface_Inputs"]
    for cell, value in QUICK_VALUES.items():
        inputs[cell] = value
        inputs[cell].fill = PatternFill("solid", fgColor=COLORS["yellow"])
    inputs["E2"] = "All editable values live on this tab. Subsurface_Dashboard is output-only."
    inputs["E2"].font = Font(color=COLORS["muted"])
    inputs["E2"].alignment = Alignment(wrap_text=True, vertical="top")

    # Add a detection threshold input used by new detectability margin graphs.
    row = 43
    inputs.cell(row, 1).value = "Radar"
    inputs.cell(row, 2).value = "Detection threshold"
    inputs.cell(row, 3).value = -45
    inputs.cell(row, 4).value = "dB"
    inputs.cell(row, 5).value = "Simple threshold for treating a return as easier to see."
    inputs.cell(row, 3).fill = PatternFill("solid", fgColor=COLORS["yellow"])
    for col in range(1, 6):
        inputs.cell(row, col).alignment = Alignment(wrap_text=True, vertical="top")


def extend_live_data(wb):
    ws = wb["Subsurface_Live_Data"]
    additions = [
        (23, "detection_threshold_db"),
        (24, "lens_echo_margin_db"),
        (25, "ocean_echo_margin_db"),
        (26, "best_deep_return_margin_db"),
        (27, "zero_margin_reference"),
    ]
    for col, header in additions:
        ws.cell(1, col).value = header
        ws.cell(1, col).fill = PatternFill("solid", fgColor=COLORS["navy"])
        ws.cell(1, col).font = Font(bold=True, color="FFFFFF")
        ws.cell(1, col).alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 21

    for row in range(2, 243):
        ws.cell(row, 21).value = f'=IF(H{row}>=Subsurface_Inputs!$C$42,"Lens visible","Weak/no lens")'
        ws.cell(row, 22).value = f'=IF(R{row}>=W{row},"Likely visible","Weak")'
        ws.cell(row, 23).value = "=Subsurface_Inputs!$C$43"
        ws.cell(row, 24).value = f"=Q{row}-W{row}"
        ws.cell(row, 25).value = f"=R{row}-W{row}"
        ws.cell(row, 26).value = f"=MAX(Q{row},R{row})-W{row}"
        ws.cell(row, 27).value = 0
        for col in range(23, 28):
            ws.cell(row, col).number_format = "0.000"
    add_borders(ws, 1, ws.max_row, 1, 27)


def add_series(chart, ws, x_col, y_col, title, color, width=22000):
    xvalues = Reference(ws, min_col=x_col, min_row=2, max_row=242)
    yvalues = Reference(ws, min_col=y_col, min_row=2, max_row=242)
    series = Series(yvalues, xvalues, title=title)
    series.graphicalProperties.line.solidFill = color
    series.graphicalProperties.line.width = width
    chart.series.append(series)


def make_chart(ws_data, title, y_title, specs, width=15.5, height=7.7, legend=True):
    chart = ScatterChart()
    chart.title = title
    chart.scatterStyle = "line"
    chart.width = width
    chart.height = height
    chart.x_axis.title = "Along-track position x (km)"
    chart.y_axis.title = y_title
    if legend:
        chart.legend.position = "b"
    else:
        chart.legend = None
    for y_col, label, color in specs:
        add_series(chart, ws_data, 1, y_col, label, color)
    return chart


def create_clean_dashboard(wb):
    delete_sheet(wb, "Subsurface_Dashboard")
    ws = wb.create_sheet("Subsurface_Dashboard", 1)
    data = wb["Subsurface_Live_Data"]

    ws["A1"] = "Europa Ice Subsurface Live Dashboard"
    ws["A1"].font = Font(bold=True, size=16, color=COLORS["dark"])
    ws["A2"] = "Output-only view: important live metrics and graphs. Change assumptions on Subsurface_Inputs."
    ws["A2"].font = Font(color=COLORS["muted"])

    kpis = [
        ["Important Data", "Value", "Unit", "Why it matters"],
        ["Average ice-ocean boundary depth", "=AVERAGE(Subsurface_Live_Data!K2:K242)", "m", "Mean proposed bottom reflector depth."],
        ["Min ice-ocean boundary depth", "=MIN(Subsurface_Live_Data!K2:K242)", "m", "Shallowest modeled bottom reflector."],
        ["Max ice-ocean boundary depth", "=MAX(Subsurface_Live_Data!K2:K242)", "m", "Deepest modeled bottom reflector."],
        ["Average ocean delay", "=AVERAGE(Subsurface_Live_Data!O2:O242)", "us", "Expected timing after surface return."],
        ["Strongest lens strength", "=MAX(Subsurface_Live_Data!H2:H242)", "0-1", "Peak possible warm/briny lens contrast."],
        ["Best lens echo margin", "=MAX(Subsurface_Live_Data!X2:X242)", "dB", "How far lens return is above/below threshold."],
        ["Best ocean echo margin", "=MAX(Subsurface_Live_Data!Y2:Y242)", "dB", "How far deep reflector is above/below threshold."],
        ["Visible lens samples", '=COUNTIF(Subsurface_Live_Data!U2:U242,"Lens visible")', "count", "How much of the pass has a strong lens flag."],
        ["Likely visible ocean samples", '=COUNTIF(Subsurface_Live_Data!V2:V242,"Likely visible")', "count", "How much of the pass clears detection threshold."],
    ]
    for r_idx, row in enumerate(kpis, start=4):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(r_idx, c_idx).value = value
    style_header(ws, 4, 1, 4)
    add_borders(ws, 4, 13, 1, 4)
    for row in range(5, 14):
        ws.cell(row, 2).number_format = "0.000"

    realism = [
        ["Realism / Proof Strength", "Score", "What this means", "How to improve"],
        ["Concept demonstration realism", 0.65, "Good for showing how radar could respond to layered ice.", "Keep formulas live and compare scenarios."],
        ["Europa prediction realism", 0.35, "Plausible structure, but synthetic and not tied to measured Europa radar/DEM.", "Use real instrument parameters, dielectric ranges, temperature/attenuation models."],
        ["Proof strength for real subsurface claim", 0.20, "This supports a hypothesis; it does not prove a real ocean/lens location.", "Validate against real observations or published constraints."],
        ["Target after upgrades", 0.55, "Could become a stronger model once uncertainty and multiple scenarios are added.", "Add sensitivity bands and noise/clutter cases."],
    ]
    for r_idx, row in enumerate(realism, start=4):
        for c_idx, value in enumerate(row, start=6):
            ws.cell(r_idx, c_idx).value = value
    style_header(ws, 4, 6, 9)
    add_borders(ws, 4, 8, 6, 9)
    for row in range(5, 9):
        ws.cell(row, 7).number_format = "0%"

    ws["F10"] = "Short answer"
    ws["F10"].font = Font(bold=True, color=COLORS["dark"])
    ws["G10"] = "The current simulation is about 35% realistic as a Europa prediction and about 20% strong as proof. It is useful for showing what to look for, not proving the real subsurface by itself."
    ws["G10"].alignment = Alignment(wrap_text=True, vertical="top")
    add_borders(ws, 10, 10, 6, 9)

    charts = [
        (
            "A16",
            make_chart(
                data,
                "Subsurface Truth Model: Icy Layers",
                "Elevation relative to reference (m)",
                [
                    (2, "Icy top surface", COLORS["orange"]),
                    (7, "Shallow internal ice layer", COLORS["green"]),
                    (10, "Possible warm/briny lens", COLORS["gold"]),
                    (12, "Possible ice-ocean boundary", COLORS["purple"]),
                ],
            ),
        ),
        (
            "J16",
            make_chart(
                data,
                "Subsurface Radar Delay After Surface Return",
                "Two-way delay (microseconds)",
                [
                    (13, "Shallow internal ice layer", COLORS["green"]),
                    (14, "Possible warm/briny lens", COLORS["gold"]),
                    (15, "Possible ice-ocean boundary", COLORS["purple"]),
                ],
            ),
        ),
        (
            "A33",
            make_chart(
                data,
                "Estimated Subsurface Echo Strength",
                "Relative echo strength (dB)",
                [
                    (16, "Shallow internal ice layer", COLORS["green"]),
                    (17, "Possible warm/briny lens", COLORS["gold"]),
                    (18, "Possible ice-ocean boundary", COLORS["purple"]),
                ],
            ),
        ),
        (
            "J33",
            make_chart(
                data,
                "Detectability Margin vs Threshold",
                "Margin above threshold (dB)",
                [
                    (24, "Lens echo margin", COLORS["gold"]),
                    (25, "Ocean echo margin", COLORS["purple"]),
                    (27, "Zero margin threshold", COLORS["gray"]),
                ],
            ),
        ),
        (
            "A50",
            make_chart(
                data,
                "Possible Warm/Briny Lens Strength",
                "Lens strength (0-1)",
                [(8, "Lens strength", COLORS["gold"])],
                legend=False,
            ),
        ),
        (
            "J50",
            make_chart(
                data,
                "Ice Shell Thickness Along Track",
                "Depth to possible boundary (m)",
                [(11, "Ice-ocean boundary depth", COLORS["purple"])],
                legend=False,
            ),
        ),
    ]
    for anchor, chart in charts:
        ws.add_chart(chart, anchor)

    for col, width in {
        "A": 34,
        "B": 16,
        "C": 12,
        "D": 54,
        "F": 34,
        "G": 14,
        "H": 50,
        "I": 44,
    }.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A4"


def append_guides(wb):
    fg = wb["Formula_Guide"]
    start = fg.max_row + 2
    rows = [
        ["detectability_margin = echo_dB - detection_threshold_dB", "Positive means the simulated return is above the simple detection threshold.", "Subsurface_Inputs detection threshold, echo estimates", "detectability margin"],
        ["realism/proof scores", "Judgment score for model credibility, not a measured science result.", "synthetic inputs and assumptions", "dashboard realism block"],
    ]
    for r, row in enumerate(rows, start=start):
        for c, value in enumerate(row, start=1):
            fg.cell(r, c).value = value
    add_borders(fg, start, start + len(rows) - 1, 1, 4)

    gg = wb["Graph_Guide"]
    start = gg.max_row + 2
    rows = [
        ["Detectability Margin vs Threshold", "Shows whether lens/ocean echoes clear a simple detection threshold.", "Useful for judging whether the simulated bottom layer would be visible.", "0 dB is the threshold line.", "Subsurface_Live_Data!A:X:Y:AA", "margin = echo - threshold", "See Subsurface_Checks"],
        ["Ice Shell Thickness Along Track", "Shows possible ice-ocean boundary depth as thickness.", "Useful because the mission question is about what is below the icy surface.", "No baseline; line is synthetic proposed thickness.", "Subsurface_Live_Data!A:K", "ocean_depth formula from Subsurface_Inputs", "See Subsurface_Checks"],
    ]
    for r, row in enumerate(rows, start=start):
        for c, value in enumerate(row, start=1):
            gg.cell(r, c).value = value
    add_borders(gg, start, start + len(rows) - 1, 1, 7)

    sf = wb["Subsurface_Formulas"]
    start = sf.max_row + 1
    rows = [
        ["detectability_margin = echo_dB - detection_threshold_dB", "Positive values are easier-to-see simulated returns.", "Subsurface_Inputs C43 and echo formulas", "lens/ocean echo margin"],
        ["realism score", "A judgment estimate: useful for communication, not a scientific proof.", "model assumptions and missing real constraints", "dashboard score"],
    ]
    for row in rows:
        sf.append(row)
    add_borders(sf, start, sf.max_row, 1, 4)


def update_checks(wb):
    ws = wb["Subsurface_Checks"]
    rows = [
        ["Detection threshold numeric", '=IF(ISNUMBER(Subsurface_Inputs!C43),"OK","Review")', "Detection threshold should be a numeric dB value."],
        ["Detectability margins finite", '=IF(COUNT(Subsurface_Live_Data!X2:Y242)=ROWS(Subsurface_Live_Data!X2:Y242)*2,"OK","Review")', "Lens/ocean margins should stay numeric."],
    ]
    for row in rows:
        ws.append(row)
    add_borders(ws, ws.max_row - len(rows) + 1, ws.max_row, 1, 3)


def main():
    wb = load_workbook(BASE_XLSX)
    set_calc_mode(wb)
    keep_existing_chart_fix(wb)
    reset_dashboard_input_links(wb)
    extend_live_data(wb)
    create_clean_dashboard(wb)
    append_guides(wb)
    update_checks(wb)
    wb.save(OUT_XLSX)
    print(OUT_XLSX)


if __name__ == "__main__":
    main()
