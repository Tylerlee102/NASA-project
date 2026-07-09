from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


BASE_XLSX = Path(r"C:\Users\tyboy\OneDrive\Documents\Nasa project\outputs\europa_ice_subsurface_simulation\v7.xlsx")
OUT_XLSX = Path(r"C:\Users\tyboy\OneDrive\Documents\Nasa project\outputs\europa_ice_subsurface_simulation\v8.xlsx")


COLORS = {
    "navy": "1F4E79",
    "dark": "1F2937",
    "muted": "526174",
    "blue": "2F6DB3",
    "green": "219653",
    "orange": "E76F24",
    "gold": "D49A22",
    "purple": "7048B8",
    "red": "C43A3A",
    "gray": "808080",
    "grid": "D4DAE3",
}


def chart_title(chart):
    try:
        rich = chart.title.tx.rich
        if rich and rich.p and rich.p[0].r:
            return rich.p[0].r[0].t or ""
    except Exception:
        return ""
    return ""


def keep_existing_chart_fix(wb):
    for chart in wb["Dashboard"]._charts:
        if chart_title(chart) == "Terrain Baseline: Total Radar Elevation Error":
            chart.legend = None


def set_calc_mode(wb):
    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass


def style_header(ws, row, min_col, max_col):
    for col in range(min_col, max_col + 1):
        cell = ws.cell(row, col)
        cell.fill = PatternFill("solid", fgColor=COLORS["navy"])
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def add_borders(ws, min_row, max_row, min_col, max_col):
    thin = Side(style="thin", color=COLORS["grid"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def set_dashboard_dimensions(ws):
    widths = {
        "A": 31,
        "B": 13,
        "C": 9,
        "D": 34,
        "E": 14,
        "F": 18,
        "G": 14,
        "H": 13,
        "I": 28,
        "J": 16,
        "K": 13,
        "L": 13,
        "M": 13,
        "N": 13,
        "O": 13,
        "P": 13,
        "Q": 13,
        "R": 13,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row, height in {
        1: 21.0,
        2: 28.0,
        4: 15.75,
        6: 26.25,
        7: 26.25,
        8: 26.25,
        9: 26.25,
        10: 26.25,
        11: 26.25,
        13: 15.75,
        14: 26.25,
        15: 26.25,
        16: 26.25,
        19: 15.75,
        20: 26.25,
        21: 26.25,
        22: 26.25,
        23: 26.25,
        25: 15.75,
        26: 13.9,
    }.items():
        ws.row_dimensions[row].height = height


def ensure_material_scatter_helpers(wb):
    ws = wb["Subsurface_Materials_Evidence"]
    ws["G1"] = "material_index"
    ws["H1"] = "material_reflection_strength_db"
    ws["I1"] = "evidence_index"
    ws["J1"] = "evidence_support_score"
    for cell in ws["G1:J1"][0]:
        cell.fill = PatternFill("solid", fgColor=COLORS["navy"])
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(wrap_text=True)
    for idx, row in enumerate(range(2, 7), start=1):
        ws.cell(row, 7).value = idx
        ws.cell(row, 8).value = f"=B{row}"
    for idx, row in enumerate(range(10, 14), start=1):
        ws.cell(row, 9).value = idx
        ws.cell(row, 10).value = f"=B{row}"
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 26
    ws.column_dimensions["I"].width = 16
    ws.column_dimensions["J"].width = 24


def add_line_series(chart, ws, x_col, y_col, label, color, min_row=2, max_row=242):
    xvalues = Reference(ws, min_col=x_col, min_row=min_row, max_row=max_row)
    yvalues = Reference(ws, min_col=y_col, min_row=min_row, max_row=max_row)
    series = Series(yvalues, xvalues, title=label)
    series.graphicalProperties.line.solidFill = color
    series.graphicalProperties.line.width = 22000
    try:
        series.smooth = False
    except Exception:
        pass
    chart.series.append(series)


def dashboard_scatter(title, x_title, y_title, x_min, x_max):
    chart = ScatterChart()
    chart.title = title
    chart.scatterStyle = "lineMarker"
    chart.width = 15
    chart.height = 7.5
    chart.x_axis.title = x_title
    chart.y_axis.title = y_title
    chart.x_axis.scaling.min = x_min
    chart.x_axis.scaling.max = x_max
    chart.x_axis.crosses = "min"
    chart.y_axis.crosses = "min"
    chart.x_axis.majorGridlines = ChartLines()
    chart.y_axis.majorGridlines = ChartLines()
    chart.legend.position = "b"
    return chart


def build_dashboard(wb):
    if "Subsurface_Dashboard" in wb.sheetnames:
        del wb["Subsurface_Dashboard"]
    ws = wb.create_sheet("Subsurface_Dashboard", 1)
    live = wb["Subsurface_Live_Data"]
    scen = wb["Subsurface_Scenario_Data"]
    material = wb["Subsurface_Materials_Evidence"]
    radargram = wb["Subsurface_Radargram_Data"]

    set_dashboard_dimensions(ws)
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    ws["A1"] = "Europa Ice Subsurface Interface"
    ws["A1"].font = Font(bold=True, size=16, color=COLORS["dark"])
    ws.merge_cells("A2:H2")
    ws["A2"] = "Live subsurface outputs from Subsurface_Inputs. The model is synthetic: it helps test what radar might see below icy topography, but it does not prove real Europa layers by itself."
    ws["A2"].font = Font(color=COLORS["muted"])
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")

    ws.merge_cells("A4:D4")
    ws["A4"] = "Key Subsurface Results"
    ws["A4"].font = Font(bold=True, size=12, color=COLORS["dark"])
    left_rows = [
        ["Output", "Value", "Unit", "What it means"],
        ["Average bottom depth", "=AVERAGE(Subsurface_Live_Data!K2:K242)", "m", "Mean proposed ice-ocean boundary depth."],
        ["Boundary uncertainty band", "=Subsurface_Inputs!$C$44", "m", "Plus/minus range around bottom reflector."],
        ["Best ocean echo margin", "=MAX(Subsurface_Live_Data!Y2:Y242)", "dB", "How far deep reflector rises above/below threshold."],
        ["Best lens echo margin", "=MAX(Subsurface_Live_Data!X2:X242)", "dB", "How far internal lens return rises above/below threshold."],
        ["Likely visible ocean samples", '=COUNTIF(Subsurface_Live_Data!V2:V242,"Likely visible")', "count", "Pass samples clearing the detection threshold."],
        ["Total evidence support", "=Subsurface_Materials_Evidence!D14", "%", "Radar + thermal + composition + magnetic/plasma support."],
    ]
    for r_idx, row in enumerate(left_rows, start=5):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(r_idx, c_idx).value = value
    style_header(ws, 5, 1, 4)
    add_borders(ws, 5, 11, 1, 4)
    for row in range(6, 12):
        ws.cell(row, 2).number_format = "0.000"

    ws.merge_cells("F4:I4")
    ws["F4"] = "Realism / Proof Strength"
    ws["F4"].font = Font(bold=True, size=12, color=COLORS["dark"])
    right_rows = [
        ["Score", "Value", "Meaning", "Improve by"],
        ["Concept realism", 0.70, "Good for showing layered-ice radar behavior.", "Compare more scenarios."],
        ["Europa prediction realism", "=MIN(0.75,0.25+Subsurface_Materials_Evidence!D14/100*0.45)", "Plausible, but synthetic and not measured.", "Use published constraints or real mission data."],
        ["Proof strength", "=MIN(0.60,0.15+Subsurface_Materials_Evidence!D14/100*0.35)", "Supports a hypothesis; does not prove a real layer.", "Validate with radar, thermal, composition, and magnetic/plasma data."],
    ]
    for r_idx, row in enumerate(right_rows, start=5):
        for c_idx, value in enumerate(row, start=6):
            ws.cell(r_idx, c_idx).value = value
    style_header(ws, 5, 6, 9)
    add_borders(ws, 5, 8, 6, 9)
    for row in range(6, 9):
        ws.cell(row, 7).number_format = "0%"

    ws.merge_cells("A13:I13")
    ws["A13"] = "Subsurface Graph Guide"
    ws["A13"].font = Font(bold=True, size=12, color=COLORS["dark"])
    graph_rows = [
        ["Graph", "What it shows", "How to read it", None, None, "Graph", "What it shows", "How to read it", None],
        ["Icy layers", "Top surface, internal ice, lens, and possible ocean boundary.", "Lower lines are deeper.", None, None, "Scenarios", "Thin / medium / thick ice shell cases.", "Different assumptions change boundary depth.", None],
        ["Uncertainty", "Mean bottom boundary with lower/upper bounds.", "Band width comes from Subsurface_Inputs.", None, None, "Control", "Ocean model compared with no-ocean control.", "Positive margin clears threshold.", None],
        ["Radargram timing", "Return delay with surface clutter and jitter.", "Later delay means deeper reflector.", None, None, "Detectability", "Lens/ocean echo margin above threshold.", "Above 0 dB is easier to detect.", None],
        ["Materials", "Reflector strength by material/interface.", "Less negative is stronger.", None, None, "Evidence", "Radar, thermal, composition, magnetic/plasma support.", "Higher percent means stronger support.", None],
    ]
    for r_idx, row in enumerate(graph_rows, start=14):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(r_idx, c_idx).value = value
    style_header(ws, 14, 1, 9)
    add_borders(ws, 14, 18, 1, 9)

    charts = []
    chart = dashboard_scatter("Subsurface Truth Model: Icy Layers", "Along-track position x (km)", "Elevation relative to reference (m)", -100, 100)
    for col, label, color in [(2, "Icy top surface", COLORS["orange"]), (7, "Shallow ice layer", COLORS["green"]), (10, "Warm/briny lens", COLORS["gold"]), (12, "Ice-ocean boundary", COLORS["purple"])]:
        add_line_series(chart, live, 1, col, label, color)
    charts.append(("A29", chart))

    chart = dashboard_scatter("Scenario Comparison: Thin / Medium / Thick Ice", "Along-track position x (km)", "Depth to possible boundary (m)", -100, 100)
    for col, label, color in [(5, "Thin shell", COLORS["green"]), (6, "Medium shell", COLORS["purple"]), (7, "Thick shell", COLORS["red"])]:
        add_line_series(chart, scen, 1, col, label, color)
    charts.append(("J29", chart))

    chart = dashboard_scatter("Boundary Uncertainty Band", "Along-track position x (km)", "Depth to possible boundary (m)", -100, 100)
    for col, label, color in [(3, "Lower bound", COLORS["green"]), (2, "Mean boundary", COLORS["purple"]), (4, "Upper bound", COLORS["red"])]:
        add_line_series(chart, scen, 1, col, label, color)
    charts.append(("A50", chart))

    chart = dashboard_scatter("Ocean Model vs No-Ocean Control", "Along-track position x (km)", "Margin above threshold (dB)", -100, 100)
    for col, label, color in [(8, "Ocean model margin", COLORS["purple"]), (9, "No-ocean control margin", COLORS["gray"]), (10, "Zero threshold", COLORS["red"])]:
        add_line_series(chart, scen, 1, col, label, color)
    charts.append(("J50", chart))

    chart = dashboard_scatter("Radargram-Style Return Timing With Clutter", "Along-track position x (km)", "Two-way delay after surface return (us)", -100, 100)
    for col, label, color in [(3, "Surface clutter upper", COLORS["gray"]), (4, "Shallow ice return", COLORS["green"]), (5, "Warm/briny lens return", COLORS["gold"]), (6, "Ocean boundary return", COLORS["purple"])]:
        add_line_series(chart, radargram, 1, col, label, color)
    charts.append(("A71", chart))

    chart = dashboard_scatter("Detectability Margin vs Threshold", "Along-track position x (km)", "Margin above threshold (dB)", -100, 100)
    for col, label, color in [(24, "Lens echo margin", COLORS["gold"]), (25, "Ocean echo margin", COLORS["purple"]), (27, "Zero margin threshold", COLORS["red"])]:
        add_line_series(chart, live, 1, col, label, color)
    charts.append(("J71", chart))

    chart = dashboard_scatter("Reflection Strength by Material / Interface", "Material/interface index", "Reflection strength before depth loss (dB)", 1, 5)
    add_line_series(chart, material, 7, 8, "Material/interface strength", COLORS["blue"], min_row=2, max_row=6)
    chart.legend = None
    charts.append(("A92", chart))

    chart = dashboard_scatter("Cross-Instrument Evidence Score", "Evidence source index", "Support score (%)", 1, 4)
    add_line_series(chart, material, 9, 10, "Evidence support score", COLORS["green"], min_row=10, max_row=13)
    chart.legend = None
    charts.append(("J92", chart))

    for anchor, chart in charts:
        ws.add_chart(chart, anchor)

    ws.freeze_panes = "A5"


def main():
    wb = load_workbook(BASE_XLSX)
    set_calc_mode(wb)
    keep_existing_chart_fix(wb)
    ensure_material_scatter_helpers(wb)
    build_dashboard(wb)
    wb.save(OUT_XLSX)
    print(OUT_XLSX)


if __name__ == "__main__":
    main()
