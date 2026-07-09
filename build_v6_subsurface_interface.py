from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import BarChart, ScatterChart, Series, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


BASE_XLSX = Path(r"C:\Users\tyboy\OneDrive\Documents\Nasa project\outputs\europa_ice_subsurface_simulation\v5.xlsx")
OUT_XLSX = Path(r"C:\Users\tyboy\OneDrive\Documents\Nasa project\outputs\europa_ice_subsurface_simulation\v6.xlsx")


COLORS = {
    "navy": "1F4E79",
    "dark": "1F2937",
    "muted": "526174",
    "blue": "2F6DB3",
    "green": "219653",
    "orange": "E76F24",
    "gold": "D49A22",
    "purple": "7048B8",
    "red": "C43A3A",
    "gray": "808080",
    "light_blue": "D9EAF7",
    "yellow": "FFF2CC",
    "grid": "D4DAE3",
}


def chart_title(chart):
    try:
        rich = chart.title.tx.rich
        if rich and rich.p and rich.p[0].r:
            return rich.p[0].r[0].t or ""
    except Exception:
        return ""
    return ""


def keep_existing_chart_fix(wb):
    for chart in wb["Dashboard"]._charts:
        if chart_title(chart) == "Terrain Baseline: Total Radar Elevation Error":
            chart.legend = None


def set_calc_mode(wb):
    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass


def delete_sheet(wb, name):
    if name in wb.sheetnames:
        del wb[name]


def style_header(ws, row, min_col, max_col):
    for col in range(min_col, max_col + 1):
        cell = ws.cell(row, col)
        cell.fill = PatternFill("solid", fgColor=COLORS["navy"])
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def add_borders(ws, min_row, max_row, min_col, max_col):
    thin = Side(style="thin", color=COLORS["grid"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def ensure_inputs(wb):
    ws = wb["Subsurface_Inputs"]
    extra_rows = [
        (44, "Uncertainty", "Boundary uncertainty band", 1500, "m", "Plus/minus range around the possible ice-ocean boundary."),
        (45, "Scenario", "Thin shell multiplier", 0.55, None, "Creates a thinner-shell scenario from the live depth."),
        (46, "Scenario", "Thick shell multiplier", 1.75, None, "Creates a thicker-shell scenario from the live depth."),
        (47, "Scenario", "No-ocean control echo", -60, "dB", "Weak baseline return for a no-bottom-reflector control case."),
        (48, "Evidence", "Radar support score", 55, "%", "How strongly radar outputs support the proposed subsurface."),
        (49, "Evidence", "Thermal support score", 35, "%", "Proxy for E-THEMIS-style warm-region support."),
        (50, "Evidence", "Composition support score", 40, "%", "Proxy for MISE-style salts/ice/organic context."),
        (51, "Evidence", "Magnetic/plasma support score", 50, "%", "Proxy for ECM/PIMS-style ocean support."),
        (52, "Evidence", "Radar evidence weight", 40, "%", "Radar should carry the most weight here."),
        (53, "Evidence", "Thermal evidence weight", 20, "%", "Thermal evidence is supporting context."),
        (54, "Evidence", "Composition evidence weight", 20, "%", "Surface composition is supporting context."),
        (55, "Evidence", "Magnetic/plasma evidence weight", 20, "%", "Interior/ocean support from non-radar instruments."),
        (56, "Radargram", "Radargram timing jitter", 1.5, "us", "Adds deterministic timing wiggle to radargram-style chart."),
        (57, "Radargram", "Surface clutter band", 12, "us", "Approximate near-surface clutter window."),
    ]
    for row, section, parameter, value, unit, note in extra_rows:
        ws.cell(row, 1).value = section
        ws.cell(row, 2).value = parameter
        ws.cell(row, 3).value = value
        ws.cell(row, 4).value = unit
        ws.cell(row, 5).value = note
        ws.cell(row, 3).fill = PatternFill("solid", fgColor=COLORS["yellow"])
        for col in range(1, 6):
            ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical="top")
    add_borders(ws, 43, 57, 1, 5)


def extend_live_data(wb):
    ws = wb["Subsurface_Live_Data"]
    for col, header in [
        (23, "detection_threshold_db"),
        (24, "lens_echo_margin_db"),
        (25, "ocean_echo_margin_db"),
        (26, "best_deep_return_margin_db"),
        (27, "zero_margin_reference"),
    ]:
        ws.cell(1, col).value = header
        ws.cell(1, col).fill = PatternFill("solid", fgColor=COLORS["navy"])
        ws.cell(1, col).font = Font(bold=True, color="FFFFFF")
    for row in range(2, 243):
        ws.cell(row, 23).value = "=Subsurface_Inputs!$C$43"
        ws.cell(row, 24).value = f"=Q{row}-W{row}"
        ws.cell(row, 25).value = f"=R{row}-W{row}"
        ws.cell(row, 26).value = f"=MAX(Q{row},R{row})-W{row}"
        ws.cell(row, 27).value = 0
    add_borders(ws, 1, ws.max_row, 1, 27)


def make_scenario_data(wb):
    delete_sheet(wb, "Subsurface_Scenario_Data")
    ws = wb.create_sheet("Subsurface_Scenario_Data", wb.sheetnames.index("Subsurface_Live_Data") + 1)
    headers = [
        "x_km",
        "mean_boundary_depth_m",
        "boundary_low_m",
        "boundary_high_m",
        "thin_shell_depth_m",
        "medium_shell_depth_m",
        "thick_shell_depth_m",
        "ocean_model_margin_db",
        "no_ocean_control_margin_db",
        "zero_margin_reference",
    ]
    ws.append(headers)
    style_header(ws, 1, 1, len(headers))
    for row in range(2, 243):
        src = row
        ws.cell(row, 1).value = f"=Subsurface_Live_Data!A{src}"
        ws.cell(row, 2).value = f"=Subsurface_Live_Data!K{src}"
        ws.cell(row, 3).value = f"=MAX(0,B{row}-Subsurface_Inputs!$C$44)"
        ws.cell(row, 4).value = f"=B{row}+Subsurface_Inputs!$C$44"
        ws.cell(row, 5).value = f"=B{row}*Subsurface_Inputs!$C$45"
        ws.cell(row, 6).value = f"=B{row}"
        ws.cell(row, 7).value = f"=B{row}*Subsurface_Inputs!$C$46"
        ws.cell(row, 8).value = f"=Subsurface_Live_Data!Y{src}"
        ws.cell(row, 9).value = f"=Subsurface_Inputs!$C$47-Subsurface_Inputs!$C$43+2*SIN(2*PI()*A{row}/42)"
        ws.cell(row, 10).value = 0
    add_borders(ws, 1, 242, 1, 10)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, 11):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 20
    return ws


def make_material_evidence(wb):
    delete_sheet(wb, "Subsurface_Materials_Evidence")
    ws = wb.create_sheet("Subsurface_Materials_Evidence", wb.sheetnames.index("Subsurface_Scenario_Data") + 1)
    material_rows = [
        ["Material / interface", "Reflection strength before depth loss (dB)", "One-way attenuation (dB/km)", "Why it matters"],
        ["Cold clean ice layer", -18, 0.5, "Weak internal contrast; often harder to see."],
        ["Warmer ice boundary", -14, 0.8, "Temperature/structure contrast can strengthen a reflector."],
        ["Dirty/salty ice layer", -10, 1.2, "Impurities can create stronger dielectric contrast and more loss."],
        ["Briny pocket/lens", -6, 2.0, "High contrast but can be lossy."],
        ["Liquid salty ocean boundary", -2, 0.0, "Strong material contrast at the bottom boundary."],
    ]
    for row in material_rows:
        ws.append(row)
    style_header(ws, 1, 1, 4)
    add_borders(ws, 1, 6, 1, 4)

    start = 9
    evidence_rows = [
        ["Evidence source", "Support score (%)", "Weight (%)", "Weighted contribution", "Notes"],
        ["Radar sounding", "=Subsurface_Inputs!$C$48", "=Subsurface_Inputs!$C$52", "=B10*C10/100", "REASON-like radar structure support."],
        ["Thermal context", "=Subsurface_Inputs!$C$49", "=Subsurface_Inputs!$C$53", "=B11*C11/100", "E-THEMIS-style warm-region support."],
        ["Surface composition", "=Subsurface_Inputs!$C$50", "=Subsurface_Inputs!$C$54", "=B12*C12/100", "MISE-style salts/ice/organics support."],
        ["Magnetic/plasma ocean support", "=Subsurface_Inputs!$C$51", "=Subsurface_Inputs!$C$55", "=B13*C13/100", "ECM/PIMS-style ocean support."],
        ["Total weighted support", None, "=SUM(C10:C13)", "=SUM(D10:D13)/MAX(1,C14)*100", "Combined evidence score."],
    ]
    for r_idx, row in enumerate(evidence_rows, start=start):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(r_idx, c_idx).value = value
    style_header(ws, start, 1, 5)
    add_borders(ws, start, start + len(evidence_rows) - 1, 1, 5)
    for col, width in {"A": 30, "B": 22, "C": 16, "D": 22, "E": 56}.items():
        ws.column_dimensions[col].width = width
    return ws


def make_radargram_data(wb):
    delete_sheet(wb, "Subsurface_Radargram_Data")
    ws = wb.create_sheet("Subsurface_Radargram_Data", wb.sheetnames.index("Subsurface_Materials_Evidence") + 1)
    headers = [
        "x_km",
        "surface_clutter_lower_us",
        "surface_clutter_upper_us",
        "shallow_return_delay_with_jitter_us",
        "lens_return_delay_with_jitter_us",
        "ocean_return_delay_with_jitter_us",
        "noise_floor_proxy",
    ]
    ws.append(headers)
    style_header(ws, 1, 1, len(headers))
    for row in range(2, 243):
        ws.cell(row, 1).value = f"=Subsurface_Live_Data!A{row}"
        ws.cell(row, 2).value = 0
        ws.cell(row, 3).value = f"=Subsurface_Inputs!$C$57+1.5*SIN(2*PI()*A{row}/35)"
        ws.cell(row, 4).value = f"=Subsurface_Live_Data!M{row}+Subsurface_Inputs!$C$56*SIN(2*PI()*A{row}/31)"
        ws.cell(row, 5).value = f'=IF(Subsurface_Live_Data!H{row}>=Subsurface_Inputs!$C$42,Subsurface_Live_Data!N{row}+Subsurface_Inputs!$C$56*SIN(2*PI()*A{row}/27),NA())'
        ws.cell(row, 6).value = f"=Subsurface_Live_Data!O{row}+Subsurface_Inputs!$C$56*SIN(2*PI()*A{row}/47)"
        ws.cell(row, 7).value = f"=-55+3*SIN(2*PI()*A{row}/19)"
    add_borders(ws, 1, 242, 1, 7)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, 8):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 22
    return ws


def add_line_series(chart, ws, x_col, y_col, label, color, width=22000):
    xvalues = Reference(ws, min_col=x_col, min_row=2, max_row=242)
    yvalues = Reference(ws, min_col=y_col, min_row=2, max_row=242)
    series = Series(yvalues, xvalues, title=label)
    series.graphicalProperties.line.solidFill = color
    series.graphicalProperties.line.width = width
    chart.series.append(series)


def scatter_chart(title, y_title, ws, specs, legend=True, width=15.5, height=7.7):
    chart = ScatterChart()
    chart.title = title
    chart.scatterStyle = "line"
    chart.width = width
    chart.height = height
    chart.x_axis.title = "Along-track position x (km)"
    chart.y_axis.title = y_title
    if legend:
        chart.legend.position = "b"
    else:
        chart.legend = None
    for col, label, color in specs:
        add_line_series(chart, ws, 1, col, label, color)
    return chart


def bar_chart(title, y_title, ws, cats_col, values_col, min_row, max_row, color):
    chart = BarChart()
    chart.title = title
    chart.type = "bar"
    chart.style = 10
    chart.width = 15.5
    chart.height = 7.7
    chart.y_axis.title = y_title
    chart.legend = None
    data = Reference(ws, min_col=values_col, min_row=min_row - 1, max_row=max_row)
    cats = Reference(ws, min_col=cats_col, min_row=min_row, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    if chart.series:
        chart.series[0].graphicalProperties.solidFill = color
    return chart


def create_dashboard(wb):
    delete_sheet(wb, "Subsurface_Dashboard")
    ws = wb.create_sheet("Subsurface_Dashboard", 1)
    live = wb["Subsurface_Live_Data"]
    scen = wb["Subsurface_Scenario_Data"]
    material = wb["Subsurface_Materials_Evidence"]
    radargram = wb["Subsurface_Radargram_Data"]

    ws["A1"] = "Europa Ice Subsurface Interface"
    ws["A1"].font = Font(bold=True, size=16, color=COLORS["dark"])
    ws["A2"] = "Output-only view: key data, realism/proof score, and live graphs. Change assumptions on Subsurface_Inputs."
    ws["A2"].font = Font(color=COLORS["muted"])

    kpis = [
        ["Important Data", "Value", "Unit", "Why it matters"],
        ["Average ice-ocean boundary depth", "=AVERAGE(Subsurface_Live_Data!K2:K242)", "m", "Mean proposed bottom reflector depth."],
        ["Boundary uncertainty band", "=Subsurface_Inputs!$C$44", "m", "Likely plus/minus range around bottom reflector."],
        ["Best ocean echo margin", "=MAX(Subsurface_Live_Data!Y2:Y242)", "dB", "How far deep reflector is above/below threshold."],
        ["Best lens echo margin", "=MAX(Subsurface_Live_Data!X2:X242)", "dB", "How far lens return is above/below threshold."],
        ["Likely visible ocean samples", '=COUNTIF(Subsurface_Live_Data!V2:V242,"Likely visible")', "count", "How much of the pass clears detection threshold."],
        ["Total evidence support", "=Subsurface_Materials_Evidence!D14", "%", "Radar + thermal + composition + magnetic/plasma support."],
    ]
    for r_idx, row in enumerate(kpis, start=4):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(r_idx, c_idx).value = value
    style_header(ws, 4, 1, 4)
    add_borders(ws, 4, 10, 1, 4)
    for row in range(5, 11):
        ws.cell(row, 2).number_format = "0.000"

    realism = [
        ["Realism / Proof Strength", "Score", "What this means", "How to improve"],
        ["Concept demonstration realism", 0.70, "Good for showing radar response to layered ice and clutter.", "Compare more scenarios and uncertainty bands."],
        ["Europa prediction realism", "=MIN(0.75,0.25+Subsurface_Materials_Evidence!D14/100*0.45)", "More realistic than v5 because it includes scenarios, material differences, and evidence inputs.", "Tie values to published constraints or real mission data."],
        ["Proof strength for real subsurface claim", "=MIN(0.60,0.15+Subsurface_Materials_Evidence!D14/100*0.35)", "Still a hypothesis test, not proof by itself.", "Validate against measured radar, thermal, composition, and magnetic/plasma observations."],
    ]
    for r_idx, row in enumerate(realism, start=4):
        for c_idx, value in enumerate(row, start=6):
            ws.cell(r_idx, c_idx).value = value
    style_header(ws, 4, 6, 9)
    add_borders(ws, 4, 7, 6, 9)
    for row in range(5, 8):
        ws.cell(row, 7).number_format = "0%"

    chart_specs = [
        (
            "A13",
            scatter_chart(
                "Subsurface Truth Model: Icy Layers",
                "Elevation relative to reference (m)",
                live,
                [(2, "Icy top surface", COLORS["orange"]), (7, "Shallow ice layer", COLORS["green"]), (10, "Warm/briny lens", COLORS["gold"]), (12, "Ice-ocean boundary", COLORS["purple"])],
            ),
        ),
        (
            "J13",
            scatter_chart(
                "Scenario Comparison: Thin / Medium / Thick Ice",
                "Depth to possible boundary (m)",
                scen,
                [(5, "Thin shell", COLORS["green"]), (6, "Medium shell", COLORS["purple"]), (7, "Thick shell", COLORS["red"])],
            ),
        ),
        (
            "A30",
            scatter_chart(
                "Boundary Uncertainty Band",
                "Depth to possible boundary (m)",
                scen,
                [(3, "Lower bound", COLORS["green"]), (2, "Mean boundary", COLORS["purple"]), (4, "Upper bound", COLORS["red"])],
            ),
        ),
        (
            "J30",
            scatter_chart(
                "Ocean Model vs No-Ocean Control",
                "Margin above threshold (dB)",
                scen,
                [(8, "Ocean model margin", COLORS["purple"]), (9, "No-ocean control margin", COLORS["gray"]), (10, "Zero threshold", COLORS["red"])],
            ),
        ),
        (
            "A47",
            scatter_chart(
                "Radargram-Style Return Timing With Clutter",
                "Two-way delay after surface return (us)",
                radargram,
                [(3, "Surface clutter upper", COLORS["gray"]), (4, "Shallow ice return", COLORS["green"]), (5, "Warm/briny lens return", COLORS["gold"]), (6, "Ocean boundary return", COLORS["purple"])],
            ),
        ),
        (
            "J47",
            scatter_chart(
                "Detectability Margin vs Threshold",
                "Margin above threshold (dB)",
                live,
                [(24, "Lens echo margin", COLORS["gold"]), (25, "Ocean echo margin", COLORS["purple"]), (27, "Zero margin threshold", COLORS["red"])],
            ),
        ),
        (
            "A64",
            bar_chart(
                "Reflection Strength by Material / Interface",
                "Pre-depth-loss reflection strength (dB)",
                material,
                1,
                2,
                2,
                6,
                COLORS["blue"],
            ),
        ),
        (
            "J64",
            bar_chart(
                "Cross-Instrument Evidence Score",
                "Support score (%)",
                material,
                1,
                2,
                10,
                13,
                COLORS["green"],
            ),
        ),
    ]
    for anchor, chart in chart_specs:
        ws.add_chart(chart, anchor)

    for col, width in {"A": 34, "B": 18, "C": 12, "D": 54, "F": 34, "G": 14, "H": 58, "I": 50}.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A4"


def append_guides(wb):
    fg = wb["Formula_Guide"]
    start = fg.max_row + 2
    rows = [
        ["scenario_depth = mean_depth * scenario_multiplier", "Thin/medium/thick ice comparison.", "Subsurface_Inputs scenario multipliers", "scenario boundary depths"],
        ["uncertainty band = boundary_depth +/- uncertainty_m", "Shows a range instead of one exact proposed boundary.", "Subsurface_Inputs boundary uncertainty", "low/high depth bounds"],
        ["no_ocean_control_margin = no_ocean_echo - detection_threshold + clutter", "Control case for comparing against the ocean model.", "no-ocean echo, detection threshold", "control margin"],
        ["evidence_score = weighted average of radar, thermal, composition, magnetic/plasma support", "Cross-instrument support estimate.", "Subsurface_Inputs evidence values and weights", "total evidence support"],
        ["radargram_delay_with_jitter = layer_delay + timing_jitter*sin(...)", "Simple radargram-style noisy timing view.", "layer delays and radargram jitter", "noisy return timing"],
    ]
    for r, row in enumerate(rows, start=start):
        for c, value in enumerate(row, start=1):
            fg.cell(r, c).value = value
    add_borders(fg, start, start + len(rows) - 1, 1, 4)

    gg = wb["Graph_Guide"]
    start = gg.max_row + 2
    rows = [
        ["Scenario Comparison: Thin / Medium / Thick Ice", "Compares possible ice shell thickness cases.", "Shows whether conclusions depend on one chosen depth.", "Medium is the live model; thin/thick are scenario multipliers.", "Subsurface_Scenario_Data!A:E:G", "scenario_depth = mean_depth*multiplier", "Live from Subsurface_Inputs"],
        ["Boundary Uncertainty Band", "Shows lower/mean/upper possible bottom reflector depths.", "Avoids pretending the boundary is exactly known.", "Mean is the live model.", "Subsurface_Scenario_Data!A:B:D", "mean +/- uncertainty", "Live from Subsurface_Inputs"],
        ["Ocean Model vs No-Ocean Control", "Compares deep-reflector detectability against a no-ocean control.", "A proposed ocean model should separate from the control.", "0 dB is threshold.", "Subsurface_Scenario_Data!A:H:J", "margin = echo - threshold", "Live from Subsurface_Inputs"],
        ["Radargram-Style Return Timing With Clutter", "Shows surface clutter, shallow return, lens return, and deep return timing.", "Closer to what radar interpretation feels like than clean depth lines.", "Delay is after surface return.", "Subsurface_Radargram_Data!A:F", "layer delay + jitter", "Live from Subsurface_Inputs"],
        ["Reflection Strength by Material / Interface", "Compares relative reflector strength of likely ice/ocean materials.", "Material contrast matters as much as shape.", "Less negative is stronger.", "Subsurface_Materials_Evidence!A:B", "lookup assumptions", "Editable through table if needed"],
        ["Cross-Instrument Evidence Score", "Radar, thermal, composition, and magnetic/plasma support inputs.", "Makes the model more realistic by not relying on radar alone.", "Higher percent means stronger support.", "Subsurface_Materials_Evidence!A:B", "weighted evidence score", "Live from Subsurface_Inputs"],
    ]
    for r, row in enumerate(rows, start=start):
        for c, value in enumerate(row, start=1):
            gg.cell(r, c).value = value
    add_borders(gg, start, start + len(rows) - 1, 1, 7)


def update_checks(wb):
    ws = wb["Subsurface_Checks"]
    rows = [
        ["Boundary uncertainty numeric", '=IF(ISNUMBER(Subsurface_Inputs!C44),"OK","Review")', "Uncertainty band should be numeric."],
        ["Evidence weights positive", '=IF(SUM(Subsurface_Inputs!C52:C55)>0,"OK","Review")', "Evidence score needs positive weights."],
        ["Scenario data present", '=IF(COUNT(Subsurface_Scenario_Data!E2:G242)=ROWS(Subsurface_Scenario_Data!E2:G242)*3,"OK","Review")', "Scenario comparison should remain numeric."],
        ["Radargram data present", '=IF(COUNT(Subsurface_Radargram_Data!C2:D242)>0,"OK","Review")', "Radargram timing helper should remain numeric."],
    ]
    start = ws.max_row + 1
    for row in rows:
        ws.append(row)
    add_borders(ws, start, ws.max_row, 1, 3)


def main():
    wb = load_workbook(BASE_XLSX)
    set_calc_mode(wb)
    keep_existing_chart_fix(wb)
    ensure_inputs(wb)
    extend_live_data(wb)
    make_scenario_data(wb)
    make_material_evidence(wb)
    make_radargram_data(wb)
    create_dashboard(wb)
    append_guides(wb)
    update_checks(wb)
    wb.save(OUT_XLSX)
    print(OUT_XLSX)


if __name__ == "__main__":
    main()
