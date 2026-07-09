from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


BASE_XLSX = Path(r"C:\Users\tyboy\OneDrive\Documents\Nasa project\outputs\europa_ice_subsurface_simulation\v3.xlsx")
OUT_XLSX = Path(r"C:\Users\tyboy\OneDrive\Documents\Nasa project\outputs\europa_ice_subsurface_simulation\v4.xlsx")


COLORS = {
    "navy": "1F4E79",
    "dark": "1F2937",
    "muted": "526174",
    "light_blue": "D9EAF7",
    "yellow": "FFF2CC",
    "grid": "D4DAE3",
}


def chart_title(chart):
    try:
        rich = chart.title.tx.rich
        if rich and rich.p and rich.p[0].r:
            return rich.p[0].r[0].t or ""
    except Exception:
        return ""
    return ""


def keep_existing_chart_fix(wb):
    for chart in wb["Dashboard"]._charts:
        if chart_title(chart) == "Terrain Baseline: Total Radar Elevation Error":
            chart.legend = None


def add_borders(ws, cell_range):
    thin = Side(style="thin", color=COLORS["grid"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws[cell_range]:
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def main():
    wb = load_workbook(BASE_XLSX)
    keep_existing_chart_fix(wb)
    dash = wb["Subsurface_Dashboard"]
    inputs = wb["Subsurface_Inputs"]

    dash["F4"] = "Quick Editable Inputs"
    dash["F4"].font = Font(bold=True, size=13, color=COLORS["dark"])
    dash["F5"] = "Change the yellow values here. These feed the live data and charts; advanced controls remain on Subsurface_Inputs."
    dash["F5"].font = Font(color=COLORS["muted"])
    dash["F5"].alignment = Alignment(wrap_text=True, vertical="top")

    headers = ["Control", "Value", "Unit", "What it changes"]
    for col, header in enumerate(headers, start=6):
        cell = dash.cell(6, col)
        cell.value = header
        cell.fill = PatternFill("solid", fgColor=COLORS["navy"])
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    quick_inputs = [
        ("Upper ice layer mean depth", 1150, "m", "Moves the shallow internal ice reflector deeper/shallower.", "C5"),
        ("Lens A strength", 0.95, "0-1", "Controls the first possible warm/briny lens contrast.", "C13"),
        ("Lens A center", -24, "km", "Moves lens A left/right along the pass.", "C14"),
        ("Lens B strength", 0.70, "0-1", "Controls the second possible warm/briny lens contrast.", "C16"),
        ("Lens B center", 30, "km", "Moves lens B left/right along the pass.", "C17"),
        ("Lens mean depth", 5100, "m", "Moves the possible warm/briny lens deeper/shallower.", "C19"),
        ("Nominal ice shell thickness", 15000, "m", "Moves the possible ice-ocean boundary deeper/shallower.", "C25"),
        ("Ocean boundary relief amplitude", 760, "m", "Changes how wavy the bottom reflector is.", "C26"),
        ("Ice refractive index", "=Inputs!$C$12", None, "Controls radar delay conversion through ice.", "C34"),
        ("One-way ice attenuation", 0.9, "dB/km", "Controls how quickly deep radar returns weaken.", "C36"),
        ("Lens display threshold", 0.18, "0-1", "Controls when the lens is labeled visible.", "C42"),
    ]

    for row_idx, (label, value, unit, note, input_cell) in enumerate(quick_inputs, start=7):
        dash.cell(row_idx, 6).value = label
        dash.cell(row_idx, 7).value = value
        dash.cell(row_idx, 8).value = unit
        dash.cell(row_idx, 9).value = note
        dash.cell(row_idx, 7).fill = PatternFill("solid", fgColor=COLORS["yellow"])
        dash.cell(row_idx, 7).number_format = "0.000"
        inputs[input_cell] = f"=Subsurface_Dashboard!G{row_idx}"
        inputs[input_cell].fill = PatternFill("solid", fgColor=COLORS["light_blue"])

    add_borders(dash, "F6:I17")
    dash.column_dimensions["F"].width = 34
    dash.column_dimensions["G"].width = 18
    dash.column_dimensions["H"].width = 12
    dash.column_dimensions["I"].width = 64
    for row in range(7, 18):
        dash.row_dimensions[row].height = 30

    inputs["E2"] = "Quick dashboard controls in Subsurface_Dashboard feed several key inputs on this tab."
    inputs["E2"].font = Font(color=COLORS["muted"])
    inputs["E2"].alignment = Alignment(wrap_text=True, vertical="top")

    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    wb.save(OUT_XLSX)
    print(OUT_XLSX)


if __name__ == "__main__":
    main()
