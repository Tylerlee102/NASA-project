from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


BASE_XLSX = Path(r"C:\Users\tyboy\OneDrive\Documents\Nasa project\outputs\europa_ice_subsurface_simulation\v2.xlsx")
OUT_XLSX = Path(r"C:\Users\tyboy\OneDrive\Documents\Nasa project\outputs\europa_ice_subsurface_simulation\v3.xlsx")


SUBSURFACE_SHEETS = [
    "Subsurface_Dashboard",
    "Subsurface_Inputs",
    "Subsurface_Live_Data",
    "Subsurface_Formulas",
    "Subsurface_Checks",
    "Subsurface_Model",
    "Subsurface_Assumptions",
    "Subsurface_Graphs",
]


COLORS = {
    "navy": "1F4E79",
    "dark": "1F2937",
    "blue": "2F6DB3",
    "green": "219653",
    "orange": "E76F24",
    "gold": "D49A22",
    "purple": "7048B8",
    "red": "C43A3A",
    "light_blue": "D9EAF7",
    "light_green": "E6F4EA",
    "light_yellow": "FFF2CC",
    "light_red": "FCE4D6",
    "grid": "D4DAE3",
    "muted": "526174",
}


def delete_if_exists(wb, name):
    if name in wb.sheetnames:
        del wb[name]


def set_calc_mode(wb):
    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass


def style_title(ws, cell, title, subtitle=None):
    ws[cell] = title
    ws[cell].font = Font(bold=True, size=16, color=COLORS["dark"])
    if subtitle:
        row = ws[cell].row + 1
        col = ws[cell].column
        ws.cell(row, col).value = subtitle
        ws.cell(row, col).font = Font(color=COLORS["muted"])
        ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical="top")


def style_header_row(ws, row, start_col=1, end_col=None, fill=COLORS["navy"]):
    end_col = end_col or ws.max_column
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(bold=True, color="FFFFFF" if fill == COLORS["navy"] else COLORS["dark"])
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def add_borders(ws, min_row, max_row, min_col, max_col):
    thin = Side(style="thin", color=COLORS["grid"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def chart_title(chart):
    try:
        rich = chart.title.tx.rich
        if rich and rich.p and rich.p[0].r:
            return rich.p[0].r[0].t or ""
    except Exception:
        return ""
    return ""


def ensure_fixed_existing_dashboard(wb):
    dashboard = wb["Dashboard"]
    for chart in dashboard._charts:
        if chart_title(chart) == "Terrain Baseline: Total Radar Elevation Error":
            chart.legend = None


def make_inputs(wb):
    ws = wb.create_sheet("Subsurface_Inputs", 2)
    style_title(
        ws,
        "A1",
        "Europa Ice Subsurface Editable Inputs",
        "Change the yellow values in column C. Subsurface_Live_Data and Subsurface_Dashboard update from these formulas.",
    )
    rows = [
        ["Section", "Parameter", "Value", "Unit", "Notes"],
        ["Surface", "Use generated icy topography", "TRUE", None, "Surface/topography is modeled as icy terrain, not rock."],
        ["Shallow ice", "Upper ice layer mean depth", 1150, "m", "Depth below the local icy surface."],
        ["Shallow ice", "Upper ice sine amplitude", 160, "m", "Smooth undulation in shallow internal ice."],
        ["Shallow ice", "Upper ice phase shift", 10, "km", "Moves the shallow layer pattern left/right."],
        ["Shallow ice", "Upper ice sine wavelength", 42, "km", "Longer number means smoother/larger features."],
        ["Shallow ice", "Upper ice cosine amplitude", 60, "m", "Secondary shallow-layer variation."],
        ["Shallow ice", "Upper ice cosine wavelength", 23, "km", "Secondary shallow-layer spacing."],
        ["Shallow ice", "Surface coupling factor", 0.04, None, "How much surface relief carries into shallow ice."],
        [None, None, None, None, None],
        ["Warm/briny lens", "Lens A strength", 0.95, "0-1", "Local possible warm/briny contrast, not proven."],
        ["Warm/briny lens", "Lens A center", -24, "km", "Along-track center of first possible lens."],
        ["Warm/briny lens", "Lens A width", 10, "km", "Controls how broad lens A is."],
        ["Warm/briny lens", "Lens B strength", 0.70, "0-1", "Second possible lens."],
        ["Warm/briny lens", "Lens B center", 30, "km", "Along-track center of second possible lens."],
        ["Warm/briny lens", "Lens B width", 8, "km", "Controls how broad lens B is."],
        ["Warm/briny lens", "Lens mean depth", 5100, "m", "Depth below local surface."],
        ["Warm/briny lens", "Lens depth wave amplitude", 520, "m", "Broad depth variation."],
        ["Warm/briny lens", "Lens depth phase shift", 6, "km", "Moves lens-depth wave left/right."],
        ["Warm/briny lens", "Lens depth wavelength", 82, "km", "Lens-depth wave spacing."],
        ["Warm/briny lens", "Lens uplift when strong", 240, "m", "Strong lens sits slightly shallower."],
        [None, None, None, None, None],
        ["Ice-ocean boundary", "Nominal ice shell thickness", 15000, "m", "Possible bottom reflector depth."],
        ["Ice-ocean boundary", "Ocean boundary sine amplitude", 760, "m", "Broad thickness variation."],
        ["Ice-ocean boundary", "Ocean boundary sine phase", 26, "km", "Moves basal pattern left/right."],
        ["Ice-ocean boundary", "Ocean boundary sine wavelength", 135, "km", "Large-scale basal relief."],
        ["Ice-ocean boundary", "Ocean boundary cosine amplitude", 330, "m", "Secondary basal relief."],
        ["Ice-ocean boundary", "Ocean boundary cosine phase", 8, "km", "Moves secondary basal pattern."],
        ["Ice-ocean boundary", "Ocean boundary cosine wavelength", 62, "km", "Secondary basal spacing."],
        ["Ice-ocean boundary", "Surface anti-coupling factor", 0.14, None, "Surface highs slightly reduce thickness."],
        [None, None, None, None, None],
        ["Radar", "Ice refractive index", "=Inputs!$C$12", None, "Uses workbook ice index by default."],
        ["Radar", "Speed of light", 299792458, "m/s", "Used for two-way delay."],
        ["Radar", "One-way ice attenuation", 0.9, "dB/km", "Simple signal loss through ice."],
        ["Radar", "Shallow layer base echo", -10, "dB", "Relative reflector strength before attenuation."],
        ["Radar", "Lens base echo", -24, "dB", "Relative lens strength before lens multiplier and attenuation."],
        ["Radar", "Lens strength echo bonus", 9, "dB", "A stronger lens creates a stronger reflector."],
        ["Radar", "Ocean boundary base echo", -6, "dB", "Large material contrast from ice to possible ocean."],
        ["Radar", "Basal roughness penalty", 0.004, "dB/m", "Rougher boundary weakens coherent echo."],
        ["Radar", "Lens display threshold", 0.18, "0-1", "Below this, lens is treated as too weak for chart focus."],
    ]
    for row in rows:
        ws.append(row)

    style_header_row(ws, 3, 1, 5)
    add_borders(ws, 3, ws.max_row, 1, 5)
    for row in range(4, ws.max_row + 1):
        if ws.cell(row, 2).value:
            ws.cell(row, 3).fill = PatternFill("solid", fgColor=COLORS["light_yellow"])
    for col, width in {"A": 22, "B": 34, "C": 18, "D": 12, "E": 72}.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A5"
    return ws


def make_live_data(wb):
    ws = wb.create_sheet("Subsurface_Live_Data", 4)
    headers = [
        "x_km",
        "surface_height_m",
        "nadir_height_m",
        "spacecraft_altitude_km",
        "target_y_km",
        "upper_ice_layer_depth_m",
        "upper_ice_layer_elevation_m",
        "lens_strength_0to1",
        "briny_warm_lens_depth_m",
        "briny_warm_lens_elevation_m",
        "ice_ocean_boundary_depth_m",
        "ice_ocean_boundary_elevation_m",
        "upper_ice_layer_delay_us",
        "briny_warm_lens_delay_us",
        "ice_ocean_boundary_delay_us",
        "upper_ice_layer_echo_db",
        "briny_warm_lens_echo_db",
        "ice_ocean_boundary_echo_db",
        "ice_thickness_variation_m",
        "ocean_to_shallow_depth_ratio",
        "lens_visible_flag",
        "ocean_echo_flag",
    ]
    ws.append(headers)
    style_header_row(ws, 1, 1, len(headers))

    for r in range(2, 243):
        src = r
        formulas = [
            f"=Model_Data!A{src}",
            f"=Model_Data!Y{src}",
            f"=Model_Data!X{src}",
            f"=Model_Data!B{src}",
            "=Inputs!$C$6",
            (
                "=Subsurface_Inputs!$C$5"
                "+Subsurface_Inputs!$C$6*SIN(2*PI()*(A{r}+Subsurface_Inputs!$C$7)/Subsurface_Inputs!$C$8)"
                "+Subsurface_Inputs!$C$9*COS(2*PI()*A{r}/Subsurface_Inputs!$C$10)"
                "+Subsurface_Inputs!$C$11*(B{r}-AVERAGE($B$2:$B$242))"
            ).format(r=r),
            f"=B{r}-F{r}",
            (
                "=MIN(1,MAX(0,"
                "Subsurface_Inputs!$C$13*EXP(-0.5*((A{r}-Subsurface_Inputs!$C$14)/Subsurface_Inputs!$C$15)^2)"
                "+Subsurface_Inputs!$C$16*EXP(-0.5*((A{r}-Subsurface_Inputs!$C$17)/Subsurface_Inputs!$C$18)^2)))"
            ).format(r=r),
            (
                "=Subsurface_Inputs!$C$19"
                "+Subsurface_Inputs!$C$20*SIN(2*PI()*(A{r}-Subsurface_Inputs!$C$21)/Subsurface_Inputs!$C$22)"
                "-Subsurface_Inputs!$C$23*H{r}"
            ).format(r=r),
            f"=B{r}-I{r}",
            (
                "=Subsurface_Inputs!$C$25"
                "+Subsurface_Inputs!$C$26*SIN(2*PI()*(A{r}+Subsurface_Inputs!$C$27)/Subsurface_Inputs!$C$28)"
                "+Subsurface_Inputs!$C$29*COS(2*PI()*(A{r}-Subsurface_Inputs!$C$30)/Subsurface_Inputs!$C$31)"
                "-Subsurface_Inputs!$C$32*(B{r}-AVERAGE($B$2:$B$242))"
            ).format(r=r),
            f"=B{r}-K{r}",
            f"=2*Subsurface_Inputs!$C$34*F{r}/Subsurface_Inputs!$C$35*1000000",
            f"=2*Subsurface_Inputs!$C$34*I{r}/Subsurface_Inputs!$C$35*1000000",
            f"=2*Subsurface_Inputs!$C$34*K{r}/Subsurface_Inputs!$C$35*1000000",
            f"=Subsurface_Inputs!$C$37-2*Subsurface_Inputs!$C$36*(F{r}/1000)",
            f"=Subsurface_Inputs!$C$38+Subsurface_Inputs!$C$39*H{r}-2*Subsurface_Inputs!$C$36*(I{r}/1000)",
            f"=Subsurface_Inputs!$C$40-2*Subsurface_Inputs!$C$36*(K{r}/1000)-Subsurface_Inputs!$C$41*ABS(L{r}-AVERAGE($L$2:$L$242))",
            f"=K{r}-AVERAGE($K$2:$K$242)",
            f"=K{r}/F{r}",
            f'=IF(H{r}>=Subsurface_Inputs!$C$42,"Lens visible","Weak/no lens")',
            f'=IF(R{r}>-45,"Likely visible","Weak")',
        ]
        ws.append(formulas)

    add_borders(ws, 1, ws.max_row, 1, ws.max_column)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 18
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column <= 20:
                cell.number_format = "0.000"
    return ws


def make_checks(wb):
    ws = wb.create_sheet("Subsurface_Checks")
    style_title(ws, "A1", "Subsurface Formula Checks")
    rows = [
        ["Check", "Live result", "Why it matters"],
        ["Upper layer above lens", '=IF(MIN(Subsurface_Live_Data!I2:I242)>MAX(Subsurface_Live_Data!F2:F242),"OK","Review")', "Lens should be deeper than shallow internal ice."],
        ["Ocean deeper than lens", '=IF(MIN(Subsurface_Live_Data!K2:K242)>MAX(Subsurface_Live_Data!I2:I242),"OK","Review")', "Possible ocean boundary should be deepest reflector."],
        ["Positive depths", '=IF(MIN(Subsurface_Live_Data!F2:K242)>0,"OK","Review")', "Depths should not go negative."],
        ["Lens strength in range", '=IF(AND(MIN(Subsurface_Live_Data!H2:H242)>=0,MAX(Subsurface_Live_Data!H2:H242)<=1),"OK","Review")', "Lens strength is intended as 0 to 1."],
        ["Echo values finite", '=IF(COUNT(Subsurface_Live_Data!P2:R242)=ROWS(Subsurface_Live_Data!P2:R242)*3,"OK","Review")', "Echo estimate columns should remain numeric."],
        ["Workbook warning", "Synthetic data for model testing; not measured Europa data.", "Keeps the interpretation honest."],
    ]
    for row in rows:
        ws.append(row)
    style_header_row(ws, 2, 1, 3)
    add_borders(ws, 2, ws.max_row, 1, 3)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 80
    return ws


def make_formula_sheet(wb):
    ws = wb.create_sheet("Subsurface_Formulas")
    style_title(
        ws,
        "A1",
        "Subsurface Simulation Formulas",
        "These formulas drive the live subsurface data and dashboard charts.",
    )
    rows = [
        ["Formula", "Meaning", "Inputs used", "Output"],
        ["surface_height = Model_Data!h_target", "Uses the existing generated icy topography as the surface.", "Model_Data, topography controls", "surface height"],
        ["upper_depth = mean + A1*sin(2*pi*(x+phase)/L1) + A2*cos(2*pi*x/L2) + coupling*(surface-avg_surface)", "Synthetic shallow internal ice layer.", "Subsurface_Inputs shallow ice controls", "upper_ice_layer_depth_m"],
        ["lens_strength = MIN(1, MAX(0, lensA*gaussianA + lensB*gaussianB))", "Where a possible warm/briny lens is strongest.", "lens strengths, centers, widths", "lens_strength_0to1"],
        ["lens_depth = mean + wave_amp*sin(2*pi*(x-phase)/wavelength) - uplift*lens_strength", "Possible warm/briny ice lens depth.", "lens controls", "briny_warm_lens_depth_m"],
        ["ocean_depth = nominal + A*sin(2*pi*(x+phase)/L) + B*cos(2*pi*(x-phase)/L2) - coupling*(surface-avg_surface)", "Possible ice-ocean boundary depth.", "ice-ocean controls", "ice_ocean_boundary_depth_m"],
        ["elevation = surface_height - depth", "Converts depth below surface to elevation profile.", "surface and layer depth", "layer elevation"],
        ["delay_us = 2*n*depth/c*1e6", "Two-way radar delay after the surface return.", "ice refractive index, speed of light", "layer delays"],
        ["upper_echo = base - 2*attenuation*(depth_km)", "Simple shallow layer echo estimate.", "base echo, attenuation, depth", "upper echo dB"],
        ["lens_echo = base + bonus*lens_strength - 2*attenuation*(depth_km)", "Lens echo gets stronger where lens contrast is stronger.", "lens strength and attenuation", "lens echo dB"],
        ["ocean_echo = base - 2*attenuation*(depth_km) - roughness_penalty*ABS(boundary - avg_boundary)", "Deep reflector echo with attenuation and roughness penalty.", "ocean inputs and boundary elevation", "ocean echo dB"],
    ]
    for row in rows:
        ws.append(row)
    style_header_row(ws, 3, 1, 4)
    add_borders(ws, 3, ws.max_row, 1, 4)
    for col, width in {"A": 70, "B": 54, "C": 42, "D": 30}.items():
        ws.column_dimensions[col].width = width
    return ws


def add_series(chart, ws, x_col, y_col, title, color):
    xvalues = Reference(ws, min_col=x_col, min_row=2, max_row=242)
    yvalues = Reference(ws, min_col=y_col, min_row=2, max_row=242)
    series = Series(yvalues, xvalues, title=title)
    series.graphicalProperties.line.solidFill = color
    series.graphicalProperties.line.width = 22000
    chart.series.append(series)


def make_scatter_chart(ws_data, title, y_title, series_specs):
    chart = ScatterChart()
    chart.title = title
    chart.scatterStyle = "line"
    chart.height = 8.2
    chart.width = 16
    chart.x_axis.title = "Along-track position x (km)"
    chart.y_axis.title = y_title
    chart.legend.position = "b"
    for y_col, label, color in series_specs:
        add_series(chart, ws_data, 1, y_col, label, color)
    return chart


def make_dashboard(wb, ws_data):
    ws = wb.create_sheet("Subsurface_Dashboard", 1)
    style_title(
        ws,
        "A1",
        "Europa Ice Subsurface Live Dashboard",
        "Live formulas from Subsurface_Inputs and Model_Data. Existing dashboard charts remain unchanged.",
    )
    ws["A4"] = "Key Live Outputs"
    ws["A4"].font = Font(bold=True, color=COLORS["dark"], size=13)
    kpi_rows = [
        ["Metric", "Value", "Unit", "Meaning"],
        ["Average ice-ocean boundary depth", "=AVERAGE(Subsurface_Live_Data!K2:K242)", "m", "Mean possible bottom reflector depth."],
        ["Min ice-ocean boundary depth", "=MIN(Subsurface_Live_Data!K2:K242)", "m", "Shallowest possible bottom reflector."],
        ["Max ice-ocean boundary depth", "=MAX(Subsurface_Live_Data!K2:K242)", "m", "Deepest possible bottom reflector."],
        ["Average ocean delay", "=AVERAGE(Subsurface_Live_Data!O2:O242)", "us", "Expected timing after surface return."],
        ["Strongest lens strength", "=MAX(Subsurface_Live_Data!H2:H242)", "0-1", "Peak possible warm/briny lens contrast."],
        ["Average ocean echo", "=AVERAGE(Subsurface_Live_Data!R2:R242)", "dB", "Mean estimated deep reflector strength."],
        ["Strongest ocean echo", "=MAX(Subsurface_Live_Data!R2:R242)", "dB", "Best-case deep reflector return."],
        ["Visible lens samples", '=COUNTIF(Subsurface_Live_Data!U2:U242,"Lens visible")', "count", "How much of the pass has a strong lens flag."],
    ]
    for r_idx, row in enumerate(kpi_rows, start=5):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(r_idx, c_idx).value = value
    style_header_row(ws, 5, 1, 4)
    add_borders(ws, 5, 13, 1, 4)
    for row in range(6, 14):
        ws.cell(row, 2).number_format = "0.000"

    summary_rows = [
        ["What this tab is for", "Use this to test how a mostly-ice subsurface could change radar returns."],
        ["What is editable", "Change yellow cells in Subsurface_Inputs."],
        ["Important warning", "This is synthetic model data, not real measured Europa radar data."],
    ]
    for r_idx, row in enumerate(summary_rows, start=5):
        ws.cell(r_idx, 6).value = row[0]
        ws.cell(r_idx, 7).value = row[1]
    add_borders(ws, 5, 7, 6, 7)
    ws.cell(5, 6).fill = PatternFill("solid", fgColor=COLORS["light_blue"])
    ws.cell(5, 6).font = Font(bold=True)

    chart1 = make_scatter_chart(
        ws_data,
        "Subsurface Truth Model: Icy Layers",
        "Elevation relative to reference (m)",
        [
            (2, "Icy top surface", COLORS["orange"]),
            (7, "Shallow internal ice layer", COLORS["green"]),
            (10, "Possible warm/briny lens", COLORS["gold"]),
            (12, "Possible ice-ocean boundary", COLORS["purple"]),
        ],
    )
    ws.add_chart(chart1, "A16")

    chart2 = make_scatter_chart(
        ws_data,
        "Subsurface Radar Delay After Surface Return",
        "Two-way delay (microseconds)",
        [
            (13, "Shallow internal ice layer", COLORS["green"]),
            (14, "Possible warm/briny lens", COLORS["gold"]),
            (15, "Possible ice-ocean boundary", COLORS["purple"]),
        ],
    )
    ws.add_chart(chart2, "J16")

    chart3 = make_scatter_chart(
        ws_data,
        "Estimated Subsurface Echo Strength",
        "Relative echo strength (dB)",
        [
            (16, "Shallow internal ice layer", COLORS["green"]),
            (17, "Possible warm/briny lens", COLORS["gold"]),
            (18, "Possible ice-ocean boundary", COLORS["purple"]),
        ],
    )
    ws.add_chart(chart3, "A33")

    chart4 = make_scatter_chart(
        ws_data,
        "Possible Warm/Briny Lens Strength",
        "Lens strength (0-1)",
        [(8, "Lens strength", COLORS["gold"])],
    )
    chart4.legend = None
    ws.add_chart(chart4, "J33")

    for col, width in {"A": 32, "B": 18, "C": 12, "D": 54, "F": 26, "G": 78}.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A5"
    return ws


def append_formula_guide(wb):
    ws = wb["Formula_Guide"]
    start = ws.max_row + 2
    rows = [
        ["Subsurface simulation formulas", None, None, None],
        ["upper_depth = mean + sinusoid + cosine + surface coupling", "Live shallow internal ice layer depth.", "Subsurface_Inputs, Model_Data h_target", "upper_ice_layer_depth_m"],
        ["lens_strength = MIN(1,MAX(0,lensA Gaussian + lensB Gaussian))", "Possible warm/briny lens strength along track.", "lens strengths, centers, widths", "lens_strength_0to1"],
        ["lens_depth = mean + sinusoid - uplift*lens_strength", "Possible lens depth below icy surface.", "lens controls", "briny_warm_lens_depth_m"],
        ["ocean_depth = nominal + basal waves - surface coupling", "Possible ice-ocean boundary depth.", "ocean boundary controls", "ice_ocean_boundary_depth_m"],
        ["delay_us = 2*n*depth/c*1e6", "Two-way delay after the surface return.", "ice refractive index, speed of light", "subsurface layer delay"],
        ["echo_dB = base contrast - two-way attenuation - roughness penalty", "Simple radar echo strength estimate.", "attenuation and layer depth/roughness", "relative echo strength"],
    ]
    for r_idx, row in enumerate(rows, start=start):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(r_idx, c_idx).value = value
    ws.cell(start, 1).font = Font(bold=True, color=COLORS["dark"], size=13)
    style_header_row(ws, start + 1, 1, 4)
    add_borders(ws, start + 1, start + len(rows) - 1, 1, 4)


def append_graph_guide(wb):
    ws = wb["Graph_Guide"]
    start = ws.max_row + 2
    rows = [
        ["Subsurface Dashboard Graphs", None, None, None, None, None, None],
        ["Graph", "What it shows", "Why it matters", "How to read the baseline", "Source range", "Formula basis", "Audit status"],
        ["Subsurface Truth Model: Icy Layers", "Icy topography plus shallow ice, lens, and possible ocean boundary elevations.", "Shows the hidden simulated structure before radar distortion.", "Surface/topography is icy terrain, not rock.", "Subsurface_Live_Data!A:L", "Depth formulas from Subsurface_Inputs.", "See Subsurface_Checks"],
        ["Subsurface Radar Delay", "Two-way delay after the surface return for each modeled layer.", "Shows when each reflector would appear in a radargram.", "0 us means the surface return; larger delay means deeper ice.", "Subsurface_Live_Data!A:M:O", "delay_us = 2*n*depth/c*1e6", "See Subsurface_Checks"],
        ["Estimated Subsurface Echo Strength", "Relative echo strength for shallow ice, lens, and possible ocean boundary.", "Shows which returns may be easiest or hardest to detect.", "Higher / less negative dB is stronger.", "Subsurface_Live_Data!A:P:R", "base contrast - attenuation - roughness penalty", "See Subsurface_Checks"],
        ["Possible Warm/Briny Lens Strength", "Strength of possible local internal lens contrast.", "Highlights where internal ice may produce a distinct reflector.", "0 means absent/weak; 1 means strongest modeled contrast.", "Subsurface_Live_Data!A:H", "Gaussian lens strength fields.", "See Subsurface_Checks"],
    ]
    for r_idx, row in enumerate(rows, start=start):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(r_idx, c_idx).value = value
    ws.cell(start, 1).font = Font(bold=True, color=COLORS["dark"], size=13)
    style_header_row(ws, start + 1, 1, 7)
    add_borders(ws, start + 1, start + len(rows) - 1, 1, 7)


def main():
    wb = load_workbook(BASE_XLSX)
    set_calc_mode(wb)
    ensure_fixed_existing_dashboard(wb)

    for sheet in SUBSURFACE_SHEETS:
        delete_if_exists(wb, sheet)

    make_inputs(wb)
    live_data = make_live_data(wb)
    make_checks(wb)
    make_formula_sheet(wb)
    make_dashboard(wb, live_data)
    append_formula_guide(wb)
    append_graph_guide(wb)

    wb.save(OUT_XLSX)
    print(OUT_XLSX)


if __name__ == "__main__":
    main()
