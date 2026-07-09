from pathlib import Path
import math

import numpy as np
import openpyxl
from PIL import Image, ImageDraw, ImageFont


INPUT_XLSX = Path(
    r"C:\Users\tyboy\Downloads\parabolic-motion-radar-model-baseline-and-runs-dashboard-native-excel-charts-fixed (1).xlsx"
)
OUTPUT_DIR = Path(r"C:\Users\tyboy\OneDrive\Documents\Nasa project\outputs\python_graph_previews")


COLORS = {
    "ink": (28, 37, 54),
    "muted": (92, 103, 120),
    "grid": (218, 224, 232),
    "axis": (80, 91, 110),
    "blue": (32, 99, 181),
    "orange": (230, 111, 36),
    "green": (35, 150, 82),
    "purple": (118, 77, 176),
    "red": (196, 58, 58),
    "gold": (214, 155, 36),
    "safe": (41, 143, 88),
    "unsafe": (204, 72, 72),
    "band": (245, 168, 81, 72),
    "background": (255, 255, 255),
}


def load_font(size, bold=False):
    candidates = [
        r"C:\Windows\Fonts\arialbd.ttf" if bold else r"C:\Windows\Fonts\arial.ttf",
        r"C:\Windows\Fonts\calibrib.ttf" if bold else r"C:\Windows\Fonts\calibri.ttf",
    ]
    for candidate in candidates:
        try:
            return ImageFont.truetype(candidate, size)
        except Exception:
            pass
    return ImageFont.load_default()


FONT_TITLE = load_font(34, bold=True)
FONT_SUBTITLE = load_font(18)
FONT_LABEL = load_font(18, bold=True)
FONT_TICK = load_font(15)
FONT_NOTE = load_font(15)


def sheet_to_columns(ws):
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    columns = {h: [] for h in headers if h is not None}
    for row in rows[1:]:
        for idx, header in enumerate(headers):
            if header is None:
                continue
            columns[header].append(row[idx] if idx < len(row) else None)
    return columns


def numeric_array(values):
    out = []
    for value in values:
        try:
            if value is None:
                out.append(np.nan)
            else:
                out.append(float(value))
        except Exception:
            out.append(np.nan)
    return np.asarray(out, dtype=float)


def draw_text(draw, xy, text, font, fill=COLORS["ink"]):
    draw.text(xy, str(text), font=font, fill=fill)


def fmt_value(value):
    if not np.isfinite(value):
        return ""
    a = abs(value)
    if a >= 1000:
        return f"{value:,.0f}"
    if a >= 100:
        return f"{value:.0f}"
    if a >= 10:
        return f"{value:.1f}"
    return f"{value:.2f}"


def nice_range(values, pad=0.08):
    vals = np.asarray(values, dtype=float)
    vals = vals[np.isfinite(vals)]
    if vals.size == 0:
        return 0.0, 1.0
    lo = float(vals.min())
    hi = float(vals.max())
    if math.isclose(lo, hi):
        spread = abs(lo) * 0.1 if lo else 1.0
        return lo - spread, hi + spread
    spread = hi - lo
    return lo - spread * pad, hi + spread * pad


def line_points(x, y, plot_box, x_range, y_range):
    left, top, right, bottom = plot_box
    x0, x1 = x_range
    y0, y1 = y_range
    pts = []
    for xv, yv in zip(x, y):
        if not (np.isfinite(xv) and np.isfinite(yv)):
            continue
        px = left + (xv - x0) / (x1 - x0) * (right - left)
        py = bottom - (yv - y0) / (y1 - y0) * (bottom - top)
        pts.append((px, py))
    return pts


def draw_legend(draw, items, x, y, max_width):
    cursor_x = x
    cursor_y = y
    row_h = 26
    for label, color in items:
        text_w = draw.textlength(label, font=FONT_NOTE)
        item_w = 34 + text_w + 22
        if cursor_x + item_w > x + max_width:
            cursor_x = x
            cursor_y += row_h
        draw.line((cursor_x, cursor_y + 12, cursor_x + 24, cursor_y + 12), fill=color, width=4)
        draw_text(draw, (cursor_x + 32, cursor_y + 3), label, FONT_NOTE, COLORS["muted"])
        cursor_x += item_w
    return cursor_y + row_h


def draw_axes(draw, plot_box, x_range, y_range, x_label, y_label, y_ticks=6):
    left, top, right, bottom = plot_box
    draw.rectangle(plot_box, outline=COLORS["axis"], width=2)
    for i in range(y_ticks):
        t = i / (y_ticks - 1)
        y_val = y_range[0] + t * (y_range[1] - y_range[0])
        py = bottom - t * (bottom - top)
        draw.line((left, py, right, py), fill=COLORS["grid"], width=1)
        label = fmt_value(y_val)
        tw = draw.textlength(label, font=FONT_TICK)
        draw_text(draw, (left - tw - 10, py - 8), label, FONT_TICK, COLORS["muted"])
    for i in range(7):
        t = i / 6
        x_val = x_range[0] + t * (x_range[1] - x_range[0])
        px = left + t * (right - left)
        draw.line((px, top, px, bottom), fill=(236, 240, 245), width=1)
        label = fmt_value(x_val)
        tw = draw.textlength(label, font=FONT_TICK)
        draw_text(draw, (px - tw / 2, bottom + 10), label, FONT_TICK, COLORS["muted"])
    draw_text(draw, ((left + right) / 2 - draw.textlength(x_label, font=FONT_LABEL) / 2, bottom + 40), x_label, FONT_LABEL)
    draw_text(draw, (left, top - 30), y_label, FONT_LABEL)


def draw_line_chart(
    image,
    box,
    title,
    subtitle,
    series,
    x_label,
    y_label,
    y_range=None,
    band=None,
    hlines=None,
):
    draw = ImageDraw.Draw(image)
    left, top, right, bottom = box
    draw_text(draw, (left, top), title, FONT_TITLE)
    if subtitle:
        draw_text(draw, (left, top + 42), subtitle, FONT_SUBTITLE, COLORS["muted"])

    legend_items = [(item["label"], item["color"]) for item in series]
    legend_bottom = draw_legend(draw, legend_items, left, top + 76, right - left)

    plot_box = (left + 90, legend_bottom + 18, right - 28, bottom - 72)
    x_values = np.concatenate([item["x"][np.isfinite(item["x"])] for item in series])
    y_values = np.concatenate([item["y"][np.isfinite(item["y"])] for item in series])
    if band is not None:
        y_values = np.concatenate([y_values, band["low"][np.isfinite(band["low"])], band["high"][np.isfinite(band["high"])]])
    if hlines:
        y_values = np.concatenate([y_values, np.asarray([line["y"] for line in hlines], dtype=float)])
    x_range = nice_range(x_values, pad=0.02)
    if y_range is None:
        y_range = nice_range(y_values)

    draw_axes(draw, plot_box, x_range, y_range, x_label, y_label)

    if band is not None:
        upper = line_points(band["x"], band["high"], plot_box, x_range, y_range)
        lower = line_points(band["x"], band["low"], plot_box, x_range, y_range)
        if len(upper) > 1 and len(lower) > 1:
            overlay = Image.new("RGBA", image.size, (255, 255, 255, 0))
            odraw = ImageDraw.Draw(overlay)
            odraw.polygon(upper + lower[::-1], fill=COLORS["band"])
            image.alpha_composite(overlay)
            draw = ImageDraw.Draw(image)

    if hlines:
        for line in hlines:
            pts = line_points(np.asarray([x_range[0], x_range[1]]), np.asarray([line["y"], line["y"]]), plot_box, x_range, y_range)
            if len(pts) == 2:
                draw.line(pts, fill=line.get("color", COLORS["axis"]), width=line.get("width", 2))
                draw_text(draw, (plot_box[2] - 220, pts[0][1] - 22), line.get("label", ""), FONT_NOTE, line.get("color", COLORS["axis"]))

    for item in series:
        pts = line_points(item["x"], item["y"], plot_box, x_range, y_range)
        if len(pts) > 1:
            draw.line(pts, fill=item["color"], width=item.get("width", 4), joint="curve")


def color_interp(a, b, t):
    return tuple(int(a[i] + (b[i] - a[i]) * t) for i in range(3))


def terrain_color(value, lo, hi):
    if not np.isfinite(value):
        return (230, 230, 230)
    t = (value - lo) / (hi - lo) if hi != lo else 0.5
    t = max(0.0, min(1.0, t))
    stops = [
        (0.00, (59, 93, 160)),
        (0.24, (128, 176, 205)),
        (0.43, (232, 238, 230)),
        (0.58, (137, 174, 97)),
        (0.78, (219, 185, 101)),
        (1.00, (137, 92, 64)),
    ]
    for idx in range(1, len(stops)):
        if t <= stops[idx][0]:
            t0, c0 = stops[idx - 1]
            t1, c1 = stops[idx]
            return color_interp(c0, c1, (t - t0) / (t1 - t0))
    return stops[-1][1]


def make_topography_function(inputs):
    def cell_bool(value):
        return value is True or str(value).upper() == "TRUE" or value == 1

    factor = 1.0 if cell_bool(inputs["topography_on"]) else 0.0
    pi = np.pi

    def h_total(x, y):
        x = np.asarray(x, dtype=float)
        y = np.asarray(y, dtype=float)
        h_project = inputs["S_dem"] * (
            115 * np.exp(-(((y - 25) ** 2) / (4.5**2)))
            + 70 * np.exp(-(((y + 38) ** 2) / (7**2))) * (0.55 + 0.45 * np.sin(2 * pi * x / 45))
            + 24 * np.sin(2 * pi * y / 9 + x / 18) * np.exp(-(((y + 5) ** 2) / (22**2)))
            + 18 * np.sin(2 * pi * x / 70)
        )
        h_ridge = (
            inputs["A_ridge"]
            * np.exp(-(((x - inputs["ridge_x0"]) ** 2) / (2 * inputs["ridge_sigma_x"] ** 2)))
            * np.exp(-(((y - inputs["ridge_y0"]) ** 2) / (2 * inputs["ridge_sigma_y"] ** 2)))
        )
        h_crater = -inputs["D_crater"] * np.exp(
            -((((x - inputs["crater_x0"]) ** 2) + ((y - inputs["crater_y0"]) ** 2)) / (2 * inputs["crater_sigma"] ** 2))
        )
        h_chaos = inputs["A_chaos"] * (
            0.55 * np.sin(2 * pi * x / inputs["Lx"])
            + 0.35 * np.cos(2 * pi * y / inputs["Ly"])
            + 0.2 * np.sin(2 * pi * (x + y) / (0.5 * (inputs["Lx"] + inputs["Ly"])))
        )
        h_rough = inputs["A_rough"] * np.sin(2 * pi * x / inputs["Lx"]) * np.sin(2 * pi * y / inputs["Ly"])
        h_trough = (
            -inputs["D_trough"]
            * np.exp(-(((y - inputs["trough_y0"]) ** 2) / (2 * inputs["trough_sigma_y"] ** 2)))
            * np.exp(-(((x - inputs["trough_x0"]) ** 2) / (2 * inputs["trough_sigma_x"] ** 2)))
        )
        smooth_seeded = (
            0.45 * np.sin(2 * pi * ((x / inputs["L_seed_x"]) + inputs["terrain_seed"] * 0.137))
            + 0.35 * np.cos(2 * pi * ((y / inputs["L_seed_y"]) + inputs["terrain_seed"] * 0.173))
            + 0.2
            * np.sin(
                2
                * pi
                * ((x + y) / (0.5 * (inputs["L_seed_x"] + inputs["L_seed_y"])) + inputs["terrain_seed"] * 0.097)
            )
        )
        raw_hash = np.sin(x * 12.9898 + y * 78.233 + inputs["terrain_seed"] * 37.719) * 43758.5453
        deterministic_hash = 2 * (raw_hash - np.floor(raw_hash)) - 1
        h_seeded = inputs["A_seeded"] * (0.7 * smooth_seeded + 0.3 * deterministic_hash)
        return factor * (h_project + h_ridge + h_crater + h_chaos + h_rough + h_trough + h_seeded)

    return h_total


def draw_topography_map(path, model, inputs):
    x = model["x_km"]
    x_min, x_max = float(np.nanmin(x)), float(np.nanmax(x))
    target_y = float(inputs["target_y_km"])
    y_min = min(-70.0, target_y - 20.0)
    y_max = max(70.0, target_y + 20.0)

    xs = np.linspace(x_min, x_max, 300)
    ys = np.linspace(y_min, y_max, 220)
    xx, yy = np.meshgrid(xs, ys)
    h = make_topography_function(inputs)(xx, yy)

    lo, hi = float(np.nanpercentile(h, 2)), float(np.nanpercentile(h, 98))
    rgb = np.zeros((h.shape[0], h.shape[1], 3), dtype=np.uint8)
    for row in range(h.shape[0]):
        for col in range(h.shape[1]):
            rgb[row, col, :] = terrain_color(h[row, col], lo, hi)

    image = Image.new("RGBA", (1400, 880), COLORS["background"] + (255,))
    draw = ImageDraw.Draw(image)
    draw_text(draw, (54, 36), "Generated Topography Map", FONT_TITLE)
    draw_text(
        draw,
        (54, 80),
        "Synthetic surface used by the model; green is nadir, orange is the off-nadir target path.",
        FONT_SUBTITLE,
        COLORS["muted"],
    )

    plot_box = (100, 135, 1230, 760)
    heat = Image.fromarray(rgb, mode="RGB").resize((plot_box[2] - plot_box[0], plot_box[3] - plot_box[1]), Image.Resampling.BILINEAR)
    image.paste(heat.convert("RGBA"), plot_box[:2])
    draw = ImageDraw.Draw(image)
    draw.rectangle(plot_box, outline=COLORS["axis"], width=2)

    def map_x(xv):
        return plot_box[0] + (xv - x_min) / (x_max - x_min) * (plot_box[2] - plot_box[0])

    def map_y(yv):
        return plot_box[3] - (yv - y_min) / (y_max - y_min) * (plot_box[3] - plot_box[1])

    for xv in np.linspace(x_min, x_max, 7):
        px = map_x(xv)
        draw.line((px, plot_box[1], px, plot_box[3]), fill=(255, 255, 255, 115), width=1)
        label = fmt_value(xv)
        draw_text(draw, (px - draw.textlength(label, font=FONT_TICK) / 2, plot_box[3] + 10), label, FONT_TICK, COLORS["muted"])
    for yv in np.linspace(y_min, y_max, 7):
        py = map_y(yv)
        draw.line((plot_box[0], py, plot_box[2], py), fill=(255, 255, 255, 115), width=1)
        label = fmt_value(yv)
        draw_text(draw, (plot_box[0] - draw.textlength(label, font=FONT_TICK) - 10, py - 8), label, FONT_TICK, COLORS["muted"])

    nadir_y = map_y(0.0)
    target_line_y = map_y(target_y)
    draw.line((plot_box[0], nadir_y, plot_box[2], nadir_y), fill=COLORS["green"], width=5)
    draw.line((plot_box[0], target_line_y, plot_box[2], target_line_y), fill=COLORS["orange"], width=5)
    draw_text(draw, (plot_box[0] + 12, nadir_y + 8), "Nadir path y = 0 km", FONT_NOTE, COLORS["green"])
    draw_text(draw, (plot_box[0] + 12, target_line_y + 8), f"Off-nadir target y = {target_y:g} km", FONT_NOTE, COLORS["orange"])
    draw_text(draw, ((plot_box[0] + plot_box[2]) / 2 - 105, plot_box[3] + 44), "Along-track position x (km)", FONT_LABEL)
    draw_text(draw, (plot_box[0], plot_box[1] - 30), "Cross-track position y (km)", FONT_LABEL)

    colorbar = (1260, 135, 1300, 760)
    for i in range(colorbar[3] - colorbar[1]):
        t = 1 - i / (colorbar[3] - colorbar[1] - 1)
        color = terrain_color(lo + t * (hi - lo), lo, hi)
        draw.line((colorbar[0], colorbar[1] + i, colorbar[2], colorbar[1] + i), fill=color, width=1)
    draw.rectangle(colorbar, outline=COLORS["axis"], width=1)
    draw_text(draw, (colorbar[2] + 10, colorbar[1] - 4), f"{hi:.0f} m", FONT_TICK, COLORS["muted"])
    draw_text(draw, (colorbar[2] + 10, colorbar[3] - 16), f"{lo:.0f} m", FONT_TICK, COLORS["muted"])
    draw_text(draw, (colorbar[0] - 8, colorbar[3] + 16), "height", FONT_TICK, COLORS["muted"])

    image.convert("RGB").save(path, quality=95)


def draw_prf_chart(path, prf_rows, required_prf):
    image = Image.new("RGBA", (1200, 760), COLORS["background"] + (255,))
    draw = ImageDraw.Draw(image)
    draw_text(draw, (54, 36), "PRF Safety Check", FONT_TITLE)
    draw_text(draw, (54, 80), "Compares candidate pulse repetition frequencies against the simple no-alias floor.", FONT_SUBTITLE, COLORS["muted"])

    prf_values = np.asarray([float(row["PRF (Hz)"]) for row in prf_rows], dtype=float)
    statuses = [str(row["Status"]) for row in prf_rows]
    labels = [f"{int(v):,} Hz" for v in prf_values]

    plot_box = (110, 140, 1120, 565)
    values_log = np.log10(prf_values)
    floor_log = math.log10(required_prf)
    y_min = math.log10(30)
    y_max = math.log10(max(prf_values.max(), required_prf) * 1.8)
    draw.rectangle(plot_box, outline=COLORS["axis"], width=2)
    tick_values = [30, 50, 100, 500, 1000, 3000, 5000]
    for tick in tick_values:
        if tick < 30 or tick > 6000:
            continue
        t = (math.log10(tick) - y_min) / (y_max - y_min)
        py = plot_box[3] - t * (plot_box[3] - plot_box[1])
        draw.line((plot_box[0], py, plot_box[2], py), fill=COLORS["grid"], width=1)
        label = f"{tick:,}"
        draw_text(draw, (plot_box[0] - draw.textlength(label, font=FONT_TICK) - 10, py - 8), label, FONT_TICK, COLORS["muted"])

    floor_y = plot_box[3] - (floor_log - y_min) / (y_max - y_min) * (plot_box[3] - plot_box[1])
    draw.line((plot_box[0], floor_y, plot_box[2], floor_y), fill=COLORS["purple"], width=4)
    floor_label = f"Required floor: {required_prf:,.0f} Hz"
    floor_label_box = (
        plot_box[0] + 14,
        floor_y - 34,
        plot_box[0] + 34 + draw.textlength(floor_label, font=FONT_NOTE),
        floor_y - 8,
    )
    draw.rounded_rectangle(floor_label_box, radius=5, fill=(255, 255, 255), outline=(230, 230, 235))
    draw_text(draw, (floor_label_box[0] + 10, floor_label_box[1] + 5), floor_label, FONT_NOTE, COLORS["purple"])

    slot = (plot_box[2] - plot_box[0]) / len(prf_values)
    bar_w = slot * 0.48
    for idx, (value, value_log, label, status) in enumerate(zip(prf_values, values_log, labels, statuses)):
        cx = plot_box[0] + slot * idx + slot / 2
        bar_top = plot_box[3] - (value_log - y_min) / (y_max - y_min) * (plot_box[3] - plot_box[1])
        color = COLORS["safe"] if value >= required_prf else COLORS["unsafe"]
        draw.rounded_rectangle((cx - bar_w / 2, bar_top, cx + bar_w / 2, plot_box[3]), radius=6, fill=color)
        draw_text(draw, (cx - draw.textlength(label, font=FONT_LABEL) / 2, plot_box[3] + 16), label, FONT_LABEL)
        status_color = COLORS["safe"] if value >= required_prf else COLORS["unsafe"]
        status_short = "OK" if value >= required_prf else "Below floor"
        draw_text(draw, (cx - draw.textlength(status_short, font=FONT_NOTE) / 2, bar_top - 26), status_short, FONT_NOTE, status_color)
        draw_text(draw, (cx - draw.textlength(status, font=FONT_TICK) / 2, plot_box[3] + 44), status, FONT_TICK, COLORS["muted"])

    draw_text(draw, (plot_box[0], plot_box[1] - 30), "PRF Hz (log scale)", FONT_LABEL)
    draw_text(draw, ((plot_box[0] + plot_box[2]) / 2 - 100, plot_box[3] + 112), "Candidate PRF setting", FONT_LABEL)
    draw_text(draw, (54, 720), "Log scale is used so the low PRF options remain visible beside the 3,000 Hz setting.", FONT_NOTE, COLORS["muted"])
    image.convert("RGB").save(path, quality=95)


def draw_preview_graphs():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(INPUT_XLSX, data_only=True, read_only=True)
    model_cols = sheet_to_columns(wb["Model_Data"])
    prf_ws = wb["PRF_Results"]
    input_ws = wb["Inputs"]
    dashboard_ws = wb["Dashboard"]

    model = {name: numeric_array(values) for name, values in model_cols.items()}
    inputs = {}
    for row in range(27, 53):
        symbol = input_ws.cell(row, 2).value
        value = input_ws.cell(row, 3).value
        if symbol:
            try:
                inputs[symbol] = float(value)
            except Exception:
                inputs[symbol] = value
    inputs["target_y_km"] = float(input_ws["C6"].value)

    prf_rows = []
    prf_headers = [prf_ws.cell(3, col).value for col in range(1, 7)]
    for row_idx in range(4, 7):
        row = {header: prf_ws.cell(row_idx, col).value for col, header in enumerate(prf_headers, start=1)}
        if row.get("PRF (Hz)") is not None:
            row["Status"] = row["Doppler sampling status"]
            prf_rows.append(row)
    required_prf = float(dashboard_ws["B16"].value)

    x = model["x_km"]

    look_image = Image.new("RGBA", (1400, 1050), COLORS["background"] + (255,))
    draw_line_chart(
        look_image,
        (54, 30, 1348, 505),
        "Look Angle Along Curved Satellite Path",
        "Shows how the off-nadir viewing angle changes along the parabolic pass.",
        [
            {"label": "Flat geometry look angle", "x": x, "y": model["look_angle_deg"], "color": COLORS["blue"]},
            {"label": "Topography-adjusted look angle", "x": x, "y": model["look_angle_topo_deg"], "color": COLORS["orange"]},
        ],
        "Along-track position x (km)",
        "Look angle (deg)",
    )
    draw_line_chart(
        look_image,
        (54, 535, 1348, 1015),
        "Off-Nadir Slant Range",
        "Shows whether terrain changes the actual radar path length to the target.",
        [
            {"label": "Flat off-nadir range", "x": x, "y": model["R_off_km"], "color": COLORS["blue"]},
            {"label": "Topography-adjusted off-nadir range", "x": x, "y": model["R_off_topo_km"], "color": COLORS["orange"]},
        ],
        "Along-track position x (km)",
        "Slant range (km)",
    )
    look_path = OUTPUT_DIR / "01_look_angle_and_slant_range.png"
    look_image.convert("RGB").save(look_path, quality=95)

    band_image = Image.new("RGBA", (1400, 820), COLORS["background"] + (255,))
    draw_line_chart(
        band_image,
        (54, 36, 1348, 785),
        "Apparent Depth With Topography Uncertainty Band",
        "Orange band uses the workbook's terrain vertical uncertainty converted into apparent-depth uncertainty.",
        [
            {"label": "Flat apparent depth", "x": x, "y": model["apparent_depth_m"], "color": COLORS["blue"]},
            {"label": "Topography-adjusted apparent depth", "x": x, "y": model["apparent_depth_topo_m"], "color": COLORS["orange"]},
        ],
        "Along-track position x (km)",
        "Apparent depth (m)",
        band={"x": x, "low": model["topo_depth_low_m"], "high": model["topo_depth_high_m"]},
        hlines=[{"y": 0, "color": COLORS["axis"], "width": 2, "label": "Nadir reference"}],
    )
    band_path = OUTPUT_DIR / "02_depth_uncertainty_band.png"
    band_image.convert("RGB").save(band_path, quality=95)

    topo_path = OUTPUT_DIR / "03_generated_topography_map.png"
    draw_topography_map(topo_path, model, inputs)

    prf_path = OUTPUT_DIR / "04_prf_safety_check.png"
    draw_prf_chart(prf_path, prf_rows, required_prf)

    print("Generated preview graphs:")
    for path in [look_path, band_path, topo_path, prf_path]:
        print(path)


if __name__ == "__main__":
    draw_preview_graphs()
