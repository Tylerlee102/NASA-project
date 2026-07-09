from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


BASE_XLSX = Path(r"C:\Users\tyboy\OneDrive\Documents\Nasa project\outputs\europa_ice_subsurface_simulation\v11.xlsx")
OUT_XLSX = Path(r"C:\Users\tyboy\OneDrive\Documents\Nasa project\outputs\europa_ice_subsurface_simulation\v12.xlsx")

COLORS = {
    "navy": "1F4E79",
    "dark": "1F2937",
    "muted": "526174",
    "grid": "D4DAE3",
    "light_blue": "D9EAF7",
}


def style_header(ws, row, min_col, max_col):
    for col in range(min_col, max_col + 1):
        cell = ws.cell(row, col)
        cell.fill = PatternFill("solid", fgColor=COLORS["navy"])
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def add_borders(ws, min_row, max_row, min_col, max_col):
    thin = Side(style="thin", color=COLORS["grid"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def delete_sheet(wb, name):
    if name in wb.sheetnames:
        del wb[name]


def set_calc_mode(wb):
    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass


def write_chart_block(ws, start_row, title, specs, source_start=2, source_end=242):
    ws.cell(start_row, 1).value = title
    ws.cell(start_row, 1).font = Font(bold=True, size=12, color=COLORS["dark"])
    header_row = start_row + 1
    for idx, spec in enumerate(specs):
        col = 1 + idx * 2
        ws.cell(header_row, col).value = f"{spec['name']} x"
        ws.cell(header_row, col + 1).value = spec["name"]
    style_header(ws, header_row, 1, len(specs) * 2)

    output_row = start_row + 2
    for source_row in range(source_start, source_end + 1):
        for idx, spec in enumerate(specs):
            col = 1 + idx * 2
            ws.cell(output_row, col).value = f"={spec['sheet']}!{spec['x_col']}{source_row}"
            ws.cell(output_row, col + 1).value = f"={spec['sheet']}!{spec['y_col']}{source_row}"
        output_row += 1

    add_borders(ws, header_row, output_row - 1, 1, len(specs) * 2)
    return output_row - 1


def make_chart_data(wb):
    delete_sheet(wb, "Subsurface_Chart_Data")
    insert_at = wb.sheetnames.index("Subsurface_Radargram_Data") + 1
    ws = wb.create_sheet("Subsurface_Chart_Data", insert_at)
    ws.sheet_view.showGridLines = False

    blocks = [
        (
            1,
            "Subsurface Truth Model: Icy Layers",
            [
                {"name": "Icy top surface", "sheet": "Subsurface_Live_Data", "x_col": "A", "y_col": "B"},
                {"name": "Shallow ice layer", "sheet": "Subsurface_Live_Data", "x_col": "A", "y_col": "G"},
                {"name": "Warm/briny lens", "sheet": "Subsurface_Live_Data", "x_col": "A", "y_col": "J"},
                {"name": "Ice-ocean boundary", "sheet": "Subsurface_Live_Data", "x_col": "A", "y_col": "L"},
            ],
            2,
            242,
        ),
        (
            246,
            "Scenario Comparison: Thin / Medium / Thick Ice",
            [
                {"name": "Thin shell", "sheet": "Subsurface_Scenario_Data", "x_col": "A", "y_col": "E"},
                {"name": "Medium shell", "sheet": "Subsurface_Scenario_Data", "x_col": "A", "y_col": "F"},
                {"name": "Thick shell", "sheet": "Subsurface_Scenario_Data", "x_col": "A", "y_col": "G"},
            ],
            2,
            242,
        ),
        (
            491,
            "Boundary Uncertainty Band",
            [
                {"name": "Lower bound", "sheet": "Subsurface_Scenario_Data", "x_col": "A", "y_col": "C"},
                {"name": "Mean boundary", "sheet": "Subsurface_Scenario_Data", "x_col": "A", "y_col": "B"},
                {"name": "Upper bound", "sheet": "Subsurface_Scenario_Data", "x_col": "A", "y_col": "D"},
            ],
            2,
            242,
        ),
        (
            736,
            "Ocean Model vs No-Ocean Control",
            [
                {"name": "Ocean model margin", "sheet": "Subsurface_Scenario_Data", "x_col": "A", "y_col": "H"},
                {"name": "No-ocean control margin", "sheet": "Subsurface_Scenario_Data", "x_col": "A", "y_col": "I"},
                {"name": "Zero threshold", "sheet": "Subsurface_Scenario_Data", "x_col": "A", "y_col": "J"},
            ],
            2,
            242,
        ),
        (
            981,
            "Radargram-Style Return Timing With Clutter",
            [
                {"name": "Surface clutter upper", "sheet": "Subsurface_Radargram_Data", "x_col": "A", "y_col": "C"},
                {"name": "Shallow ice return", "sheet": "Subsurface_Radargram_Data", "x_col": "A", "y_col": "D"},
                {"name": "Warm/briny lens return", "sheet": "Subsurface_Radargram_Data", "x_col": "A", "y_col": "E"},
                {"name": "Ocean boundary return", "sheet": "Subsurface_Radargram_Data", "x_col": "A", "y_col": "F"},
            ],
            2,
            242,
        ),
        (
            1226,
            "Detectability Margin vs Threshold",
            [
                {"name": "Lens echo margin", "sheet": "Subsurface_Live_Data", "x_col": "A", "y_col": "X"},
                {"name": "Ocean echo margin", "sheet": "Subsurface_Live_Data", "x_col": "A", "y_col": "Y"},
                {"name": "Zero margin threshold", "sheet": "Subsurface_Live_Data", "x_col": "A", "y_col": "AA"},
            ],
            2,
            242,
        ),
        (
            1471,
            "Reflection Strength by Material / Interface",
            [
                {"name": "Material/interface strength", "sheet": "Subsurface_Materials_Evidence", "x_col": "G", "y_col": "H"},
            ],
            2,
            6,
        ),
        (
            1490,
            "Cross-Instrument Evidence Score",
            [
                {"name": "Evidence support score", "sheet": "Subsurface_Materials_Evidence", "x_col": "I", "y_col": "J"},
            ],
            10,
            13,
        ),
    ]

    for start_row, title, specs, source_start, source_end in blocks:
        write_chart_block(ws, start_row, title, specs, source_start, source_end)

    for col in range(1, 10):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 24
    ws.freeze_panes = "A3"
    return ws


def append_formula_chart_section(wb):
    ws = wb["Subsurface_Formulas"]
    start = ws.max_row + 2
    rows = [
        ["Live dashboard graph formulas", None, None, None],
        ["Graph", "Live chart-data range", "How it updates", "Formula/source basis"],
        [
            "Subsurface Truth Model: Icy Layers",
            "Subsurface_Chart_Data!A3:H243",
            "Chart-data cells point to Subsurface_Live_Data; changing Subsurface_Inputs recalculates the layer elevations and redraws the chart.",
            "surface B, shallow G, lens J, boundary L from Subsurface_Live_Data",
        ],
        [
            "Scenario Comparison: Thin / Medium / Thick Ice",
            "Subsurface_Chart_Data!A248:F488",
            "Scenario depth formulas update from the live boundary depth and input multipliers.",
            "thin E, medium F, thick G from Subsurface_Scenario_Data",
        ],
        [
            "Boundary Uncertainty Band",
            "Subsurface_Chart_Data!A493:F733",
            "Uncertainty bounds update from the live boundary depth and the boundary uncertainty input.",
            "low C, mean B, high D from Subsurface_Scenario_Data",
        ],
        [
            "Ocean Model vs No-Ocean Control",
            "Subsurface_Chart_Data!A738:F978",
            "Margins update from the detection threshold, ocean echo model, and no-ocean control input.",
            "ocean margin H, no-ocean margin I, zero J from Subsurface_Scenario_Data",
        ],
        [
            "Radargram-Style Return Timing With Clutter",
            "Subsurface_Chart_Data!A983:H1223",
            "Timing curves update from layer depths, refractive index, speed of light, and radargram jitter/clutter inputs.",
            "clutter C, shallow D, lens E, ocean F from Subsurface_Radargram_Data",
        ],
        [
            "Detectability Margin vs Threshold",
            "Subsurface_Chart_Data!A1228:F1468",
            "Echo margins update from live lens/ocean echo formulas minus detection threshold.",
            "lens X, ocean Y, zero AA from Subsurface_Live_Data",
        ],
        [
            "Reflection Strength by Material / Interface",
            "Subsurface_Chart_Data!A1473:B1477",
            "Material strength chart updates from the material/evidence table.",
            "material index G and strength H from Subsurface_Materials_Evidence",
        ],
        [
            "Cross-Instrument Evidence Score",
            "Subsurface_Chart_Data!A1492:B1495",
            "Evidence score chart updates from the editable support-score inputs.",
            "evidence index I and score J from Subsurface_Materials_Evidence",
        ],
    ]
    for r_idx, row in enumerate(rows, start=start):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(r_idx, c_idx).value = value
            ws.cell(r_idx, c_idx).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(start, 1).font = Font(bold=True, size=13, color=COLORS["dark"])
    style_header(ws, start + 1, 1, 4)
    add_borders(ws, start + 1, start + len(rows) - 1, 1, 4)
    for col, width in {"A": 40, "B": 34, "C": 70, "D": 58}.items():
        ws.column_dimensions[col].width = width


def main():
    wb = load_workbook(BASE_XLSX)
    set_calc_mode(wb)
    make_chart_data(wb)
    append_formula_chart_section(wb)
    wb.save(OUT_XLSX)
    print(OUT_XLSX)


if __name__ == "__main__":
    main()
