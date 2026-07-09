from __future__ import annotations

import html
import json
import math
import os
from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont


ROOT = Path(__file__).resolve().parent
PROJECT_ROOT = ROOT.parents[1]
WORKBOOK = PROJECT_ROOT / "outputs" / "europa_ice_subsurface_simulation" / "v19.xlsx"
CHART_DIR = ROOT / "charts"
CHART_DIR.mkdir(parents=True, exist_ok=True)

RESULTS_CSV = ROOT / "dirty_ice_results.csv"
SUMMARY_CSV = ROOT / "scenario_summary.csv"
SENSITIVITY_CSV = ROOT / "attenuation_sensitivity_grid.csv"
CROSSREF_CSV = ROOT / "paper_cross_reference.csv"
REPORT_HTML = ROOT / "dirty_ice_research_report.html"
RUN_METADATA = ROOT / "run_metadata.json"


FREQUENCIES = {
    "HF_9MHz": {"mhz": 9.0, "resolution_m": 300.0, "role": "deep sounding"},
    "VHF_60MHz": {"mhz": 60.0, "resolution_m": 30.0, "role": "shallow/high-resolution sounding"},
}

ICE_EPS = 3.15
OCEAN_EPS = 80.0
DETECTION_THRESHOLD_DB = -45.0
SYSTEM_OFFSET_DB = -10.0

FONT_FAMILY = ["Aptos", "Inter", "Segoe UI", "DejaVu Sans", "Arial", "sans-serif"]
MONO_FONT_FAMILY = ["Consolas", "DejaVu Sans Mono", "monospace"]
TOKENS = {
    "surface": "#FCFCFD",
    "panel": "#FFFFFF",
    "ink": "#1F2430",
    "muted": "#6F768A",
    "grid": "#E6E8F0",
    "axis": "#D7DBE7",
}
COLORS = {
    "blue": "#5477C4",
    "blue_light": "#CEDFFE",
    "orange": "#CC6F47",
    "orange_light": "#FFBDA1",
    "olive": "#71B436",
    "olive_light": "#BEEB96",
    "pink": "#BD569B",
    "gold": "#B8A037",
    "neutral": "#7A828F",
    "dark": "#2E4780",
}


@dataclass(frozen=True)
class Layer:
    top_m: float
    thickness_m: float
    eps: float
    attenuation_db_km_hf: float
    label: str

    @property
    def bottom_m(self) -> float:
        return self.top_m + self.thickness_m


def setup_style() -> None:
    return None


def hex_to_rgb(color: str) -> tuple[int, int, int]:
    color = color.lstrip("#")
    return tuple(int(color[i : i + 2], 16) for i in (0, 2, 4))


def blend_color(start: str, end: str, t: float) -> tuple[int, int, int]:
    a = np.array(hex_to_rgb(start), dtype=float)
    b = np.array(hex_to_rgb(end), dtype=float)
    rgb = a + (b - a) * max(0, min(1, t))
    return tuple(int(v) for v in rgb)


def get_font(size: int, bold: bool = False):
    candidates = [
        r"C:\Windows\Fonts\arialbd.ttf" if bold else r"C:\Windows\Fonts\arial.ttf",
        r"C:\Windows\Fonts\segoeuib.ttf" if bold else r"C:\Windows\Fonts\segoeui.ttf",
        r"C:\Windows\Fonts\calibrib.ttf" if bold else r"C:\Windows\Fonts\calibri.ttf",
    ]
    for path in candidates:
        try:
            return ImageFont.truetype(path, size=size)
        except OSError:
            continue
    return ImageFont.load_default()


def text_width(draw: ImageDraw.ImageDraw, text: str, font) -> int:
    bbox = draw.textbbox((0, 0), text, font=font)
    return bbox[2] - bbox[0]


def draw_text(draw, xy, text, fill=None, size=18, bold=False, anchor=None):
    font = get_font(size, bold)
    draw.text(xy, text, fill=fill or hex_to_rgb(TOKENS["ink"]), font=font, anchor=anchor)


def draw_chart_header(draw, title: str, subtitle: str, width: int):
    draw_text(draw, (70, 34), title, size=28, bold=True)
    draw_text(draw, (70, 74), subtitle, fill=hex_to_rgb(TOKENS["muted"]), size=17)
    draw.line((70, 112, width - 70, 112), fill=hex_to_rgb(TOKENS["grid"]), width=2)


def save_image(img: Image.Image, name: str) -> Path:
    path = CHART_DIR / name
    img.save(path)
    return path


def new_chart_canvas(width=1400, height=850) -> tuple[Image.Image, ImageDraw.ImageDraw]:
    img = Image.new("RGB", (width, height), hex_to_rgb(TOKENS["surface"]))
    draw = ImageDraw.Draw(img)
    draw.rounded_rectangle((36, 24, width - 36, height - 28), radius=14, fill=hex_to_rgb(TOKENS["panel"]), outline=hex_to_rgb(TOKENS["grid"]), width=2)
    return img, draw


def scale_value(value, vmin, vmax, pix_min, pix_max):
    if vmax == vmin:
        return (pix_min + pix_max) / 2
    return pix_min + (value - vmin) * (pix_max - pix_min) / (vmax - vmin)


def nice_label(label: str) -> str:
    return label.replace("_", " ")


def draw_legend(draw, items, x, y):
    for idx, (label, color) in enumerate(items):
        yy = y + idx * 28
        draw.rounded_rectangle((x, yy, x + 22, yy + 14), radius=3, fill=hex_to_rgb(color))
        draw_text(draw, (x + 32, yy - 2), label, fill=hex_to_rgb(TOKENS["muted"]), size=15)


def draw_axes(draw, left, top, right, bottom):
    draw.line((left, bottom, right, bottom), fill=hex_to_rgb(TOKENS["axis"]), width=2)
    draw.line((left, top, left, bottom), fill=hex_to_rgb(TOKENS["axis"]), width=2)


def get_cell_value(ws, coord):
    value = ws[coord].value
    if isinstance(value, str) and value.startswith("="):
        return None
    return value


def load_inputs() -> dict:
    wb = load_workbook(WORKBOOK, data_only=False, read_only=True)
    inputs = wb["Inputs"]
    subs = wb["Subsurface_Inputs"]

    p = {
        "z0_km": float(inputs["C5"].value),
        "target_y_km": float(inputs["C6"].value),
        "x_min_km": float(inputs["C7"].value),
        "x_max_km": float(inputs["C8"].value),
        "delta_z_edge_km": float(inputs["C9"].value),
        "x_edge_km": float(inputs["C10"].value),
        "speed_km_s": float(inputs["C11"].value),
        "ice_index": float(inputs["C12"].value),
        "vhf_wavelength_m": float(inputs["C13"].value),
        "hf_wavelength_m": float(inputs["C14"].value),
        "c_m_s": float(inputs["C21"].value),
        "radius_km": float(inputs["C52"].value),
        "topography_on": str(inputs["C27"].value).upper() == "TRUE",
        "ridge_height_m": float(inputs["C28"].value),
        "ridge_x0_km": float(inputs["C29"].value),
        "ridge_y0_km": float(inputs["C30"].value),
        "ridge_sigma_x_km": float(inputs["C31"].value),
        "ridge_sigma_y_km": float(inputs["C32"].value),
        "crater_depth_m": float(inputs["C33"].value),
        "crater_x0_km": float(inputs["C34"].value),
        "crater_y0_km": float(inputs["C35"].value),
        "crater_sigma_km": float(inputs["C36"].value),
        "chaos_amp_m": float(inputs["C37"].value),
        "rough_amp_m": float(inputs["C38"].value),
        "rough_lx_km": float(inputs["C39"].value),
        "rough_ly_km": float(inputs["C40"].value),
        "trough_depth_m": float(inputs["C41"].value),
        "trough_x0_km": float(inputs["C42"].value),
        "trough_y0_km": float(inputs["C43"].value),
        "trough_sigma_y_km": float(inputs["C44"].value),
        "trough_sigma_x_km": float(inputs["C45"].value),
        "terrain_uncertainty_m": float(inputs["C46"].value),
        "terrain_seed": float(inputs["C47"].value),
        "seeded_amp_m": float(inputs["C48"].value),
        "seeded_lx_km": float(inputs["C49"].value),
        "seeded_ly_km": float(inputs["C50"].value),
        "dem_scale": float(inputs["C51"].value),
        "upper_mean_m": float(subs["C5"].value),
        "upper_sine_amp_m": float(subs["C6"].value),
        "upper_phase_km": float(subs["C7"].value),
        "upper_wavelength_km": float(subs["C8"].value),
        "upper_cos_amp_m": float(subs["C9"].value),
        "upper_cos_wavelength_km": float(subs["C10"].value),
        "upper_surface_coupling": float(subs["C11"].value),
        "lens_a_strength": float(subs["C13"].value),
        "lens_a_center_km": float(subs["C14"].value),
        "lens_a_width_km": float(subs["C15"].value),
        "lens_b_strength": float(subs["C16"].value),
        "lens_b_center_km": float(subs["C17"].value),
        "lens_b_width_km": float(subs["C18"].value),
        "lens_mean_depth_m": float(subs["C19"].value),
        "lens_depth_amp_m": float(subs["C20"].value),
        "lens_depth_phase_km": float(subs["C21"].value),
        "lens_depth_wavelength_km": float(subs["C22"].value),
        "lens_uplift_m": float(subs["C23"].value),
        "ocean_nominal_m": float(subs["C25"].value),
        "ocean_sine_amp_m": float(subs["C26"].value),
        "ocean_sine_phase_km": float(subs["C27"].value),
        "ocean_sine_wavelength_km": float(subs["C28"].value),
        "ocean_cos_amp_m": float(subs["C29"].value),
        "ocean_cos_phase_km": float(subs["C30"].value),
        "ocean_cos_wavelength_km": float(subs["C31"].value),
        "ocean_surface_anticoupling": float(subs["C32"].value),
        "workbook_attenuation_db_km": float(subs["C36"].value),
        "detection_threshold_db": float(subs["C43"].value),
    }
    wb.close()
    return p


def fract(v):
    return v - np.floor(v)


def terrain_height(x_km: np.ndarray, y_km: float, p: dict) -> np.ndarray:
    x = np.asarray(x_km, dtype=float)
    y = float(y_km)
    if not p["topography_on"]:
        return np.zeros_like(x)

    dem = (
        115 * np.exp(-(((y) - 25) ** 2) / (4.5**2))
        + 70 * np.exp(-(((y) + 38) ** 2) / (7**2)) * (0.55 + 0.45 * np.sin(2 * np.pi * x / 45))
        + 24 * np.sin(2 * np.pi * y / 9 + x / 18) * np.exp(-(((y) + 5) ** 2) / (22**2))
        + 18 * np.sin(2 * np.pi * x / 70)
    )

    ridge = (
        p["ridge_height_m"]
        * np.exp(-((x - p["ridge_x0_km"]) ** 2) / (2 * p["ridge_sigma_x_km"] ** 2))
        * np.exp(-((y - p["ridge_y0_km"]) ** 2) / (2 * p["ridge_sigma_y_km"] ** 2))
    )
    crater = -p["crater_depth_m"] * np.exp(
        -(((x - p["crater_x0_km"]) ** 2) + ((y - p["crater_y0_km"]) ** 2)) / (2 * p["crater_sigma_km"] ** 2)
    )
    chaos = p["chaos_amp_m"] * (
        0.55 * np.sin(2 * np.pi * x / p["rough_lx_km"])
        + 0.35 * np.cos(2 * np.pi * y / p["rough_ly_km"])
        + 0.2 * np.sin(2 * np.pi * (x + y) / (0.5 * (p["rough_lx_km"] + p["rough_ly_km"])))
    )
    rough = p["rough_amp_m"] * np.sin(2 * np.pi * x / p["rough_lx_km"]) * np.sin(2 * np.pi * y / p["rough_ly_km"])
    trough = -p["trough_depth_m"] * np.exp(-((y - p["trough_y0_km"]) ** 2) / (2 * p["trough_sigma_y_km"] ** 2)) * np.exp(
        -((x - p["trough_x0_km"]) ** 2) / (2 * p["trough_sigma_x_km"] ** 2)
    )
    seeded_wave = (
        0.45 * np.sin(2 * np.pi * ((x / p["seeded_lx_km"]) + p["terrain_seed"] * 0.137))
        + 0.35 * np.cos(2 * np.pi * ((y / p["seeded_ly_km"]) + p["terrain_seed"] * 0.173))
        + 0.2
        * np.sin(
            2
            * np.pi
            * ((x + y) / (0.5 * (p["seeded_lx_km"] + p["seeded_ly_km"])) + p["terrain_seed"] * 0.097)
        )
    )
    pseudo_random = 2 * fract(np.sin(x * 12.9898 + y * 78.233 + p["terrain_seed"] * 37.719) * 43758.5453) - 1
    seeded = p["seeded_amp_m"] * (0.7 * seeded_wave + 0.3 * pseudo_random)

    return p["dem_scale"] * dem + ridge + crater + chaos + rough + trough + seeded


def compute_baseline(p: dict) -> pd.DataFrame:
    x = np.linspace(p["x_min_km"], p["x_max_km"], 241)
    surface = terrain_height(x, p["target_y_km"], p)
    surface_mean = surface.mean()

    upper = (
        p["upper_mean_m"]
        + p["upper_sine_amp_m"] * np.sin(2 * np.pi * (x + p["upper_phase_km"]) / p["upper_wavelength_km"])
        + p["upper_cos_amp_m"] * np.cos(2 * np.pi * x / p["upper_cos_wavelength_km"])
        + p["upper_surface_coupling"] * (surface - surface_mean)
    )
    lens_strength = np.clip(
        p["lens_a_strength"] * np.exp(-0.5 * ((x - p["lens_a_center_km"]) / p["lens_a_width_km"]) ** 2)
        + p["lens_b_strength"] * np.exp(-0.5 * ((x - p["lens_b_center_km"]) / p["lens_b_width_km"]) ** 2),
        0,
        1,
    )
    lens_depth = (
        p["lens_mean_depth_m"]
        + p["lens_depth_amp_m"] * np.sin(2 * np.pi * (x - p["lens_depth_phase_km"]) / p["lens_depth_wavelength_km"])
        - p["lens_uplift_m"] * lens_strength
    )
    ocean_depth = (
        p["ocean_nominal_m"]
        + p["ocean_sine_amp_m"] * np.sin(2 * np.pi * (x + p["ocean_sine_phase_km"]) / p["ocean_sine_wavelength_km"])
        + p["ocean_cos_amp_m"] * np.cos(2 * np.pi * (x - p["ocean_cos_phase_km"]) / p["ocean_cos_wavelength_km"])
        - p["ocean_surface_anticoupling"] * (surface - surface_mean)
    )

    slope = np.gradient(surface, x * 1000)
    slope_deg = np.degrees(np.arctan(slope))

    return pd.DataFrame(
        {
            "x_km": x,
            "surface_height_m": surface,
            "surface_slope_deg": slope_deg,
            "upper_depth_m": upper,
            "lens_strength": lens_strength,
            "lens_depth_m": lens_depth,
            "ocean_depth_m": ocean_depth,
        }
    )


def attenuation_scale(freq_mhz: float) -> float:
    return math.sqrt(freq_mhz / 9.0)


def reflection_power(eps_a: float, eps_b: float) -> float:
    return reflection_amplitude(eps_a, eps_b) ** 2


def reflection_amplitude(eps_a: float, eps_b: float) -> float:
    na = math.sqrt(max(eps_a, 1e-9))
    nb = math.sqrt(max(eps_b, 1e-9))
    return (nb - na) / (nb + na)


def power_to_db(power: float) -> float:
    return 10 * math.log10(max(power, 1e-18))


def gaussian_patch(x_km: float, center_km: float, width_km: float) -> float:
    return math.exp(-0.5 * ((x_km - center_km) / width_km) ** 2)


def layers_for_scenario(scenario: str, row: pd.Series) -> list[Layer]:
    x = float(row.x_km)
    lens_strength = float(row.lens_strength)
    lens_depth = float(row.lens_depth_m)
    ocean_depth = float(row.ocean_depth_m)
    layers: list[Layer] = []

    if scenario in ("salt_reference", "complex_dirty_ice"):
        salt_depth_1 = min(max(0.5 * ocean_depth, 1800), ocean_depth - 800)
        salt_depth_2 = min(10000 + 350 * math.sin(2 * math.pi * x / 80), ocean_depth - 700)
        layers.extend(
            [
                Layer(salt_depth_1, 1.0, 5.0, 2.0, "1 m hydrated salt layer"),
                Layer(salt_depth_2, 0.2, 5.0, 2.0, "0.2 m melt-derived salt layer"),
            ]
        )

    if scenario in ("mixed_dirty_thin_layers", "complex_dirty_ice"):
        mix_envelope = max(gaussian_patch(x, -30, 18), gaussian_patch(x, 22, 20), 0.45 * gaussian_patch(x, 48, 13))
        if mix_envelope <= 0.55:
            mix_envelope = 0.0
        stack_center = 0.72 * ocean_depth + 520 * math.sin(2 * math.pi * (x + 14) / 78)
        # Multiple sub-resolution interfaces: the exact values are simulation parameters,
        # not claims that these exact layers exist on Europa.
        offsets = [-210, -125, -55, 25, 95, 185]
        eps_values = [4.0, 6.8, 3.7, 7.5, 4.9, 6.1]
        atten_values = [3.5, 5.0, 2.5, 6.5, 4.0, 5.5]
        thickness_values = [22, 18, 35, 16, 42, 24]
        if mix_envelope > 0:
            for off, eps, att, thick in zip(offsets, eps_values, atten_values, thickness_values):
                contrast_scale = 0.45 + 0.55 * mix_envelope
                eps_scaled = ICE_EPS + (eps - ICE_EPS) * contrast_scale
                att_scaled = att * (0.55 + 0.70 * mix_envelope)
                thick_scaled = thick * (0.65 + 0.55 * mix_envelope)
                top = min(max(stack_center + off, 900), ocean_depth - 900)
                layers.append(Layer(top, thick_scaled, eps_scaled, att_scaled, "stacked dirty thin layer"))

    if scenario in ("briny_lens_masking", "complex_dirty_ice") and lens_strength > 0.08:
        thickness = 80 + 720 * lens_strength
        eps = 10 + 18 * lens_strength
        att = 6 + 18 * lens_strength
        top = min(max(lens_depth - thickness / 2, 600), ocean_depth - thickness - 600)
        layers.append(Layer(top, thickness, eps, att, "warm/briny conductive lens"))

    if scenario in ("non_salt_impurity_mix", "complex_dirty_ice"):
        # Acid/dust/organics are deliberately represented as an uncertain family:
        # their dielectric properties at Europa conditions are less well constrained.
        non_salt_envelope = max(gaussian_patch(x, -45, 15), 0.7 * gaussian_patch(x, 5, 24), gaussian_patch(x, 36, 17))
        if non_salt_envelope > 0.28:
            phase = math.sin(2 * math.pi * (x + 30) / 65)
            depths = [
                0.22 * ocean_depth + 220 * phase,
                0.42 * ocean_depth - 160 * phase,
                0.66 * ocean_depth + 180 * math.cos(2 * math.pi * x / 50),
            ]
            for depth, eps, att, thick, label in [
                (depths[0], 4.4, 3.2, 70, "radiolytic/acid-rich dirty layer"),
                (depths[1], 5.8, 4.8, 120, "dust/organic mixed dirty layer"),
                (depths[2], 7.2, 6.2, 55, "high-contrast dirty inclusion layer"),
            ]:
                contrast_scale = 0.50 + 0.50 * non_salt_envelope
                eps_scaled = ICE_EPS + (eps - ICE_EPS) * contrast_scale
                att_scaled = att * (0.55 + 0.65 * non_salt_envelope)
                thick_scaled = thick * (0.70 + 0.45 * non_salt_envelope)
                top = min(max(depth, 300), ocean_depth - thick_scaled - 800)
                layers.append(Layer(top, thick_scaled, eps_scaled, att_scaled, label))

    return sorted(layers, key=lambda layer: layer.top_m)


def attenuation_to_depth(depth_m: float, layers: list[Layer], freq_mhz: float, clean_att_hf_db_km: float) -> float:
    scale = attenuation_scale(freq_mhz)
    clean_att = clean_att_hf_db_km * scale
    depth_remaining_m = max(depth_m, 0.0)
    layer_loss = 0.0
    layer_overlap_total = 0.0
    for layer in layers:
        overlap = max(0.0, min(depth_m, layer.bottom_m) - layer.top_m)
        if overlap > 0:
            layer_overlap_total += overlap
            layer_loss += layer.attenuation_db_km_hf * scale * (overlap / 1000.0)
    clean_path_m = max(0.0, depth_remaining_m - layer_overlap_total)
    return clean_att * (clean_path_m / 1000.0) + layer_loss


def interface_records(layers: list[Layer], ocean_depth_m: float) -> list[dict]:
    records = []
    last_eps = ICE_EPS
    for layer in layers:
        records.append({"depth_m": layer.top_m, "eps_before": last_eps, "eps_after": layer.eps, "label": f"top {layer.label}"})
        records.append({"depth_m": layer.bottom_m, "eps_before": layer.eps, "eps_after": ICE_EPS, "label": f"base {layer.label}"})
        last_eps = ICE_EPS
    records.append({"depth_m": ocean_depth_m, "eps_before": ICE_EPS, "eps_after": OCEAN_EPS, "label": "true ice-ocean boundary"})
    return sorted(records, key=lambda item: item["depth_m"])


def strongest_false_echo(
    layers: list[Layer],
    ocean_depth_m: float,
    freq_mhz: float,
    resolution_m: float,
    clean_att_hf_db_km: float,
) -> tuple[float, float, str]:
    records = [rec for rec in interface_records(layers, ocean_depth_m) if rec["label"] != "true ice-ocean boundary"]
    if not records:
        return -999.0, np.nan, "none"

    best_db = -999.0
    best_depth = np.nan
    best_label = "none"
    for i, rec in enumerate(records):
        center = rec["depth_m"]
        group = [r for r in records if abs(r["depth_m"] - center) <= resolution_m / 2]
        amps = []
        labels = []
        for r in group:
            refl_amp = reflection_amplitude(r["eps_before"], r["eps_after"])
            one_way_loss = attenuation_to_depth(r["depth_m"], layers, freq_mhz, clean_att_hf_db_km)
            # phase of a two-way return in ice
            wavelength_ice_m = 299_792_458.0 / (freq_mhz * 1_000_000.0) / math.sqrt(ICE_EPS)
            phase = 4 * math.pi * r["depth_m"] / wavelength_ice_m
            amp = refl_amp * 10 ** ((SYSTEM_OFFSET_DB - 2 * one_way_loss) / 20.0) * complex(math.cos(phase), math.sin(phase))
            amps.append(amp)
            labels.append(r["label"])
        combined_power_db = 20 * math.log10(max(abs(sum(amps)), 1e-12))
        if combined_power_db > best_db:
            best_db = combined_power_db
            best_depth = center
            best_label = "; ".join(sorted(set(labels)))[:120]
    return best_db, best_depth, best_label


def ocean_echo(
    layers: list[Layer],
    ocean_depth_m: float,
    freq_mhz: float,
    clean_att_hf_db_km: float,
) -> float:
    refl_db = power_to_db(reflection_power(ICE_EPS, OCEAN_EPS))
    one_way_loss = attenuation_to_depth(ocean_depth_m, layers, freq_mhz, clean_att_hf_db_km)
    return SYSTEM_OFFSET_DB + refl_db - 2 * one_way_loss


def classify(ocean_db: float, false_db: float, false_depth_m: float, ocean_depth_m: float) -> str:
    ocean_detected = ocean_db >= DETECTION_THRESHOLD_DB
    false_detected = false_db >= DETECTION_THRESHOLD_DB
    deep_confuser = False
    if np.isfinite(false_depth_m) and ocean_depth_m > 0:
        false_depth_ratio = false_depth_m / ocean_depth_m
        deep_confuser = 0.62 <= false_depth_ratio <= 1.05
    if ocean_detected and (false_db < ocean_db - 3):
        return "clear ocean boundary"
    if ocean_detected and false_detected and not deep_confuser:
        return "clear ocean boundary"
    if ocean_detected and false_detected and deep_confuser and false_db >= ocean_db - 3 and false_db <= ocean_db + 3:
        return "ambiguous"
    if false_detected and deep_confuser and false_db > ocean_db + 3:
        return "false layer stronger"
    if not ocean_detected and false_detected and deep_confuser:
        return "ocean hidden, false layer visible"
    return "weak/no deep detection"


def run_simulation(baseline: pd.DataFrame) -> pd.DataFrame:
    scenarios = {
        "clean_ice_control": {
            "description": "No dirty layers; only clean ice attenuation.",
            "clean_att_hf_db_km": 0.55,
        },
        "salt_reference": {
            "description": "REASON-like meter-scale hydrated salt target layers with epsilon near 5.",
            "clean_att_hf_db_km": 0.65,
        },
        "mixed_dirty_thin_layers": {
            "description": "Sub-resolution stack of dirty layers; tests false bright reflectors.",
            "clean_att_hf_db_km": 0.75,
        },
        "briny_lens_masking": {
            "description": "Localized conductive briny/warm lens that reflects and attenuates deeper energy.",
            "clean_att_hf_db_km": 0.75,
        },
        "non_salt_impurity_mix": {
            "description": "Acid/dust/organic-rich layers; high uncertainty family.",
            "clean_att_hf_db_km": 0.85,
        },
        "complex_dirty_ice": {
            "description": "Combined salt + dirty stack + briny lens + non-salt layers.",
            "clean_att_hf_db_km": 0.85,
        },
    }

    rows = []
    for _, row in baseline.iterrows():
        for scenario, cfg in scenarios.items():
            layers = layers_for_scenario(scenario, row)
            for freq_name, fcfg in FREQUENCIES.items():
                false_db, false_depth, false_label = strongest_false_echo(
                    layers,
                    float(row.ocean_depth_m),
                    fcfg["mhz"],
                    fcfg["resolution_m"],
                    cfg["clean_att_hf_db_km"],
                )
                ocean_db = ocean_echo(layers, float(row.ocean_depth_m), fcfg["mhz"], cfg["clean_att_hf_db_km"])
                strongest_depth = false_depth if false_db > ocean_db else float(row.ocean_depth_m)
                inferred_error = strongest_depth - float(row.ocean_depth_m) if not np.isnan(strongest_depth) else np.nan
                rows.append(
                    {
                        "x_km": row.x_km,
                        "scenario": scenario,
                        "frequency": freq_name,
                        "frequency_mhz": fcfg["mhz"],
                        "frequency_role": fcfg["role"],
                        "surface_height_m": row.surface_height_m,
                        "surface_slope_deg": row.surface_slope_deg,
                        "upper_depth_m": row.upper_depth_m,
                        "lens_strength": row.lens_strength,
                        "lens_depth_m": row.lens_depth_m,
                        "ocean_depth_m": row.ocean_depth_m,
                        "layer_count": len(layers),
                        "ocean_echo_db": ocean_db,
                        "false_echo_db": false_db,
                        "false_depth_m": false_depth,
                        "false_label": false_label,
                        "false_minus_ocean_db": false_db - ocean_db,
                        "ocean_margin_db": ocean_db - DETECTION_THRESHOLD_DB,
                        "false_margin_db": false_db - DETECTION_THRESHOLD_DB,
                        "inferred_depth_if_strongest_m": strongest_depth,
                        "inferred_depth_error_m": inferred_error,
                        "classification": classify(ocean_db, false_db, false_depth, float(row.ocean_depth_m)),
                    }
                )
    return pd.DataFrame(rows)


def summarize(results: pd.DataFrame) -> pd.DataFrame:
    grouped = []
    for (scenario, frequency), g in results.groupby(["scenario", "frequency"]):
        counts = g["classification"].value_counts(normalize=True).mul(100)
        grouped.append(
            {
                "scenario": scenario,
                "frequency": frequency,
                "points": len(g),
                "clear_ocean_pct": counts.get("clear ocean boundary", 0.0),
                "ambiguous_pct": counts.get("ambiguous", 0.0),
                "false_stronger_pct": counts.get("false layer stronger", 0.0),
                "hidden_false_visible_pct": counts.get("ocean hidden, false layer visible", 0.0),
                "weak_no_deep_detection_pct": counts.get("weak/no deep detection", 0.0),
                "median_ocean_margin_db": g["ocean_margin_db"].median(),
                "min_ocean_margin_db": g["ocean_margin_db"].min(),
                "max_false_minus_ocean_db": g["false_minus_ocean_db"].max(),
                "max_abs_inferred_depth_error_m": g["inferred_depth_error_m"].abs().max(),
                "median_layer_count": g["layer_count"].median(),
            }
        )
    return pd.DataFrame(grouped).sort_values(["scenario", "frequency"])


def sensitivity_grid(baseline: pd.DataFrame) -> pd.DataFrame:
    thicknesses = [1, 5, 10, 30, 60, 100, 200, 500]
    attenuations = [1, 2, 4, 6, 8, 12, 16, 20]
    rows = []
    sample = baseline.iloc[::4].copy()
    for thickness in thicknesses:
        for att in attenuations:
            classifications = []
            ocean_margins = []
            false_over = []
            for _, row in sample.iterrows():
                layer_top = min(0.72 * row.ocean_depth_m + 400 * math.sin(2 * math.pi * row.x_km / 80), row.ocean_depth_m - thickness - 800)
                layers = [Layer(layer_top, thickness, 7.5, att, "sensitivity dirty layer")]
                # Use HF because it is the deep-sounding band most likely to reach the ocean.
                false_db, false_depth, _ = strongest_false_echo(layers, row.ocean_depth_m, 9.0, 300.0, 0.75)
                ocean_db = ocean_echo(layers, row.ocean_depth_m, 9.0, 0.75)
                classifications.append(classify(ocean_db, false_db, false_depth, row.ocean_depth_m))
                ocean_margins.append(ocean_db - DETECTION_THRESHOLD_DB)
                false_over.append(false_db - ocean_db)
            s = pd.Series(classifications)
            rows.append(
                {
                    "dirty_layer_thickness_m": thickness,
                    "dirty_layer_attenuation_db_km_hf": att,
                    "ambiguous_or_false_pct": 100
                    * s.isin(["ambiguous", "false layer stronger", "ocean hidden, false layer visible"]).mean(),
                    "ocean_hidden_pct": 100 * s.isin(["ocean hidden, false layer visible", "weak/no deep detection"]).mean(),
                    "median_ocean_margin_db": float(np.median(ocean_margins)),
                    "median_false_minus_ocean_db": float(np.median(false_over)),
                }
            )
    return pd.DataFrame(rows)


def paper_cross_reference() -> pd.DataFrame:
    rows = [
        {
            "source": "Blankenship et al. 2024, REASON",
            "paper_or_source_claim": "REASON is a dual-frequency 9/60 MHz radar and is designed to characterize brines/salts, search for an ice-ocean interface, and handle uncertainty in key Europa properties.",
            "simulation_choice": "Model uses HF_9MHz and VHF_60MHz; dirty layers alter reflectivity/attenuation; outputs include ocean-vs-false-layer ambiguity.",
            "status": "Directly aligned",
            "link": "https://link.springer.com/article/10.1007/s11214-024-01072-3",
        },
        {
            "source": "Blankenship et al. 2024, thin layers",
            "paper_or_source_claim": "Thin layers can cause constructive/destructive interference depending on thickness and permittivity; REASON considers 0.5 m and 200 m near-surface modes.",
            "simulation_choice": "False reflector calculation coherently sums interfaces inside the radar vertical-resolution window.",
            "status": "Simplified forward model",
            "link": "https://link.springer.com/article/10.1007/s11214-024-01072-3",
        },
        {
            "source": "Blankenship et al. 2024, salt layers",
            "paper_or_source_claim": "Salt layers are modeled as potential target interfaces; paper adopts epsilon=5 and meter-scale salt layers.",
            "simulation_choice": "Salt_reference scenario uses epsilon=5, 1.0 m and 0.2 m salt layers.",
            "status": "Directly aligned",
            "link": "https://link.springer.com/article/10.1007/s11214-024-01072-3",
        },
        {
            "source": "Blankenship et al. 2024, mushy layer",
            "paper_or_source_claim": "REASON is not designed to penetrate an increasingly mushy brine layer because it is highly conductive and attenuating, but the top interface becomes more reflective.",
            "simulation_choice": "Briny_lens_masking scenario makes the lens both reflective and high-loss, then checks whether it hides the ocean echo.",
            "status": "Qualitatively aligned",
            "link": "https://link.springer.com/article/10.1007/s11214-024-01072-3",
        },
        {
            "source": "Lalich et al. 2021 / Mars analog",
            "paper_or_source_claim": "Bright radar reflections under Mars polar ice can be explained by interference between multiple layer boundaries without liquid water.",
            "simulation_choice": "Mixed_dirty_thin_layers tests whether stacked interfaces can outshine the ocean boundary in a Europa-like setup.",
            "status": "Analog support, not Europa proof",
            "link": "https://arxiv.org/abs/2107.03497",
        },
        {
            "source": "Pettinelli et al. 2015 review",
            "paper_or_source_claim": "Dielectric properties of Jovian satellite ice analogs are central to radar interpretation and remain measurement/model dependent.",
            "simulation_choice": "Non_salt_impurity_mix varies epsilon and attenuation for less-constrained acid/dust/organic/radiolytic families.",
            "status": "Uncertainty bracket",
            "link": "https://doi.org/10.1002/2014RG000463",
        },
    ]
    return pd.DataFrame(rows)


def clamp(value: float, low: float, high: float) -> float:
    return max(low, min(high, value))


def format_tick(value: float) -> str:
    if abs(value) >= 100:
        return f"{value:.0f}"
    if abs(value) >= 10:
        return f"{value:.0f}"
    return f"{value:.1f}".rstrip("0").rstrip(".")


def draw_plot_frame(
    draw: ImageDraw.ImageDraw,
    left: int,
    top: int,
    right: int,
    bottom: int,
    x_min: float,
    x_max: float,
    ticks: list[float],
    x_label: str,
    percent: bool = False,
) -> None:
    draw_axes(draw, left, top, right, bottom)
    for tick in ticks:
        x = int(scale_value(tick, x_min, x_max, left, right))
        draw.line((x, top, x, bottom), fill=hex_to_rgb(TOKENS["grid"]), width=1)
        label = f"{tick:.0f}%" if percent else format_tick(tick)
        draw_text(draw, (x, bottom + 14), label, fill=hex_to_rgb(TOKENS["muted"]), size=14, anchor="mt")
    draw_text(draw, ((left + right) // 2, bottom + 58), x_label, fill=hex_to_rgb(TOKENS["muted"]), size=17, anchor="mm")


def draw_y_grid(
    draw: ImageDraw.ImageDraw,
    left: int,
    top: int,
    right: int,
    bottom: int,
    y_min: float,
    y_max: float,
    ticks: list[float],
) -> None:
    for tick in ticks:
        y = int(scale_value(tick, y_min, y_max, bottom, top))
        draw.line((left, y, right, y), fill=hex_to_rgb(TOKENS["grid"]), width=1)
        draw_text(draw, (left - 12, y), format_tick(tick), fill=hex_to_rgb(TOKENS["muted"]), size=14, anchor="rm")


def draw_boxplot(
    draw: ImageDraw.ImageDraw,
    values: np.ndarray,
    y: int,
    left: int,
    right: int,
    x_min: float,
    x_max: float,
    fill: str,
    outline: str,
    height: int = 34,
) -> None:
    values = np.array([v for v in values if np.isfinite(v)], dtype=float)
    if values.size == 0:
        return
    q0, q1, q2, q3, q4 = np.percentile(values, [0, 25, 50, 75, 100])
    x0 = int(scale_value(q0, x_min, x_max, left, right))
    x1 = int(scale_value(q1, x_min, x_max, left, right))
    x2 = int(scale_value(q2, x_min, x_max, left, right))
    x3 = int(scale_value(q3, x_min, x_max, left, right))
    x4 = int(scale_value(q4, x_min, x_max, left, right))
    color = hex_to_rgb(outline)
    draw.line((x0, y, x4, y), fill=color, width=3)
    draw.line((x0, y - height // 3, x0, y + height // 3), fill=color, width=3)
    draw.line((x4, y - height // 3, x4, y + height // 3), fill=color, width=3)
    draw.rounded_rectangle((x1, y - height // 2, x3, y + height // 2), radius=4, fill=hex_to_rgb(fill), outline=color, width=2)
    draw.line((x2, y - height // 2 - 4, x2, y + height // 2 + 4), fill=hex_to_rgb(COLORS["dark"]), width=3)


def plot_outcomes(summary: pd.DataFrame) -> Path:
    order = [
        "clean_ice_control",
        "salt_reference",
        "mixed_dirty_thin_layers",
        "briny_lens_masking",
        "non_salt_impurity_mix",
        "complex_dirty_ice",
    ]
    plot_df = summary.copy()
    plot_df["risk_or_ambiguity_pct"] = (
        plot_df["ambiguous_pct"] + plot_df["false_stronger_pct"] + plot_df["hidden_false_visible_pct"]
    )
    img, draw = new_chart_canvas(1500, 900)
    draw_chart_header(
        draw,
        "Dirty-ice ambiguity rises when layers are mixed or briny",
        "Share of along-track points where the strongest radar return is ambiguous, false, or hides the ocean boundary.",
        1500,
    )
    left, top, right, bottom = 430, 165, 1320, 725
    ticks = [0, 20, 40, 60, 80, 100]
    draw_plot_frame(draw, left, top, right, bottom, 0, 100, ticks, "Ambiguous / false-layer outcome share", percent=True)
    band_height = 24
    group_gap = (bottom - top) / len(order)
    for idx, scenario in enumerate(order):
        center = int(top + group_gap * (idx + 0.5))
        label = nice_label(scenario)
        draw_text(draw, (left - 24, center), label, fill=hex_to_rgb(TOKENS["ink"]), size=17, anchor="rm")
        values = [
            ("HF_9MHz", COLORS["blue"], int(center - band_height - 3)),
            ("VHF_60MHz", COLORS["orange"], int(center + 3)),
        ]
        for band, color, y0 in values:
            pct = float(plot_df[(plot_df.scenario == scenario) & (plot_df.frequency == band)]["risk_or_ambiguity_pct"].iloc[0])
            x1 = int(scale_value(pct, 0, 100, left, right))
            draw.rounded_rectangle((left, y0, x1, y0 + band_height), radius=5, fill=hex_to_rgb(color))
            label_x = min(x1 + 10, right - 44)
            draw_text(draw, (label_x, y0 + band_height // 2), f"{pct:.1f}%", fill=hex_to_rgb(TOKENS["ink"]), size=14, anchor="lm")
    draw_legend(draw, [("HF 9 MHz", COLORS["blue"]), ("VHF 60 MHz", COLORS["orange"])], 1160, 760)
    return save_image(img, "scenario_outcome_risk.png")


def plot_echo_track(results: pd.DataFrame) -> Path:
    plot_df = results[(results["scenario"] == "complex_dirty_ice") & (results["frequency"] == "HF_9MHz")].copy()
    plot_df = plot_df.sort_values("x_km")
    img, draw = new_chart_canvas(1500, 900)
    draw_chart_header(
        draw,
        "False dirty-layer echoes can compete with the ocean return",
        "Complex dirty-ice scenario, HF 9 MHz. Higher values are easier to detect; threshold is the model cutoff.",
        1500,
    )
    left, top, right, bottom = 120, 165, 1320, 720
    x_min, x_max = float(plot_df["x_km"].min()), float(plot_df["x_km"].max())
    y_values = np.concatenate([
        plot_df["ocean_echo_db"].to_numpy(dtype=float),
        plot_df["false_echo_db"].to_numpy(dtype=float),
        np.array([DETECTION_THRESHOLD_DB]),
    ])
    y_min = math.floor((float(np.nanmin(y_values)) - 8) / 10) * 10
    y_max = math.ceil((float(np.nanmax(y_values)) + 8) / 10) * 10
    x_ticks = list(np.linspace(x_min, x_max, 7))
    y_ticks = list(np.linspace(y_min, y_max, 7))
    draw_plot_frame(draw, left, top, right, bottom, x_min, x_max, x_ticks, "Along-track position x (km)")
    draw_y_grid(draw, left, top, right, bottom, y_min, y_max, y_ticks)
    draw_text(draw, (left, top - 28), "Relative echo power (dB)", fill=hex_to_rgb(TOKENS["muted"]), size=17, anchor="lm")

    x_vals = plot_df["x_km"].to_numpy(dtype=float)
    ocean_vals = plot_df["ocean_echo_db"].to_numpy(dtype=float)
    false_vals = plot_df["false_echo_db"].to_numpy(dtype=float)
    risk_classes = {"ambiguous", "false layer stronger", "ocean hidden, false layer visible"}
    classes = plot_df["classification"].tolist()
    for i in range(len(plot_df) - 1):
        if classes[i] in risk_classes or classes[i + 1] in risk_classes:
            x0 = int(scale_value(x_vals[i], x_min, x_max, left, right))
            x1 = int(scale_value(x_vals[i + 1], x_min, x_max, left, right))
            draw.rectangle((x0, top, x1, bottom), fill=blend_color(TOKENS["panel"], COLORS["orange_light"], 0.32))

    threshold_y = int(scale_value(DETECTION_THRESHOLD_DB, y_min, y_max, bottom, top))
    for x0 in range(left, right, 18):
        draw.line((x0, threshold_y, min(x0 + 10, right), threshold_y), fill=hex_to_rgb(COLORS["neutral"]), width=2)

    ocean_points = [(int(scale_value(x, x_min, x_max, left, right)), int(scale_value(y, y_min, y_max, bottom, top))) for x, y in zip(x_vals, ocean_vals)]
    false_points = [(int(scale_value(x, x_min, x_max, left, right)), int(scale_value(y, y_min, y_max, bottom, top))) for x, y in zip(x_vals, false_vals)]
    draw.line(ocean_points, fill=hex_to_rgb(COLORS["blue"]), width=4, joint="curve")
    draw.line(false_points, fill=hex_to_rgb(COLORS["orange"]), width=4, joint="curve")
    draw_legend(
        draw,
        [
            ("True ocean boundary echo", COLORS["blue"]),
            ("Strongest dirty-layer false echo", COLORS["orange"]),
            ("Threshold / shaded deep-risk zones", COLORS["neutral"]),
        ],
        850,
        760,
    )
    return save_image(img, "complex_dirty_ice_hf_echo_track.png")


def plot_depth_error(results: pd.DataFrame) -> Path:
    scenario_order = ["salt_reference", "mixed_dirty_thin_layers", "briny_lens_masking", "complex_dirty_ice"]
    plot_df = results[(results["frequency"] == "HF_9MHz") & (results["scenario"].isin(scenario_order))].copy()
    plot_df["abs_error_km"] = plot_df["inferred_depth_error_m"].abs() / 1000
    img, draw = new_chart_canvas(1500, 860)
    draw_chart_header(
        draw,
        "False strongest echoes create kilometer-scale depth mistakes",
        "HF 9 MHz; error is zero when the true ocean echo remains strongest.",
        1500,
    )
    left, top, right, bottom = 430, 170, 1290, 690
    max_error = max(1.0, float(plot_df["abs_error_km"].max()))
    x_max = math.ceil(max_error * 1.15)
    ticks = list(np.linspace(0, x_max, 6))
    draw_plot_frame(draw, left, top, right, bottom, 0, x_max, ticks, "Absolute depth error if strongest echo is used (km)")
    group_gap = (bottom - top) / len(scenario_order)
    for idx, scenario in enumerate(scenario_order):
        y = int(top + group_gap * (idx + 0.5))
        values = plot_df[plot_df["scenario"] == scenario]["abs_error_km"].to_numpy(dtype=float)
        draw_text(draw, (left - 24, y), nice_label(scenario), fill=hex_to_rgb(TOKENS["ink"]), size=17, anchor="rm")
        draw_boxplot(draw, values, y, left, right, 0, x_max, COLORS["blue_light"], COLORS["blue"])
    return save_image(img, "hf_depth_error_distribution.png")


def plot_sensitivity(sens: pd.DataFrame) -> Path:
    pivot = sens.pivot(
        index="dirty_layer_attenuation_db_km_hf",
        columns="dirty_layer_thickness_m",
        values="ambiguous_or_false_pct",
    ).sort_index(ascending=True)
    img, draw = new_chart_canvas(1450, 900)
    draw_chart_header(
        draw,
        "Risk threshold depends on both layer thickness and loss",
        "Single dirty layer near 70% of the modeled ocean depth, HF 9 MHz; values show ambiguity or false-boundary risk.",
        1450,
    )
    left, top, right, bottom = 190, 165, 1125, 700
    rows = list(pivot.index)
    cols = list(pivot.columns)
    cell_w = (right - left) / len(cols)
    cell_h = (bottom - top) / len(rows)
    max_value = max(1.0, float(np.nanmax(pivot.values)))
    for i, attenuation in enumerate(rows):
        for j, thickness in enumerate(cols):
            value = float(pivot.loc[attenuation, thickness])
            x0 = int(left + j * cell_w)
            x1 = int(left + (j + 1) * cell_w)
            y0 = int(top + i * cell_h)
            y1 = int(top + (i + 1) * cell_h)
            t = clamp(value / max_value, 0, 1)
            fill = blend_color("#FFFFFF", COLORS["orange"], t)
            draw.rectangle((x0, y0, x1, y1), fill=fill, outline=hex_to_rgb(TOKENS["grid"]))
            draw_text(draw, ((x0 + x1) // 2, (y0 + y1) // 2), f"{value:.0f}", fill=hex_to_rgb(TOKENS["ink"]), size=15, anchor="mm")
    for j, thickness in enumerate(cols):
        x = int(left + (j + 0.5) * cell_w)
        draw_text(draw, (x, bottom + 14), format_tick(thickness), fill=hex_to_rgb(TOKENS["muted"]), size=14, anchor="mt")
    for i, attenuation in enumerate(rows):
        y = int(top + (i + 0.5) * cell_h)
        draw_text(draw, (left - 14, y), format_tick(attenuation), fill=hex_to_rgb(TOKENS["muted"]), size=14, anchor="rm")
    draw_text(draw, ((left + right) // 2, bottom + 58), "Dirty layer thickness (m)", fill=hex_to_rgb(TOKENS["muted"]), size=17, anchor="mm")
    draw_text(draw, (left, top - 28), "Dirty layer attenuation at 9 MHz (dB/km)", fill=hex_to_rgb(TOKENS["muted"]), size=17, anchor="lm")
    cbar_left, cbar_top, cbar_right, cbar_bottom = 1190, 205, 1240, 665
    for y in range(cbar_top, cbar_bottom):
        t = 1 - (y - cbar_top) / (cbar_bottom - cbar_top)
        draw.line((cbar_left, y, cbar_right, y), fill=blend_color("#FFFFFF", COLORS["orange"], t))
    draw.rectangle((cbar_left, cbar_top, cbar_right, cbar_bottom), outline=hex_to_rgb(TOKENS["grid"]), width=1)
    draw_text(draw, (1258, cbar_top), f"{max_value:.0f}%", fill=hex_to_rgb(TOKENS["muted"]), size=14, anchor="lm")
    draw_text(draw, (1258, cbar_bottom), "0%", fill=hex_to_rgb(TOKENS["muted"]), size=14, anchor="lm")
    draw_text(draw, (1190, 172), "Ambiguous / false outcome", fill=hex_to_rgb(TOKENS["muted"]), size=16, anchor="lm")
    return save_image(img, "dirty_layer_sensitivity_heatmap.png")


def plot_frequency_margin(results: pd.DataFrame) -> Path:
    scenario_order = ["clean_ice_control", "salt_reference", "complex_dirty_ice"]
    plot_df = results[results["scenario"].isin(scenario_order)].copy()
    img, draw = new_chart_canvas(1500, 860)
    draw_chart_header(
        draw,
        "HF is the deep-sounding band; VHF loses the deep ocean first",
        "Positive margin means the ocean boundary is detectable in this simplified model.",
        1500,
    )
    left, top, right, bottom = 410, 170, 1285, 680
    values = plot_df["ocean_margin_db"].to_numpy(dtype=float)
    x_min = math.floor((min(float(np.nanmin(values)), 0.0) - 5) / 10) * 10
    x_max = math.ceil((max(float(np.nanmax(values)), 0.0) + 5) / 10) * 10
    ticks = [tick for tick in np.arange(math.ceil(x_min / 30) * 30, x_max + 1, 30)]
    ticks = sorted(set([float(tick) for tick in ticks] + [0.0]))
    draw_plot_frame(draw, left, top, right, bottom, x_min, x_max, ticks, "Ocean echo margin above detection threshold (dB)")
    zero_x = int(scale_value(0, x_min, x_max, left, right))
    for y0 in range(top, bottom, 18):
        draw.line((zero_x, y0, zero_x, min(y0 + 10, bottom)), fill=hex_to_rgb(COLORS["neutral"]), width=2)
    group_gap = (bottom - top) / len(scenario_order)
    for idx, scenario in enumerate(scenario_order):
        center = int(top + group_gap * (idx + 0.5))
        draw_text(draw, (left - 26, center), nice_label(scenario), fill=hex_to_rgb(TOKENS["ink"]), size=17, anchor="rm")
        hf_vals = plot_df[(plot_df["scenario"] == scenario) & (plot_df["frequency"] == "HF_9MHz")]["ocean_margin_db"].to_numpy(dtype=float)
        vhf_vals = plot_df[(plot_df["scenario"] == scenario) & (plot_df["frequency"] == "VHF_60MHz")]["ocean_margin_db"].to_numpy(dtype=float)
        draw_boxplot(draw, hf_vals, center - 24, left, right, x_min, x_max, COLORS["blue_light"], COLORS["blue"], height=28)
        draw_boxplot(draw, vhf_vals, center + 24, left, right, x_min, x_max, COLORS["orange_light"], COLORS["orange"], height=28)
    draw_legend(draw, [("HF 9 MHz", COLORS["blue"]), ("VHF 60 MHz", COLORS["orange"])], 1120, 730)
    return save_image(img, "frequency_ocean_margin.png")


def build_report(
    baseline: pd.DataFrame,
    results: pd.DataFrame,
    summary: pd.DataFrame,
    sensitivity: pd.DataFrame,
    crossref: pd.DataFrame,
    chart_paths: dict[str, Path],
) -> None:
    def img(path: Path, alt: str) -> str:
        rel = path.relative_to(ROOT).as_posix()
        return f'<img src="{html.escape(rel)}" alt="{html.escape(alt)}" />'

    hf_complex = summary[(summary.scenario == "complex_dirty_ice") & (summary.frequency == "HF_9MHz")].iloc[0]
    hf_clean = summary[(summary.scenario == "clean_ice_control") & (summary.frequency == "HF_9MHz")].iloc[0]
    hf_salt = summary[(summary.scenario == "salt_reference") & (summary.frequency == "HF_9MHz")].iloc[0]
    hf_lens = summary[(summary.scenario == "briny_lens_masking") & (summary.frequency == "HF_9MHz")].iloc[0]

    source_rows = "\n".join(
        f"<tr><td>{html.escape(r.source)}</td><td>{html.escape(r.paper_or_source_claim)}</td><td>{html.escape(r.simulation_choice)}</td><td>{html.escape(r.status)}</td><td><a href=\"{html.escape(r.link)}\">link</a></td></tr>"
        for r in crossref.itertuples()
    )

    summary_rows = "\n".join(
        f"<tr><td>{html.escape(r.scenario)}</td><td>{html.escape(r.frequency)}</td><td>{r.clear_ocean_pct:.1f}%</td><td>{(r.ambiguous_pct + r.false_stronger_pct + r.hidden_false_visible_pct):.1f}%</td><td>{r.weak_no_deep_detection_pct:.1f}%</td><td>{r.median_ocean_margin_db:.1f}</td><td>{r.max_abs_inferred_depth_error_m/1000:.2f}</td></tr>"
        for r in summary.itertuples()
    )

    html_text = f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>Dirty Ice Radar Simulation for Europa</title>
  <style>
    body {{ margin: 0; background: #FCFCFD; color: #1F2430; font-family: Aptos, Inter, Segoe UI, Arial, sans-serif; line-height: 1.5; }}
    main {{ max-width: 1120px; margin: 0 auto; padding: 36px 28px 56px; }}
    h1 {{ font-size: 30px; margin: 0 0 12px; }}
    h2 {{ font-size: 21px; margin: 34px 0 10px; border-top: 1px solid #E6E8F0; padding-top: 24px; }}
    h3 {{ font-size: 16px; margin: 22px 0 8px; }}
    p, li {{ font-size: 15px; }}
    .summary {{ background: #EAF1FE; border-left: 5px solid #5477C4; padding: 14px 18px; border-radius: 6px; }}
    .metric-grid {{ display: grid; grid-template-columns: repeat(4, minmax(0, 1fr)); gap: 12px; margin: 18px 0 8px; }}
    .metric {{ background: #FFFFFF; border: 1px solid #E6E8F0; border-radius: 6px; padding: 14px; }}
    .metric strong {{ display: block; font-size: 24px; color: #2E4780; }}
    .metric span {{ color: #6F768A; font-size: 13px; }}
    figure {{ margin: 22px 0 26px; background: #FFFFFF; border: 1px solid #E6E8F0; border-radius: 6px; padding: 14px; }}
    figure img {{ max-width: 100%; height: auto; display: block; }}
    figcaption {{ color: #6F768A; font-size: 13px; margin-top: 8px; }}
    table {{ width: 100%; border-collapse: collapse; font-size: 13px; margin: 16px 0 22px; background: #FFFFFF; }}
    th, td {{ border: 1px solid #E6E8F0; padding: 8px 9px; vertical-align: top; }}
    th {{ background: #F4F5F7; text-align: left; }}
    code {{ font-family: Consolas, monospace; background: #F4F5F7; padding: 2px 4px; border-radius: 4px; }}
    .note {{ color: #464C55; background: #FFF4C2; border-left: 5px solid #B8A037; padding: 12px 16px; border-radius: 6px; }}
    @media (max-width: 800px) {{ .metric-grid {{ grid-template-columns: 1fr 1fr; }} main {{ padding: 24px 16px; }} }}
  </style>
</head>
<body>
<main>
  <h1>Dirty Ice Radar Simulation for Europa</h1>
  <section class="summary">
    <strong>Technical summary.</strong>
    The simulation supports the research gap you picked: meter-scale salt layers alone are not the most interesting missing piece, but patchy mixed dirty ice and briny lenses can make the deep radar interpretation less reliable. In the HF 9 MHz model, the complex dirty-ice case creates {hf_complex.ambiguous_pct + hf_complex.false_stronger_pct + hf_complex.hidden_false_visible_pct:.1f}% deep false-boundary risk and {hf_complex.weak_no_deep_detection_pct:.1f}% weak/no deep-ocean detection, while the clean-ice and simple salt-reference cases keep the HF ocean boundary clear.
  </section>

  <div class="metric-grid">
    <div class="metric"><strong>{hf_complex.ambiguous_pct + hf_complex.false_stronger_pct + hf_complex.hidden_false_visible_pct:.1f}%</strong><span>HF complex dirty-ice deep false-boundary risk</span></div>
    <div class="metric"><strong>{hf_complex.weak_no_deep_detection_pct:.1f}%</strong><span>HF complex dirty-ice weak/no deep detection</span></div>
    <div class="metric"><strong>{hf_complex.max_abs_inferred_depth_error_m/1000:.2f} km</strong><span>Max false-depth error if strongest echo is trusted</span></div>
    <div class="metric"><strong>{baseline.ocean_depth_m.mean()/1000:.1f} km</strong><span>Mean simulated ice-ocean boundary depth</span></div>
  </div>

  <h2>Key findings with visual evidence</h2>
  <p><strong>The best problem case is combined dirty ice, not simple salt alone.</strong> The REASON paper already treats salt layers as possible target interfaces, including meter-scale layers with permittivity around 5. This simulation matches that reference case, then adds the under-studied part: multiple dirty material types and stacked thin interfaces below radar resolution.</p>
  <figure>{img(chart_paths["outcomes"], "Scenario outcome risk chart")}<figcaption>Chart 1. Outcome risk is the share of points where deep false or ambiguous returns compete with the real ocean boundary, or where the ocean is hidden while a deep false layer remains visible.</figcaption></figure>

  <p><strong>False dirty-layer echoes can compete with the ocean return in the complex case.</strong> The track plot compares the true ocean-boundary echo with the strongest false dirty-layer echo. Areas where the false echo approaches or exceeds the ocean echo are where interpretation becomes risky.</p>
  <figure>{img(chart_paths["echo_track"], "Complex dirty ice HF echo track")}<figcaption>Chart 2. HF 9 MHz deep-sounding case. The threshold is the simplified detection cutoff; shaded bands mark deep false-boundary risk, not every shallow internal reflector.</figcaption></figure>

  <p><strong>When a false reflector wins, the depth mistake can become kilometer-scale.</strong> This does not prove that Europa will produce those exact false layers; it shows why realistic dirty-mixture distributions are a strong student simulation target.</p>
  <figure>{img(chart_paths["depth_error"], "Depth error distribution")}<figcaption>Chart 3. Error is computed only from this simplified strongest-return interpretation rule.</figcaption></figure>

  <p><strong>Dirty-layer thickness and loss both matter.</strong> Thin or weakly lossy layers are less dangerous; thick and high-loss layers can hide the ocean or make false interfaces dominate. This is the simulation space worth exploring next.</p>
  <figure>{img(chart_paths["sensitivity"], "Dirty layer sensitivity heatmap")}<figcaption>Chart 4. The sensitivity grid uses one deep dirty layer near 70% of the modeled ocean depth and shows the share of sampled points with ambiguity or false-boundary risk.</figcaption></figure>

  <p><strong>HF and VHF should not be interpreted the same way.</strong> The VHF band is higher-resolution but loses deep ocean detectability first; the HF band is the relevant deep-sounding band for the ice-ocean boundary.</p>
  <figure>{img(chart_paths["frequency_margin"], "Frequency ocean margin comparison")}<figcaption>Chart 5. Positive margin means the ocean boundary is above the model detection threshold.</figcaption></figure>

  <h2>Scope, data, and metric definitions</h2>
  <p>The baseline geometry comes from your current workbook model, reimplemented in Python because the latest workbook formulas do not store reliable cached values for direct reading. The track has 241 points from -60 km to 60 km and keeps your generated topography, shallow layer, briny lens, and ice-ocean boundary logic.</p>
  <ul>
    <li><strong>Ocean echo dB:</strong> relative radar power from the true ice-ocean boundary after two-way attenuation.</li>
    <li><strong>False echo dB:</strong> strongest simulated non-ocean return from dirty layers, including coherent summing of interfaces inside a radar-resolution window.</li>
    <li><strong>Ambiguous / false-risk:</strong> a false echo deep enough to confuse the bottom/ocean boundary is within 3 dB of the ocean echo, stronger than it, or visible while the ocean is hidden.</li>
    <li><strong>Detection threshold:</strong> {DETECTION_THRESHOLD_DB:.0f} dB, matching the simplified threshold style already used in the workbook.</li>
  </ul>

  <h2>Model specification and assumptions</h2>
  <p>The model is a first-order radar interpretation stress test, not a full REASON processing pipeline. It combines interface reflectivity from dielectric contrast, one-way attenuation by material path length, and a simple coherent thin-layer interference approximation. It uses the REASON bands of 9 MHz and 60 MHz.</p>
  <table>
    <thead><tr><th>Scenario</th><th>Frequency</th><th>Clear ocean</th><th>Ambiguous / false risk</th><th>Weak/no deep detection</th><th>Median ocean margin (dB)</th><th>Max inferred depth error (km)</th></tr></thead>
    <tbody>{summary_rows}</tbody>
  </table>

  <h2>Cross-reference against research papers</h2>
  <p>The simulation is intentionally conservative about claims. It uses paper-backed anchors for frequencies, salt-layer permittivity, thin-layer interference, and briny/mushy high-loss behavior, then marks non-salt dirty materials as an uncertainty bracket.</p>
  <table>
    <thead><tr><th>Source</th><th>What the source supports</th><th>How the simulation uses it</th><th>Status</th><th>Link</th></tr></thead>
    <tbody>{source_rows}</tbody>
  </table>

  <h2>Limitations, uncertainty, and robustness checks</h2>
  <p class="note"><strong>Main limitation:</strong> non-salt dirty ice is not well constrained at Europa temperatures and REASON frequencies. The model should be treated as a sensitivity tool: it tells you which combinations of layer thickness, dielectric contrast, and attenuation become dangerous, not that those exact layers exist.</p>
  <ul>
    <li>The attenuation values are scenario ranges chosen to bracket paper-backed behavior; they are not lab-measured constants for every material.</li>
    <li>The thin-layer interference model is simplified and does not replace full-wave electromagnetic modeling.</li>
    <li>The simulation assumes nadir-style sounding and does not include full spacecraft SAR focusing, clutter migration, plasma distortion, or thermal gradient physics.</li>
    <li>The strongest-return interpretation rule is deliberately simple; real mission interpretation will use multi-instrument and multi-pass context.</li>
  </ul>

  <h2>Recommended next steps</h2>
  <ol>
    <li>Turn the complex dirty-ice scenario into a parameter sweep: layer count, layer spacing, dielectric constant, and attenuation.</li>
    <li>Add a material library for specific candidates: NaCl-rich ice, MgSO4 hydrates, sulfuric acid hydrate, dust/organics, and briny/mushy inclusions.</li>
    <li>Compare HF and VHF separately: HF for deep ocean detection, VHF for shallow false-reflector discrimination.</li>
    <li>Use the claim: <em>mixed dirty ice could create ambiguous radar signatures that hide or mimic the true ice-ocean boundary under some layer configurations.</em></li>
  </ol>

  <h2>Further questions</h2>
  <p>The next research question should be: which realistic material mixture produces the highest false-reflector risk while still being plausible for Europa? That will need more lab-property values than the public mission summary gives.</p>
</main>
</body>
</html>
"""
    REPORT_HTML.write_text(html_text, encoding="utf-8")


def main():
    setup_style()
    p = load_inputs()
    baseline = compute_baseline(p)
    results = run_simulation(baseline)
    summary = summarize(results)
    sensitivity = sensitivity_grid(baseline)
    crossref = paper_cross_reference()

    baseline.to_csv(ROOT / "baseline_from_current_model.csv", index=False)
    results.to_csv(RESULTS_CSV, index=False)
    summary.to_csv(SUMMARY_CSV, index=False)
    sensitivity.to_csv(SENSITIVITY_CSV, index=False)
    crossref.to_csv(CROSSREF_CSV, index=False)

    chart_paths = {
        "outcomes": plot_outcomes(summary),
        "echo_track": plot_echo_track(results),
        "depth_error": plot_depth_error(results),
        "sensitivity": plot_sensitivity(sensitivity),
        "frequency_margin": plot_frequency_margin(results),
    }
    build_report(baseline, results, summary, sensitivity, crossref, chart_paths)

    metadata = {
        "workbook_used": str(WORKBOOK),
        "outputs": {
            "report_html": str(REPORT_HTML),
            "results_csv": str(RESULTS_CSV),
            "summary_csv": str(SUMMARY_CSV),
            "sensitivity_csv": str(SENSITIVITY_CSV),
            "cross_reference_csv": str(CROSSREF_CSV),
            "charts": {k: str(v) for k, v in chart_paths.items()},
        },
        "key_assumptions": {
            "frequencies_mhz": {k: v["mhz"] for k, v in FREQUENCIES.items()},
            "detection_threshold_db": DETECTION_THRESHOLD_DB,
            "ice_eps": ICE_EPS,
            "ocean_eps": OCEAN_EPS,
            "system_offset_db": SYSTEM_OFFSET_DB,
        },
    }
    RUN_METADATA.write_text(json.dumps(metadata, indent=2), encoding="utf-8")
    print(json.dumps(metadata, indent=2))


if __name__ == "__main__":
    main()
