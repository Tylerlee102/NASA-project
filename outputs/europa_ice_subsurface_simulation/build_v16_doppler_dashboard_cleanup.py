from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
IN_FILE = BASE_DIR / "v15.xlsx"
OUT_FILE = BASE_DIR / "v16.xlsx"

DASH_SHEET = "Doppler_Depth_Inversion"
DATA_SHEET = "Doppler_Depth_Data"
FORMULA_SHEET = "Subsurface_Formulas"

DATA_START_ROW = 2
SOURCE_START_ROW = 2
SOURCE_ROWS = 241
DATA_END_ROW = DATA_START_ROW + SOURCE_ROWS - 1


def style_cell(cell, fill=None, font=None, align=None, border=None):
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    if align is not None:
        cell.alignment = align
    if border is not None:
        cell.border = border


def add_series(chart, ws, col, title, color, width=18000, dash=None):
    x_values = Reference(ws, min_col=1, min_row=DATA_START_ROW, max_row=DATA_END_ROW)
    y_values = Reference(ws, min_col=col, min_row=DATA_START_ROW, max_row=DATA_END_ROW)
    series = Series(y_values, x_values, title=title)
    series.graphicalProperties.line.solidFill = color
    series.graphicalProperties.line.width = width
    if dash:
        series.graphicalProperties.line.dashStyle = dash
    series.marker.symbol = "none"
    chart.series.append(series)


def base_chart(title, y_title, reverse_y=False):
    chart = ScatterChart()
    chart.scatterStyle = "line"
    chart.title = title
    chart.x_axis.title = "Along-track position x (km)"
    chart.y_axis.title = y_title
    chart.legend.position = "b"
    chart.height = 10.2
    chart.width = 23.8
    chart.style = 2
    chart.display_blanks = "gap"
    if reverse_y:
        chart.y_axis.scaling.orientation = "maxMin"
        chart.y_axis.crosses = "max"
    return chart


def add_dashboard_charts(ws, data_ws):
    angle = base_chart(
        "Doppler-Inverted Look Angle vs Existing Geometry",
        "Look angle (deg)",
    )
    add_series(angle, data_ws, 4, "Doppler angle from VHF shift", "ED7D31", 18000)
    add_series(angle, data_ws, 5, "Existing model geometry angle", "4472C4", 16000, "dash")
    ws.add_chart(angle, "A14")

    depth = base_chart(
        "Raw Slant Depth vs Doppler-Corrected Ocean Depth",
        "Depth below surface (m, positive down)",
        reverse_y=True,
    )
    add_series(depth, data_ws, 8, "True simulated ocean depth", "404040", 16000)
    add_series(depth, data_ws, 12, "Raw slant depth from echo delay", "ED7D31", 16000, "dash")
    add_series(depth, data_ws, 13, "Doppler-corrected actual depth", "70AD47", 18000)
    ws.add_chart(depth, "J14")

    error = base_chart(
        "Depth Error Before and After Angle Correction",
        "Depth error (m)",
    )
    add_series(error, data_ws, 14, "Uncorrected slant-depth error", "ED7D31", 16000)
    add_series(error, data_ws, 15, "Corrected depth error", "70AD47", 18000)
    ws.add_chart(error, "A34")

    layers = base_chart(
        "Corrected Layer Depths From Doppler Angle",
        "Depth below surface (m, positive down)",
        reverse_y=True,
    )
    add_series(layers, data_ws, 19, "Corrected upper-layer depth", "70AD47", 16000)
    add_series(layers, data_ws, 18, "Corrected briny lens depth", "4472C4", 16000)
    add_series(layers, data_ws, 13, "Corrected ocean boundary depth", "ED7D31", 18000)
    ws.add_chart(layers, "J34")


def build_data_sheet(wb, insert_at):
    if DATA_SHEET in wb.sheetnames:
        del wb[DATA_SHEET]
    ws = wb.create_sheet(DATA_SHEET, insert_at)
    ws.sheet_view.showGridLines = False

    navy = "17365D"
    pale_blue = "D9EAF7"
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name="Aptos", size=8, bold=True, color="FFFFFF")
    small_font = Font(name="Aptos", size=8, color="1F1F1F")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    ws.merge_cells("A1:S1")
    ws["A1"] = "Detailed Doppler Depth Data - feeds the Doppler dashboard charts"
    style_cell(
        ws["A1"],
        PatternFill("solid", fgColor=navy),
        Font(name="Aptos Display", size=14, bold=True, color="FFFFFF"),
        Alignment(horizontal="left", vertical="center"),
    )
    ws.row_dimensions[1].height = 24

    header_row = 3
    data_row_start = 4
    global DATA_START_ROW, DATA_END_ROW
    DATA_START_ROW = data_row_start
    DATA_END_ROW = DATA_START_ROW + SOURCE_ROWS - 1

    headers = [
        "x_km",
        "vhf_doppler_hz",
        "radial_velocity_m_s",
        "doppler_angle_deg",
        "geometry_look_angle_deg",
        "true_upper_depth_m",
        "true_lens_depth_m",
        "true_ocean_depth_m",
        "upper_slant_delay_us",
        "lens_slant_delay_us",
        "ocean_slant_delay_us",
        "raw_ocean_slant_depth_m",
        "doppler_corrected_ocean_depth_m",
        "uncorrected_ocean_error_m",
        "corrected_ocean_error_m",
        "correction_removed_m",
        "raw_lens_slant_depth_m",
        "doppler_corrected_lens_depth_m",
        "doppler_corrected_upper_depth_m",
    ]
    for c, header in enumerate(headers, 1):
        cell = ws.cell(header_row, c, header)
        style_cell(cell, PatternFill("solid", fgColor=navy), header_font, center, border)

    for offset in range(SOURCE_ROWS):
        r = DATA_START_ROW + offset
        src = SOURCE_START_ROW + offset
        formulas = [
            f"=Subsurface_Live_Data!A{src}",
            f"=Model_Data!P{src}",
            f"=ABS(B{r})*Inputs!$C$13/2",
            f"=DEGREES(ASIN(MIN(1,C{r}/(Inputs!$C$11*1000))))",
            f"=Model_Data!I{src}",
            f"=Subsurface_Live_Data!F{src}",
            f"=Subsurface_Live_Data!I{src}",
            f"=Subsurface_Live_Data!K{src}",
            f"=Subsurface_Live_Data!M{src}/COS(RADIANS(D{r}))",
            f"=Subsurface_Live_Data!N{src}/COS(RADIANS(D{r}))",
            f"=Subsurface_Live_Data!O{src}/COS(RADIANS(D{r}))",
            f"=Inputs!$C$21*K{r}/(2*Inputs!$C$12*1000000)",
            f"=L{r}*COS(RADIANS(D{r}))",
            f"=L{r}-H{r}",
            f"=M{r}-H{r}",
            f"=N{r}-O{r}",
            f"=Inputs!$C$21*J{r}/(2*Inputs!$C$12*1000000)",
            f"=Q{r}*COS(RADIANS(D{r}))",
            f"=Inputs!$C$21*I{r}/(2*Inputs!$C$12*1000000)*COS(RADIANS(D{r}))",
        ]
        for c, formula in enumerate(formulas, 1):
            cell = ws.cell(r, c, formula)
            style_cell(cell, None, small_font, right, border)

    widths = [10, 13, 15, 14, 16, 15, 15, 15, 16, 16, 16, 18, 21, 18, 18, 17, 17, 21, 21]
    for c, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = width

    formats = {
        "A": "0.0",
        "B": "0.000",
        "C": "0.000",
        "D": "0.000",
        "E": "0.000",
        "F": "0.0",
        "G": "0.0",
        "H": "0.0",
        "I": "0.000",
        "J": "0.000",
        "K": "0.000",
        "L": "0.0",
        "M": "0.0",
        "N": "0.000",
        "O": "0.000000",
        "P": "0.000",
        "Q": "0.0",
        "R": "0.0",
        "S": "0.0",
    }
    for col, fmt in formats.items():
        for row in range(DATA_START_ROW, DATA_END_ROW + 1):
            ws[f"{col}{row}"].number_format = fmt

    ws.freeze_panes = f"A{DATA_START_ROW}"
    ws.auto_filter.ref = f"A{header_row}:S{DATA_END_ROW}"
    return ws


def build_dashboard_sheet(wb, insert_at, data_ws):
    if DASH_SHEET in wb.sheetnames:
        del wb[DASH_SHEET]
    ws = wb.create_sheet(DASH_SHEET, insert_at)
    ws.sheet_view.showGridLines = False

    navy = "17365D"
    blue = "1F4E79"
    pale_blue = "D9EAF7"
    pale_green = "E2F0D9"
    pale_orange = "FCE4D6"
    gray = "F2F2F2"
    white = "FFFFFF"
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title_font = Font(name="Aptos Display", size=18, bold=True, color=white)
    sub_font = Font(name="Aptos", size=10, color=white)
    section_font = Font(name="Aptos", size=10, bold=True, color=white)
    header_font = Font(name="Aptos", size=9, bold=True, color="1F1F1F")
    body_font = Font(name="Aptos", size=9, color="1F1F1F")
    value_font = Font(name="Aptos Display", size=14, bold=True, color="17365D")
    small_font = Font(name="Aptos", size=8, color="404040")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    ws.merge_cells("A1:S1")
    ws["A1"] = "Doppler-Angle Depth Inversion"
    style_cell(ws["A1"], PatternFill("solid", fgColor=navy), title_font, left)
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:S3")
    ws["A2"] = (
        "Main view only: charts plus key live results. The detailed row-by-row numbers that feed "
        "the graphs were moved to the Doppler_Depth_Data tab."
    )
    style_cell(ws["A2"], PatternFill("solid", fgColor="244062"), sub_font, left)

    sections = [
        ("A5:D5", "Main result"),
        ("F5:I5", "Live inputs"),
        ("K5:N5", "Sample depth point"),
        ("P5:S5", "Simplified method"),
    ]
    for rng, label in sections:
        ws.merge_cells(rng)
        cell = ws[rng.split(":")[0]]
        cell.value = label
        style_cell(cell, PatternFill("solid", fgColor=blue), section_font, left, border)

    result_rows = [
        ("Mean raw slant error", f"=SUMPRODUCT(ABS({DATA_SHEET}!N{DATA_START_ROW}:N{DATA_END_ROW}))/ROWS({DATA_SHEET}!N{DATA_START_ROW}:N{DATA_END_ROW})", "m"),
        ("Mean corrected error", f"=SUMPRODUCT(ABS({DATA_SHEET}!O{DATA_START_ROW}:O{DATA_END_ROW}))/ROWS({DATA_SHEET}!O{DATA_START_ROW}:O{DATA_END_ROW})", "m"),
        ("Max Doppler angle", f"=MAX({DATA_SHEET}!D{DATA_START_ROW}:D{DATA_END_ROW})", "deg"),
        ("Depth correction status", f'=IF(SUMPRODUCT(ABS({DATA_SHEET}!O{DATA_START_ROW}:O{DATA_END_ROW}))/ROWS({DATA_SHEET}!O{DATA_START_ROW}:O{DATA_END_ROW})<0.000001,"PASS","CHECK")', ""),
    ]
    for r, row in enumerate(result_rows, 6):
        for c, value in enumerate(row, 1):
            cell = ws.cell(r, c, value)
            fill = PatternFill("solid", fgColor=pale_green) if c == 1 else PatternFill("solid", fgColor=gray)
            font = value_font if c == 2 else body_font
            style_cell(cell, fill, font, right if c == 2 else left, border)

    input_rows = [
        ("VHF wavelength", "=Inputs!$C$13", "m"),
        ("Spacecraft speed", "=Inputs!$C$11*1000", "m/s"),
        ("Ice refractive index", "=Inputs!$C$12", "n"),
        ("Speed of light", "=Inputs!$C$21", "m/s"),
    ]
    for r, row in enumerate(input_rows, 6):
        for c, value in enumerate(row, 6):
            cell = ws.cell(r, c, value)
            fill = PatternFill("solid", fgColor=pale_blue) if c == 6 else PatternFill("solid", fgColor=gray)
            style_cell(cell, fill, body_font, right if c == 7 else left, border)

    sample_rows = [
        ("x position", f"={DATA_SHEET}!A{DATA_START_ROW}", "km"),
        ("True ocean depth", f"={DATA_SHEET}!H{DATA_START_ROW}", "m"),
        ("Raw slant depth", f"={DATA_SHEET}!L{DATA_START_ROW}", "m"),
        ("Corrected ocean depth", f"={DATA_SHEET}!M{DATA_START_ROW}", "m"),
    ]
    for r, row in enumerate(sample_rows, 6):
        for c, value in enumerate(row, 11):
            cell = ws.cell(r, c, value)
            fill = PatternFill("solid", fgColor=pale_orange) if c == 11 else PatternFill("solid", fgColor=gray)
            style_cell(cell, fill, body_font, right if c == 12 else left, border)

    method_rows = [
        ("1", "Doppler shift gives line-of-sight velocity"),
        ("2", "Velocity gives look angle"),
        ("3", "Echo delay gives slant depth"),
        ("4", "Angle correction gives actual depth"),
    ]
    for r, row in enumerate(method_rows, 6):
        ws.cell(r, 16, row[0])
        ws.cell(r, 17, row[1])
        ws.merge_cells(start_row=r, start_column=17, end_row=r, end_column=19)
        style_cell(ws.cell(r, 16), PatternFill("solid", fgColor=pale_green), body_font, center, border)
        style_cell(ws.cell(r, 17), PatternFill("solid", fgColor=gray), small_font, left, border)
        for c in [18, 19]:
            style_cell(ws.cell(r, c), PatternFill("solid", fgColor=gray), small_font, left, border)

    for cell in ["B6", "B7", "B8", "G6", "G7", "G8", "G9", "L6", "L7", "L8", "L9"]:
        ws[cell].number_format = "0.000"
    ws["B9"].number_format = "@"

    ws["A11"] = "NASA context source"
    ws["B11"] = "https://science.nasa.gov/mission/europa-clipper/spacecraft-instruments/"
    ws.merge_cells("B11:S11")
    style_cell(ws["A11"], PatternFill("solid", fgColor=pale_blue), header_font, left, border)
    style_cell(ws["B11"], None, small_font, left, border)

    for c in range(1, 20):
        ws.column_dimensions[get_column_letter(c)].width = 13
    for row in range(1, 56):
        ws.row_dimensions[row].height = 18

    add_dashboard_charts(ws, data_ws)
    return ws


def update_formula_notes(wb):
    if FORMULA_SHEET not in wb.sheetnames:
        return
    ws = wb[FORMULA_SHEET]
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                cell.value = cell.value.replace("Doppler_Depth_Inversion!B:B", "Doppler_Depth_Data!B:B")
                cell.value = cell.value.replace("Doppler_Depth_Inversion!C:C", "Doppler_Depth_Data!C:C")


def main():
    wb = load_workbook(IN_FILE)
    if DASH_SHEET in wb.sheetnames:
        dashboard_index = wb.sheetnames.index(DASH_SHEET)
    else:
        dashboard_index = wb.sheetnames.index("Subsurface_Dashboard") + 1

    if DATA_SHEET in wb.sheetnames:
        del wb[DATA_SHEET]
    if DASH_SHEET in wb.sheetnames:
        del wb[DASH_SHEET]

    data_ws = build_data_sheet(wb, dashboard_index)
    dash_ws = build_dashboard_sheet(wb, dashboard_index, data_ws)
    # Ensure the visible dashboard is directly after Subsurface_Dashboard, and data sits after it.
    wb._sheets.remove(data_ws)
    wb._sheets.insert(dashboard_index + 1, data_ws)

    update_formula_notes(wb)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(OUT_FILE)
    print(f"created {OUT_FILE}")
    print(f"dashboard={dash_ws.title}")
    print(f"data={data_ws.title}")


if __name__ == "__main__":
    main()
