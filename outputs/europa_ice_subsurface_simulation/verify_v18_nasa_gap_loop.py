from pathlib import Path
from zipfile import ZipFile
import re

from openpyxl import load_workbook


PATH = Path(__file__).resolve().parent / "v18.xlsx"
ERROR_TOKENS = ("#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A")
REQUIRED_SHEETS = ["Research_Gaps", "NASA_Coverage_Matrix", "Gap_Test_Topography"]


def main():
    wb = load_workbook(PATH, data_only=False, read_only=False)
    order = wb.sheetnames[:12]
    chart_counts = {ws.title: len(ws._charts) for ws in wb.worksheets if len(ws._charts)}
    missing = [sheet for sheet in REQUIRED_SHEETS if sheet not in wb.sheetnames]

    text_by_sheet = {}
    url_count = 0
    formula_error_hits = []
    for ws in wb.worksheets:
        values = []
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if value is not None:
                    values.append(str(value))
                if isinstance(value, str):
                    if value.startswith("https://"):
                        url_count += 1
                    for token in ERROR_TOKENS:
                        if token in value:
                            formula_error_hits.append((ws.title, cell.coordinate, token))
        text_by_sheet[ws.title] = "\n".join(values)

    required_phrases = {
        "NASA_Coverage_Matrix": [
            "Best project angle",
            "Doppler/frequency-shift measurements",
            "What not to claim",
        ],
        "Gap_Test_Topography": [
            "Live Gap Test",
            "PROBLEM CASE FOUND",
            "Depth Error if Look Angle Is Biased",
        ],
    }
    phrase_hits = {
        sheet: {phrase: phrase in text_by_sheet.get(sheet, "") for phrase in phrases}
        for sheet, phrases in required_phrases.items()
    }

    chart_error_hits = []
    new_chart_infos = []
    with ZipFile(PATH) as zf:
        for name in zf.namelist():
            if name.startswith("xl/charts/chart") and name.endswith(".xml"):
                xml = zf.read(name).decode("utf-8", errors="replace")
                for token in ERROR_TOKENS:
                    if token in xml:
                        chart_error_hits.append((name, token))
                title = ""
                m = re.search(r"<a:t>(.*?)</a:t>", xml)
                if m:
                    title = m.group(1)
                if title in [
                    "Depth Error if Look Angle Is Biased",
                    "Generated Topography Slope Used in Risk Test",
                ]:
                    refs = re.findall(r"<(?:\w+:)?f>(.*?)</(?:\w+:)?f>", xml)
                    new_chart_infos.append((title, refs))

    test = wb["Gap_Test_Topography"]
    formulas = {
        "B7_angle_bias": test["B7"].value,
        "G9_max_biased_error": test["G9"].value,
        "G12_verdict": test["G12"].value,
        "J18_biased_depth_error": test["J18"].value,
        "K18_risk_flag": test["K18"].value,
    }
    wb.close()

    print(f"path={PATH}")
    print(f"sheet_order_first_12={order}")
    print(f"missing_required_sheets={missing}")
    print(f"chart_counts={chart_counts}")
    print(f"url_count={url_count}")
    print(f"formula_error_token_count={len(formula_error_hits)}")
    print(f"chart_error_token_count={len(chart_error_hits)}")
    print(f"phrase_hits={phrase_hits}")
    print(f"key_formulas={formulas}")
    for title, refs in new_chart_infos:
        print(f"new_chart={title} refs={refs}")


if __name__ == "__main__":
    main()
