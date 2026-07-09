from pathlib import Path
from zipfile import ZipFile
import re

from openpyxl import load_workbook


PATH = Path(__file__).resolve().parent / "v15.xlsx"
ERROR_TOKENS = ("#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A")


def visible_error_count():
    wb = load_workbook(PATH, data_only=True, read_only=False)
    errors = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("#"):
                    errors.append((ws.title, cell.coordinate, value))
    wb.close()
    return errors


def formula_error_tokens():
    wb = load_workbook(PATH, data_only=False, read_only=False)
    hits = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str):
                    for token in ERROR_TOKENS:
                        if token in value:
                            hits.append((ws.title, cell.coordinate, token, value[:120]))
    doppler = wb["Doppler_Depth_Inversion"]
    sheet_order = wb.sheetnames[:8]
    chart_count = len(doppler._charts)
    chart_series_counts = [len(chart.series) for chart in doppler._charts]
    formulas = {
        "D24": doppler["D24"].value,
        "L24": doppler["L24"].value,
        "M24": doppler["M24"].value,
        "O24": doppler["O24"].value,
        "L11": doppler["L11"].value,
    }
    wb.close()
    return hits, sheet_order, chart_count, chart_series_counts, formulas


def chart_xml_scan():
    chart_infos = []
    bad_tokens = []
    new_chart_titles = (
        "Doppler-Inverted Look Angle vs Existing Geometry",
        "Raw Slant Depth vs Doppler-Corrected Ocean Depth",
        "Depth Error Before and After Angle Correction",
        "Corrected Layer Depths From Doppler Angle",
    )
    with ZipFile(PATH) as zf:
        chart_names = sorted(name for name in zf.namelist() if name.startswith("xl/charts/chart") and name.endswith(".xml"))
        for name in chart_names:
            text = zf.read(name).decode("utf-8", errors="replace")
            title_match = re.search(r"<a:t>(.*?)</a:t>", text)
            title = title_match.group(1) if title_match else ""
            series_count = len(re.findall(r"<(?:\w+:)?ser>", text))
            refs = re.findall(r"<(?:\w+:)?f>(.*?)</(?:\w+:)?f>", text)
            if title in new_chart_titles:
                chart_infos.append((name, title, series_count, refs[:8]))
            for token in ERROR_TOKENS:
                if token in text:
                    bad_tokens.append((name, token))
    return chart_infos, bad_tokens


def main():
    visible_errors = visible_error_count()
    formula_hits, sheet_order, chart_count, chart_series_counts, formulas = formula_error_tokens()
    chart_infos, chart_bad_tokens = chart_xml_scan()

    print(f"path={PATH}")
    print(f"sheet_order_first_8={sheet_order}")
    print(f"doppler_chart_count_openpyxl={chart_count}")
    print(f"doppler_chart_series_counts={chart_series_counts}")
    print(f"visible_error_count={len(visible_errors)}")
    print(f"formula_error_token_count={len(formula_hits)}")
    print(f"chart_xml_bad_token_count={len(chart_bad_tokens)}")
    print(f"doppler_formulas={formulas}")
    for info in chart_infos:
        print(f"new_chart_xml={info[0]} title={info[1]} series={info[2]} refs={info[3]}")


if __name__ == "__main__":
    main()
