from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook


PATH = Path(__file__).resolve().parent / "v17.xlsx"
SHEET = "Research_Gaps"
ERROR_TOKENS = ("#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A")


def main():
    wb = load_workbook(PATH, data_only=False, read_only=False)
    order = wb.sheetnames[:10]
    ws = wb[SHEET]
    required_phrases = [
        "Best framing for your project",
        "Topography + Doppler/look-angle depth error",
        "False bottom layer from multiple thin internal layers",
        "Briny/warm lens masks deeper ocean return",
        "Do not claim NASA forgot Doppler",
    ]
    text = "\n".join(str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None)
    phrase_hits = {phrase: (phrase in text) for phrase in required_phrases}
    urls = [cell.value for row in ws.iter_rows() for cell in row if isinstance(cell.value, str) and cell.value.startswith("https://")]

    formula_error_hits = []
    visible_formula_cells = 0
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str):
                    if value.startswith("="):
                        visible_formula_cells += 1
                    for token in ERROR_TOKENS:
                        if token in value:
                            formula_error_hits.append((sheet.title, cell.coordinate, token))

    chart_token_hits = []
    with ZipFile(PATH) as zf:
        for name in zf.namelist():
            if name.startswith("xl/charts/chart") and name.endswith(".xml"):
                chart_xml = zf.read(name).decode("utf-8", errors="replace")
                for token in ERROR_TOKENS:
                    if token in chart_xml:
                        chart_token_hits.append((name, token))

    print(f"path={PATH}")
    print(f"sheet_order_first_10={order}")
    print(f"research_gaps_index={wb.sheetnames.index(SHEET) + 1}")
    print(f"phrase_hits={phrase_hits}")
    print(f"url_count={len(urls)}")
    print(f"formula_error_token_count={len(formula_error_hits)}")
    print(f"chart_error_token_count={len(chart_token_hits)}")
    print(f"formula_cell_count={visible_formula_cells}")
    wb.close()


if __name__ == "__main__":
    main()
