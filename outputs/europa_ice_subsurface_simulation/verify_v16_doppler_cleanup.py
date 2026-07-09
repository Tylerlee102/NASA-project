from pathlib import Path
from zipfile import ZipFile
import re

from openpyxl import load_workbook


PATH = Path(__file__).resolve().parent / "v16.xlsx"
DASH_SHEET = "Doppler_Depth_Inversion"
DATA_SHEET = "Doppler_Depth_Data"
ERROR_TOKENS = ("#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A")


def workbook_scan():
    wb = load_workbook(PATH, data_only=False, read_only=False)
    order = wb.sheetnames[:9]
    dash = wb[DASH_SHEET]
    data = wb[DATA_SHEET]

    formula_error_hits = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str):
                    for token in ERROR_TOKENS:
                        if token in value:
                            formula_error_hits.append((ws.title, cell.coordinate, token))

    dash_non_empty_after_60 = []
    for row in dash.iter_rows(min_row=60):
        for cell in row:
            if cell.value not in (None, ""):
                dash_non_empty_after_60.append(cell.coordinate)

    formulas = {
        "Dashboard B6": dash["B6"].value,
        "Dashboard B7": dash["B7"].value,
        "Dashboard B9": dash["B9"].value,
        "Data D4": data["D4"].value,
        "Data M4": data["M4"].value,
    }
    chart_count = len(dash._charts)
    data_formula_rows = sum(1 for row in data.iter_rows(min_row=4, max_row=244, min_col=1, max_col=19) if any(cell.value for cell in row))
    wb.close()
    return order, formula_error_hits, dash_non_empty_after_60, formulas, chart_count, data_formula_rows


def chart_refs():
    titles = (
        "Doppler-Inverted Look Angle vs Existing Geometry",
        "Raw Slant Depth vs Doppler-Corrected Ocean Depth",
        "Depth Error Before and After Angle Correction",
        "Corrected Layer Depths From Doppler Angle",
    )
    infos = []
    bad_tokens = []
    with ZipFile(PATH) as zf:
        for name in sorted(n for n in zf.namelist() if n.startswith("xl/charts/chart") and n.endswith(".xml")):
            text = zf.read(name).decode("utf-8", errors="replace")
            title_match = re.search(r"<a:t>(.*?)</a:t>", text)
            title = title_match.group(1) if title_match else ""
            if title in titles:
                refs = re.findall(r"<(?:\w+:)?f>(.*?)</(?:\w+:)?f>", text)
                series_count = len(re.findall(r"<(?:\w+:)?ser>", text))
                infos.append((name, title, series_count, refs))
            for token in ERROR_TOKENS:
                if token in text:
                    bad_tokens.append((name, token))
    return infos, bad_tokens


def main():
    order, formula_error_hits, dash_after_60, formulas, chart_count, data_formula_rows = workbook_scan()
    infos, bad_tokens = chart_refs()
    print(f"path={PATH}")
    print(f"sheet_order_first_9={order}")
    print(f"dashboard_chart_count={chart_count}")
    print(f"data_formula_rows={data_formula_rows}")
    print(f"dashboard_non_empty_cells_after_row_60={len(dash_after_60)}")
    print(f"formula_error_token_count={len(formula_error_hits)}")
    print(f"chart_bad_token_count={len(bad_tokens)}")
    print(f"key_formulas={formulas}")
    for info in infos:
        data_ref_count = sum(DATA_SHEET in ref for ref in info[3])
        dash_ref_count = sum(DASH_SHEET in ref for ref in info[3])
        print(f"chart={info[1]} series={info[2]} data_refs={data_ref_count} dash_refs={dash_ref_count}")


if __name__ == "__main__":
    main()
