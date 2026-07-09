from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
IN_FILE = BASE_DIR / "v16.xlsx"
OUT_FILE = BASE_DIR / "v17.xlsx"
SHEET = "Research_Gaps"


def style(cell, fill=None, font=None, alignment=None, border=None):
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border


def write_row(ws, row, values, fill=None, font=None, border=None, align=None):
    for col, value in enumerate(values, 1):
        cell = ws.cell(row, col, value)
        style(cell, fill, font, align, border)


def build_sheet(wb):
    if SHEET in wb.sheetnames:
        del wb[SHEET]

    insert_at = wb.sheetnames.index("Doppler_Depth_Data") + 1
    ws = wb.create_sheet(SHEET, insert_at)
    ws.sheet_view.showGridLines = False

    navy = "17365D"
    blue = "1F4E79"
    pale_blue = "D9EAF7"
    pale_green = "E2F0D9"
    pale_orange = "FCE4D6"
    pale_red = "F4CCCC"
    gray = "F2F2F2"
    white = "FFFFFF"
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title_font = Font(name="Aptos Display", size=18, bold=True, color=white)
    section_font = Font(name="Aptos", size=11, bold=True, color=white)
    header_font = Font(name="Aptos", size=9, bold=True, color="1F1F1F")
    body_font = Font(name="Aptos", size=9, color="1F1F1F")
    bold_font = Font(name="Aptos", size=9, bold=True, color="1F1F1F")
    small_font = Font(name="Aptos", size=8, color="404040")

    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("A1:J1")
    ws["A1"] = "NASA Research Gap Check - Europa Radar Model"
    style(ws["A1"], PatternFill("solid", fgColor=navy), title_font, left)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J4")
    ws["A2"] = (
        "Best framing for your project: NASA is already checking Europa's ice thickness, surface "
        "roughness, topography, radar returns, Doppler/gravity evidence, and cross-instrument data. "
        "Your strongest angle is not 'NASA forgot this.' It is: this workbook stress-tests one "
        "specific interpretation problem where rough ice topography plus Doppler/look-angle correction "
        "can make the bottom layer depth look wrong or create a false subsurface signal."
    )
    style(ws["A2"], PatternFill("solid", fgColor="244062"), Font(name="Aptos", size=10, color=white), left)

    ws.merge_cells("A6:J6")
    ws["A6"] = "Main research question to use"
    style(ws["A6"], PatternFill("solid", fgColor=blue), section_font, left, border)
    main_rows = [
        ["Question", "Can rough/generated Europa ice topography plus Doppler/look-angle correction cause the radar-inferred bottom layer or ocean boundary to be overestimated, underestimated, or confused with a shallow briny layer?"],
        ["Why it matters", "NASA will use radar echoes, surface elevations/roughness, and gravity/Doppler-type measurements together. A small geometry error can turn a slant-path echo into the wrong actual depth."],
        ["What your model can add", "A simple live sensitivity test showing when topography amplitude, local slope, off-nadir geometry, ice refractive index, and briny lens attenuation push the bottom-layer answer outside a safe error range."],
        ["Best claim wording", "This simulation tests a possible radar interpretation risk; it does not claim NASA missed the whole problem or that Europa's ocean is proven."],
    ]
    for r, values in enumerate(main_rows, 7):
        ws.cell(r, 1, values[0])
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=10)
        ws.cell(r, 2, values[1])
        style(ws.cell(r, 1), PatternFill("solid", fgColor=pale_green), bold_font, left, border)
        for c in range(2, 11):
            style(ws.cell(r, c), PatternFill("solid", fgColor=gray), body_font, left, border)

    ws.merge_cells("A13:J13")
    ws["A13"] = "Gap candidates ranked for your workbook"
    style(ws["A13"], PatternFill("solid", fgColor=blue), section_font, left, border)
    headers = [
        "Rank",
        "Possible topic",
        "NASA/source already covers",
        "Possible under-tested angle",
        "Workbook test to add",
        "Best graph",
        "Evidence strength",
        "Difficulty",
        "Use this?",
        "Notes",
    ]
    write_row(ws, 14, headers, PatternFill("solid", fgColor=pale_blue), header_font, border, center)

    rows = [
        [
            1,
            "Topography + Doppler/look-angle depth error",
            "NASA says REASON studies ice structure/thickness plus surface elevations and roughness; NASA also uses frequency shifts for gravity/radio science.",
            "Public summaries do not show a simple combined sensitivity threshold for generated topography + off-nadir angle + radar depth conversion.",
            "Sweep topography amplitude/slope and off-nadir angle; calculate raw slant depth, corrected depth, and remaining bottom-layer error.",
            "Heatmap: local surface slope vs look angle, colored by bottom-depth error.",
            "Strong",
            "Medium",
            "YES - strongest",
            "This directly matches your current model and new Doppler tab.",
        ],
        [
            2,
            "False bottom layer from multiple thin internal layers",
            "REASON will analyze bounced radar signals to see internal features. Mars radar studies show bright reflectors can come from layer interference, not liquid water.",
            "Test whether Europa-like internal layer spacing can combine into a bright return that looks like an ocean boundary.",
            "Add repeated thin layers with small dielectric contrasts; compare true ocean echo vs false bright stacked-layer echo.",
            "False-interface count vs layer spacing and dielectric contrast.",
            "Strong analog",
            "Medium-high",
            "YES",
            "This is a clean 'problem NASA must avoid' angle without claiming they ignored it.",
        ],
        [
            3,
            "Briny/warm lens masks deeper ocean return",
            "NASA is searching for shallow subsurface water and the ice-ocean interface.",
            "A shallow conductive briny lens could absorb or reflect energy enough that the deeper bottom layer becomes weak or invisible.",
            "Sweep lens thickness, salinity/attenuation, and depth; calculate ocean echo margin after passing through lens.",
            "Echo margin vs briny lens strength.",
            "Medium-strong",
            "Medium",
            "YES",
            "Your workbook already has lens and ocean echo margins, so this is easy to extend.",
        ],
        [
            4,
            "Radar depth vs gravity/topography mismatch",
            "Gravity studies use Doppler/range measurements; gravity-topography admittance can complement radar for shell thickness.",
            "Radar sees local structure; gravity/topography can average broad structure. A local anomaly may disagree with global shell estimates.",
            "Compare local radar-inferred shell thickness with smoothed gravity/topography proxy and flag mismatches.",
            "Radar thickness minus gravity-proxy thickness along track.",
            "Medium",
            "Medium-high",
            "Maybe",
            "Good extension after the radar-only model is solid.",
        ],
        [
            5,
            "Plasma/ionosphere delay plus angle correction",
            "NASA uses PIMS to separate plasma distortion from Europa's induced magnetic signal; prior work studies Europa ionosphere effects on radar propagation.",
            "Small path-delay or phase distortion could combine with geometry correction and move a weak bottom echo.",
            "Add a tunable ionosphere/plasma delay term and compare corrected depth with/without it.",
            "Depth bias vs plasma delay and look angle.",
            "Medium",
            "Higher",
            "Later",
            "Useful but harder to explain unless the radar model is already polished.",
        ],
    ]
    for r, values in enumerate(rows, 15):
        write_row(ws, r, values, None, body_font, border, left)
        if values[8].startswith("YES"):
            for c in range(1, 11):
                ws.cell(r, c).fill = PatternFill("solid", fgColor=pale_green)
        elif values[8] == "Maybe":
            for c in range(1, 11):
                ws.cell(r, c).fill = PatternFill("solid", fgColor=pale_orange)

    ws.merge_cells("A22:J22")
    ws["A22"] = "Research loop checklist"
    style(ws["A22"], PatternFill("solid", fgColor=blue), section_font, left, border)
    loop_headers = ["Step", "What to check", "Pass condition"]
    write_row(ws, 23, loop_headers, PatternFill("solid", fgColor=pale_blue), header_font, border, center)
    loop_rows = [
        ["1", "Find a NASA/primary source that says what the mission already measures.", "You can cite the exact instrument or paper."],
        ["2", "Mark whether NASA already clearly covers the same exact issue.", "If yes, reframe as a sensitivity test, not a discovery claim."],
        ["3", "Ask what your workbook can calculate that the source summary does not show.", "It produces a number, threshold, or graph."],
        ["4", "Check whether the result could change interpretation of an ocean, briny lens, or shell thickness.", "It affects a science conclusion or measurement confidence."],
        ["5", "Write a careful claim.", "Use 'possible interpretation risk' or 'stress test,' not 'NASA did not think of it.'"],
    ]
    for r, values in enumerate(loop_rows, 24):
        write_row(ws, r, values, None, body_font, border, left)

    ws.merge_cells("A31:J31")
    ws["A31"] = "Source evidence used"
    style(ws["A31"], PatternFill("solid", fgColor=blue), section_font, left, border)
    source_headers = ["ID", "Source", "What it supports", "Link"]
    write_row(ws, 32, source_headers, PatternFill("solid", fgColor=pale_blue), header_font, border, center)
    source_rows = [
        [
            "S1",
            "NASA Europa Clipper instruments page",
            "REASON probes ice shell, studies structure/thickness, surface elevations, roughness; gravity/radio science analyzes frequency shifts.",
            "https://science.nasa.gov/mission/europa-clipper/spacecraft-instruments/",
        ],
        [
            "S2",
            "NASA Europa Clipper mission page",
            "Mission carries nine instruments plus gravity experiment; instruments operate together on each flyby; main goal is habitability below surface.",
            "https://science.nasa.gov/mission/europa-clipper/",
        ],
        [
            "S3",
            "Verma & Margot 2018, Expected Precision of Europa Clipper Gravity Measurements",
            "Radio science uses two-way Doppler and radar altimeter crossover ranges; results depend on DSN tracking and crossover performance.",
            "https://arxiv.org/abs/1801.08946",
        ],
        [
            "S4",
            "Akiba, Ermakov & Militzer 2021, gravity-topography admittance",
            "Gravity/topography can complement ice-penetrating radar, and simple Airy models may not work across all wavelengths.",
            "https://arxiv.org/abs/2105.02790",
        ],
        [
            "S5",
            "Lalich, Hayes & Poggiali 2021, bright radar reflections without liquid water",
            "Radar simulations can explain bright basal reflectors through multiple-layer interference rather than liquid water; useful analog warning.",
            "https://arxiv.org/abs/2107.03497",
        ],
    ]
    for r, values in enumerate(source_rows, 33):
        write_row(ws, r, values, None, small_font, border, left)

    ws.merge_cells("A40:J40")
    ws["A40"] = "Do not claim"
    style(ws["A40"], PatternFill("solid", fgColor="C00000"), section_font, left, border)
    warning_rows = [
        ["Do not claim NASA forgot Doppler, roughness, topography, or radar clutter.", "The sources show these are already in the mission/instrument plan."],
        ["Do not claim the model proves an ocean exists.", "The workbook is a simulated stress test, not Europa flight data."],
        ["Do not claim one graph is enough.", "The stronger project is a loop: source -> possible risk -> formula -> graph -> threshold -> careful conclusion."],
    ]
    for r, values in enumerate(warning_rows, 41):
        ws.cell(r, 1, values[0])
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=10)
        ws.cell(r, 2, values[1])
        style(ws.cell(r, 1), PatternFill("solid", fgColor=pale_red), bold_font, left, border)
        for c in range(2, 11):
            style(ws.cell(r, c), PatternFill("solid", fgColor=gray), body_font, left, border)

    widths = [8, 25, 38, 38, 36, 32, 16, 15, 16, 36]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in range(1, 45):
        ws.row_dimensions[row].height = 36
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 54
    ws.freeze_panes = "A14"
    return ws


def main():
    wb = load_workbook(IN_FILE)
    build_sheet(wb)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(OUT_FILE)
    print(f"created {OUT_FILE}")


if __name__ == "__main__":
    main()
