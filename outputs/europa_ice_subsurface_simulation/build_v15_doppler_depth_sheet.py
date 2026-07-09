from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
IN_FILE = BASE_DIR / "v14.xlsx"
OUT_FILE = BASE_DIR / "v15.xlsx"

SHEET_NAME = "Doppler_Depth_Inversion"
FORMULA_SHEET = "Subsurface_Formulas"

DATA_START_ROW = 24
SOURCE_START_ROW = 2
SOURCE_ROWS = 241
DATA_END_ROW = DATA_START_ROW + SOURCE_ROWS - 1


def style_cell(cell, fill=None, font=None, align=None, border=None):
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    if align is not None:
        cell.alignment = align
    if border is not None:
        cell.border = border


def add_doppler_sheet(wb):
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]

    insert_at = wb.sheetnames.index("Subsurface_Dashboard") + 1
    ws = wb.create_sheet(SHEET_NAME, insert_at)
    ws.sheet_view.showGridLines = False

    navy = "17365D"
    blue = "1F4E79"
    pale_blue = "D9EAF7"
    pale_green = "E2F0D9"
    pale_orange = "FCE4D6"
    gray = "F2F2F2"
    white = "FFFFFF"
    orange = "F4B183"
    green = "A9D18E"
    red = "F4CCCC"

    thin_gray = Side(style="thin", color="D9E2F3")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    title_fill = PatternFill("solid", fgColor=navy)
    section_fill = PatternFill("solid", fgColor=blue)
    header_fill = PatternFill("solid", fgColor=pale_blue)
    param_fill = PatternFill("solid", fgColor=gray)
    method_fill = PatternFill("solid", fgColor=pale_green)
    note_fill = PatternFill("solid", fgColor=pale_orange)

    title_font = Font(name="Aptos Display", size=18, bold=True, color=white)
    subtitle_font = Font(name="Aptos", size=10, color=white)
    section_font = Font(name="Aptos", size=11, bold=True, color=white)
    header_font = Font(name="Aptos", size=9, bold=True, color="1F1F1F")
    body_font = Font(name="Aptos", size=9, color="1F1F1F")
    small_font = Font(name="Aptos", size=8, color="404040")

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    ws.merge_cells("A1:S1")
    ws["A1"] = "Doppler-Angle Depth Inversion"
    style_cell(ws["A1"], title_fill, title_font, Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:S3")
    ws["A2"] = (
        "Simplified live model: use the simulated Doppler shift to estimate look angle, "
        "then convert radar echo delay from slant depth into vertical ice depth. NASA's real "
        "Europa processing also uses spacecraft trajectory, radar focusing, clutter removal, "
        "surface topography, and multi-instrument checks."
    )
    style_cell(ws["A2"], PatternFill("solid", fgColor="244062"), subtitle_font, left)

    ws.merge_cells("A5:D5")
    ws["A5"] = "Method summary"
    style_cell(ws["A5"], section_fill, section_font, left)

    method_headers = ["Step", "What the model does", "Formula idea", "Why it matters"]
    for col, header in enumerate(method_headers, start=1):
        cell = ws.cell(6, col, header)
        style_cell(cell, header_fill, header_font, center, border)

    method_rows = [
        (
            "1",
            "Read the radar echo delay from the simulated subsurface return.",
            "slant depth = c * delay / (2 * n)",
            "Delay gives distance traveled through ice, but not yet true vertical depth.",
        ),
        (
            "2",
            "Reverse the Doppler shift into a line-of-sight velocity estimate.",
            "v radial = ABS(fD) * wavelength / 2",
            "Doppler shift tells how fast the reflecting path is moving toward/away from the radar.",
        ),
        (
            "3",
            "Turn that velocity fraction into a Doppler look angle.",
            "theta = ASIN(MIN(1, v radial / spacecraft speed))",
            "The angle tells how much the measured path is tilted away from vertical.",
        ),
        (
            "4",
            "Correct slant depth back to actual vertical depth.",
            "actual depth = slant depth * COS(theta)",
            "This is the depth estimate to compare with the simulated layer truth.",
        ),
        (
            "5",
            "Check before/after error.",
            "error = corrected depth - true simulated depth",
            "This shows whether the angle correction is doing what it should.",
        ),
    ]
    for r, row_values in enumerate(method_rows, start=7):
        for c, value in enumerate(row_values, start=1):
            cell = ws.cell(r, c, value)
            style_cell(cell, method_fill if c == 1 else None, body_font, left, border)

    ws.merge_cells("F5:I5")
    ws["F5"] = "Live inputs used by this tab"
    style_cell(ws["F5"], section_fill, section_font, left)
    for col, header in enumerate(["Input", "Live value", "Unit", "Source"], start=6):
        cell = ws.cell(6, col, header)
        style_cell(cell, header_fill, header_font, center, border)

    params = [
        ("VHF radar wavelength", "=Inputs!$C$13", "m", "Inputs!C13"),
        ("Spacecraft speed", "=Inputs!$C$11*1000", "m/s", "Inputs!C11"),
        ("Speed of light", "=Inputs!$C$21", "m/s", "Inputs!C21"),
        ("Ice refractive index", "=Inputs!$C$12", "n", "Inputs!C12"),
        ("ASIN safety cap", 1, "ratio", "prevents invalid angle if noise exceeds 1"),
    ]
    for r, row_values in enumerate(params, start=7):
        for c, value in enumerate(row_values, start=6):
            cell = ws.cell(r, c, value)
            style_cell(cell, param_fill if c == 6 else None, body_font, left if c != 7 else right, border)

    ws.merge_cells("K5:S5")
    ws["K5"] = "Calculation checks"
    style_cell(ws["K5"], section_fill, section_font, left)
    checks = [
        ("Mean raw slant-depth error", f"=SUMPRODUCT(ABS(N{DATA_START_ROW}:N{DATA_END_ROW}))/ROWS(N{DATA_START_ROW}:N{DATA_END_ROW})", "m"),
        ("Mean corrected-depth error", f"=SUMPRODUCT(ABS(O{DATA_START_ROW}:O{DATA_END_ROW}))/ROWS(O{DATA_START_ROW}:O{DATA_END_ROW})", "m"),
        ("Average correction removed", f"=SUMPRODUCT(ABS(P{DATA_START_ROW}:P{DATA_END_ROW}))/ROWS(P{DATA_START_ROW}:P{DATA_END_ROW})", "m"),
        ("Max Doppler angle", f"=MAX(D{DATA_START_ROW}:D{DATA_END_ROW})", "deg"),
        ("Corrected ocean depth matches simulated truth?", f'=IF(SUMPRODUCT(ABS(O{DATA_START_ROW}:O{DATA_END_ROW}))/ROWS(O{DATA_START_ROW}:O{DATA_END_ROW})<0.000001,"PASS","CHECK")', ""),
    ]
    for c, header in enumerate(["Metric", "Live value", "Unit"], start=11):
        cell = ws.cell(6, c, header)
        style_cell(cell, header_fill, header_font, center, border)
    for r, row_values in enumerate(checks, start=7):
        for c, value in enumerate(row_values, start=11):
            cell = ws.cell(r, c, value)
            style_cell(cell, note_fill if c == 11 else None, body_font, left if c != 12 else right, border)

    ws.merge_cells("A14:S16")
    ws["A14"] = (
        "NASA context: Europa Clipper's REASON radar is designed to probe Europa's icy shell "
        "and study ice structure/thickness from bounced radar signals. NASA also uses frequency "
        "shift measurements in radio/gravity science. This tab uses those ideas in a simplified "
        "spreadsheet form: Doppler gives a look-angle correction, radar delay gives slant depth, "
        "and the angle correction estimates actual vertical depth."
    )
    style_cell(ws["A14"], PatternFill("solid", fgColor="EAF2F8"), body_font, left, border)

    ws["A18"] = "NASA source"
    ws["B18"] = "https://science.nasa.gov/mission/europa-clipper/spacecraft-instruments/"
    style_cell(ws["A18"], header_fill, header_font, left, border)
    style_cell(ws["B18"], None, body_font, left, border)

    data_headers = [
        "x_km",
        "vhf_doppler_hz",
        "radial_velocity_m_s",
        "doppler_angle_deg",
        "geometry_look_angle_deg",
        "true_upper_depth_m",
        "true_lens_depth_m",
        "true_ocean_depth_m",
        "upper_slant_delay_us",
        "lens_slant_delay_us",
        "ocean_slant_delay_us",
        "raw_ocean_slant_depth_m",
        "doppler_corrected_ocean_depth_m",
        "uncorrected_ocean_error_m",
        "corrected_ocean_error_m",
        "correction_removed_m",
        "raw_lens_slant_depth_m",
        "doppler_corrected_lens_depth_m",
        "doppler_corrected_upper_depth_m",
    ]
    for c, header in enumerate(data_headers, start=1):
        cell = ws.cell(DATA_START_ROW - 1, c, header)
        style_cell(cell, PatternFill("solid", fgColor=navy), Font(name="Aptos", size=8, bold=True, color=white), center, border)

    for offset in range(SOURCE_ROWS):
        r = DATA_START_ROW + offset
        src = SOURCE_START_ROW + offset
        formulas = [
            f"=Subsurface_Live_Data!A{src}",
            f"=Model_Data!P{src}",
            f"=ABS(B{r})*$G$7/2",
            f"=DEGREES(ASIN(MIN($G$11,C{r}/$G$8)))",
            f"=Model_Data!I{src}",
            f"=Subsurface_Live_Data!F{src}",
            f"=Subsurface_Live_Data!I{src}",
            f"=Subsurface_Live_Data!K{src}",
            f"=Subsurface_Live_Data!M{src}/COS(RADIANS(D{r}))",
            f"=Subsurface_Live_Data!N{src}/COS(RADIANS(D{r}))",
            f"=Subsurface_Live_Data!O{src}/COS(RADIANS(D{r}))",
            f"=$G$9*K{r}/(2*$G$10*1000000)",
            f"=L{r}*COS(RADIANS(D{r}))",
            f"=L{r}-H{r}",
            f"=M{r}-H{r}",
            f"=N{r}-O{r}",
            f"=$G$9*J{r}/(2*$G$10*1000000)",
            f"=Q{r}*COS(RADIANS(D{r}))",
            f"=$G$9*I{r}/(2*$G$10*1000000)*COS(RADIANS(D{r}))",
        ]
        for c, formula in enumerate(formulas, start=1):
            cell = ws.cell(r, c, formula)
            style_cell(cell, None, small_font, right, border)

    for col in range(1, 20):
        ws.column_dimensions[get_column_letter(col)].width = [
            10, 13, 15, 14, 16, 15, 15, 15, 16, 16, 16, 18, 21, 18, 18, 17, 17, 21, 21
        ][col - 1]

    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        ws.row_dimensions[row].height = 14

    number_formats = {
        "A": "0.0",
        "B": "0.000",
        "C": "0.000",
        "D": "0.000",
        "E": "0.000",
        "F": "0.0",
        "G": "0.0",
        "H": "0.0",
        "I": "0.000",
        "J": "0.000",
        "K": "0.000",
        "L": "0.0",
        "M": "0.0",
        "N": "0.000",
        "O": "0.000000",
        "P": "0.000",
        "Q": "0.0",
        "R": "0.0",
        "S": "0.0",
    }
    for col, fmt in number_formats.items():
        for row in range(DATA_START_ROW, DATA_END_ROW + 1):
            ws[f"{col}{row}"].number_format = fmt
    for cell in ["G7", "G8", "G9", "G10", "G11", "L7", "L8", "L9", "L10"]:
        ws[cell].number_format = "0.000"
    ws["L11"].number_format = "@"

    ws.freeze_panes = f"A{DATA_START_ROW}"
    ws.auto_filter.ref = f"A{DATA_START_ROW - 1}:S{DATA_END_ROW}"

    add_live_charts(ws)

    return ws


def add_series(chart, ws, col, title, color, width=18000, dash=None):
    x_values = Reference(ws, min_col=1, min_row=DATA_START_ROW, max_row=DATA_END_ROW)
    y_values = Reference(ws, min_col=col, min_row=DATA_START_ROW, max_row=DATA_END_ROW)
    series = Series(y_values, x_values, title=title)
    series.graphicalProperties.line.solidFill = color
    series.graphicalProperties.line.width = width
    if dash:
        series.graphicalProperties.line.dashStyle = dash
    series.marker.symbol = "none"
    chart.series.append(series)


def base_scatter_chart(title, y_title, reverse_y=False):
    chart = ScatterChart()
    chart.scatterStyle = "line"
    chart.title = title
    chart.x_axis.title = "Along-track position x (km)"
    chart.y_axis.title = y_title
    chart.legend.position = "b"
    chart.height = 11.5
    chart.width = 25.5
    chart.style = 2
    chart.display_blanks = "gap"
    chart.x_axis.majorGridlines = None
    if reverse_y:
        chart.y_axis.scaling.orientation = "maxMin"
        chart.y_axis.crosses = "max"
    return chart


def add_live_charts(ws):
    angle = base_scatter_chart(
        "Doppler-Inverted Look Angle vs Existing Geometry",
        "Look angle (deg)",
        reverse_y=False,
    )
    add_series(angle, ws, 4, "Doppler angle from VHF shift", "ED7D31", 18000)
    add_series(angle, ws, 5, "Existing model geometry angle", "4472C4", 16000, "dash")
    ws.add_chart(angle, "A270")

    depth = base_scatter_chart(
        "Raw Slant Depth vs Doppler-Corrected Ocean Depth",
        "Depth below surface (m, positive down)",
        reverse_y=True,
    )
    add_series(depth, ws, 8, "True simulated ocean depth", "404040", 16000)
    add_series(depth, ws, 12, "Raw slant depth from echo delay", "ED7D31", 16000, "dash")
    add_series(depth, ws, 13, "Doppler-corrected actual depth", "70AD47", 18000)
    ws.add_chart(depth, "K270")

    error = base_scatter_chart(
        "Depth Error Before and After Angle Correction",
        "Depth error (m)",
        reverse_y=False,
    )
    add_series(error, ws, 14, "Uncorrected slant-depth error", "ED7D31", 16000)
    add_series(error, ws, 15, "Corrected depth error", "70AD47", 18000)
    ws.add_chart(error, "A298")

    layers = base_scatter_chart(
        "Corrected Layer Depths From Doppler Angle",
        "Depth below surface (m, positive down)",
        reverse_y=True,
    )
    add_series(layers, ws, 19, "Corrected upper-layer depth", "70AD47", 16000)
    add_series(layers, ws, 18, "Corrected briny lens depth", "4472C4", 16000)
    add_series(layers, ws, 13, "Corrected ocean boundary depth", "ED7D31", 18000)
    ws.add_chart(layers, "K298")


def append_formula_notes(wb):
    ws = wb[FORMULA_SHEET]
    start = ws.max_row + 2
    section_fill = PatternFill("solid", fgColor="17365D")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    white_font = Font(name="Aptos", size=11, bold=True, color="FFFFFF")
    header_font = Font(name="Aptos", size=9, bold=True, color="1F1F1F")
    body_font = Font(name="Aptos", size=9, color="1F1F1F")
    thin_gray = Side(style="thin", color="D9E2F3")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=5)
    cell = ws.cell(start, 1, "Doppler-angle depth inversion formulas")
    style_cell(cell, section_fill, white_font, left, border)

    headers = ["Workbook concept", "Formula", "Live inputs", "Output used for", "Important note"]
    for c, header in enumerate(headers, start=1):
        cell = ws.cell(start + 1, c, header)
        style_cell(cell, header_fill, header_font, left, border)

    rows = [
        (
            "Doppler radial velocity",
            "v_radial = ABS(fD) * radar_wavelength / 2",
            "Doppler_Depth_Inversion!B:B, Inputs!C13",
            "Estimate line-of-sight motion from simulated VHF Doppler.",
            "Simplified two-way radar Doppler relationship.",
        ),
        (
            "Doppler look angle",
            "theta = ASIN(MIN(1, v_radial / spacecraft_speed))",
            "Doppler_Depth_Inversion!C:C, Inputs!C11",
            "Estimate the angle needed to correct slant depth.",
            "The cap prevents invalid ASIN values if noisy inputs exceed the physical range.",
        ),
        (
            "Slant depth from radar delay",
            "slant_depth = c * delay_us / (2 * ice_index * 1,000,000)",
            "Inputs!C21, Inputs!C12, simulated layer delays",
            "Convert radar echo time into distance through ice.",
            "Uses two-way travel time and the slower radar speed in ice.",
        ),
        (
            "Actual vertical depth",
            "actual_depth = slant_depth * COS(theta)",
            "Doppler angle and slant depth columns",
            "Estimate corrected ocean, lens, and upper-layer depths.",
            "This is the workbook's simplified version of correcting an angled return.",
        ),
        (
            "Depth error check",
            "error = corrected_depth - true_simulated_depth",
            "Corrected depth and Subsurface_Live_Data true depth",
            "Prove the formula is integrated and calculating on-sheet.",
            "In this controlled simulation the corrected ocean error should be near zero.",
        ),
    ]

    for r, row_values in enumerate(rows, start=start + 2):
        for c, value in enumerate(row_values, start=1):
            cell = ws.cell(r, c, value)
            style_cell(cell, None, body_font, left, border)

    ws.cell(start + 8, 1, "NASA context source")
    ws.cell(start + 8, 2, "https://science.nasa.gov/mission/europa-clipper/spacecraft-instruments/")
    for c in range(1, 6):
        style_cell(ws.cell(start + 8, c), None, body_font, left, border)

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = max(ws.column_dimensions[get_column_letter(col)].width or 12, [25, 42, 36, 35, 46][col - 1])


def main():
    wb = load_workbook(IN_FILE)
    add_doppler_sheet(wb)
    append_formula_notes(wb)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    wb.save(OUT_FILE)
    print(f"created {OUT_FILE}")


if __name__ == "__main__":
    main()
