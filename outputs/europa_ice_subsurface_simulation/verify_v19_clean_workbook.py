from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook


PATH = Path(__file__).resolve().parent / "v19.xlsx"
REMOVED = {"Research_Gaps", "NASA_Coverage_Matrix", "Gap_Test_Topography"}
ERROR_TOKENS = ("#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A")


def main():
    wb = load_workbook(PATH, data_only=False, read_only=False)
    sheets = wb.sheetnames
    still_present = [name for name in REMOVED if name in sheets]
    chart_counts = {ws.title: len(ws._charts) for ws in wb.worksheets if len(ws._charts)}
    formula_errors = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str):
                    for token in ERROR_TOKENS:
                        if token in value:
                            formula_errors.append((ws.title, cell.coordinate, token))
    wb.close()

    chart_errors = []
    with ZipFile(PATH) as zf:
        for name in zf.namelist():
            if name.startswith("xl/charts/chart") and name.endswith(".xml"):
                text = zf.read(name).decode("utf-8", errors="replace")
                for token in ERROR_TOKENS:
                    if token in text:
                        chart_errors.append((name, token))

    print(f"path={PATH}")
    print(f"research_tabs_still_present={still_present}")
    print(f"sheet_order_first_12={sheets[:12]}")
    print(f"chart_counts={chart_counts}")
    print(f"formula_error_token_count={len(formula_errors)}")
    print(f"chart_error_token_count={len(chart_errors)}")


if __name__ == "__main__":
    main()
