from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
IN_FILE = BASE_DIR / "v17.xlsx"
OUT_FILE = BASE_DIR / "v18.xlsx"

COVERAGE = "NASA_Coverage_Matrix"
TEST = "Gap_Test_Topography"
DATA = "Doppler_Depth_Data"

DATA_START = 18
DATA_ROWS = 241
DATA_END = DATA_START + DATA_ROWS - 1


def style(cell, fill=None, font=None, align=None, border=None):
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    if align is not None:
        cell.alignment = align
    if border is not None:
        cell.border = border


def row_style(ws, row, start_col, end_col, fill=None, font=None, align=None, border=None):
    for col in range(start_col, end_col + 1):
        style(ws.cell(row, col), fill, font, align, border)


def write_row(ws, row, values, fill=None, font=None, align=None, border=None):
    for col, value in enumerate(values, 1):
        cell = ws.cell(row, col, value)
        style(cell, fill, font, align, border)


def common_styles():
    thin = Side(style="thin", color="D9E2F3")
    return {
        "navy": "17365D",
        "blue": "1F4E79",
        "pale_blue": "D9EAF7",
        "pale_green": "E2F0D9",
        "pale_orange": "FCE4D6",
        "pale_red": "F4CCCC",
        "gray": "F2F2F2",
        "white": "FFFFFF",
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "title_font": Font(name="Aptos Display", size=18, bold=True, color="FFFFFF"),
        "section_font": Font(name="Aptos", size=11, bold=True, color="FFFFFF"),
        "header_font": Font(name="Aptos", size=9, bold=True, color="1F1F1F"),
        "body_font": Font(name="Aptos", size=9, color="1F1F1F"),
        "small_font": Font(name="Aptos", size=8, color="404040"),
        "value_font": Font(name="Aptos Display", size=13, bold=True, color="17365D"),
        "left": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "center": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "right": Alignment(horizontal="right", vertical="center", wrap_text=True),
    }


def build_coverage_sheet(wb, insert_at):
    if COVERAGE in wb.sheetnames:
        del wb[COVERAGE]
    ws = wb.create_sheet(COVERAGE, insert_at)
    ws.sheet_view.showGridLines = False
    s = common_styles()

    ws.merge_cells("A1:H1")
    ws["A1"] = "NASA Coverage Matrix - What Is Already Covered vs Your Testable Gap"
    style(ws["A1"], PatternFill("solid", fgColor=s["navy"]), s["title_font"], s["left"])

    ws.merge_cells("A2:H4")
    ws["A2"] = (
        "Purpose: loop through each possible project idea and separate two things: "
        "what NASA/mission papers already say they will measure, and what your workbook can still "
        "test as a focused risk or sensitivity case. This keeps your project original without "
        "overclaiming that NASA ignored a major known issue."
    )
    style(ws["A2"], PatternFill("solid", fgColor="244062"), Font(name="Aptos", size=10, color=s["white"]), s["left"])

    ws.merge_cells("A6:H6")
    ws["A6"] = "Coverage decision table"
    style(ws["A6"], PatternFill("solid", fgColor=s["blue"]), s["section_font"], s["left"], s["border"])

    headers = [
        "Topic / risk",
        "Already covered by NASA or papers?",
        "Evidence",
        "What still looks open enough for your model",
        "Workbook test",
        "Safe wording",
        "Priority",
        "Decision",
    ]
    write_row(ws, 7, headers, PatternFill("solid", fgColor=s["pale_blue"]), s["header_font"], s["center"], s["border"])

    rows = [
        [
            "Ice shell thickness and ocean detection",
            "Yes, core mission objective",
            "NASA states Europa Clipper will determine icy shell thickness, study ocean interaction, and use REASON to probe ice structure/thickness.",
            "Not a new gap by itself. Use it as the science target your error test affects.",
            "Compare true simulated ocean boundary with corrected and biased inferred boundary.",
            "My model tests how measurement geometry can bias shell/ocean depth interpretation.",
            "High",
            "Use as context, not novelty",
        ],
        [
            "Surface elevation and roughness",
            "Yes",
            "NASA states REASON will study surface elevations and roughness.",
            "Public mission summaries do not provide a simple student-level threshold for when rough topography creates unacceptable depth error.",
            "Compute local slope from generated topography and compare with depth error under angle-bias cases.",
            "The model stress-tests when rough terrain makes radar depth correction risky.",
            "Very high",
            "Best project angle",
        ],
        [
            "Doppler/frequency-shift measurements",
            "Yes for gravity/radio science",
            "NASA says gravity/radio science analyzes frequency shifts in spacecraft signals; Verma & Margot simulate two-way Doppler and crossover ranges.",
            "Radar depth correction is not the same as gravity science, but Doppler/look-angle geometry is useful for this simplified model.",
            "Use Doppler-inverted look angle, then test depth error if the angle is wrong by 1, 3, or 5 degrees.",
            "This is a simplified Doppler/look-angle sensitivity model, not NASA's full gravity pipeline.",
            "Very high",
            "Use carefully",
        ],
        [
            "False bright bottom return from internal layering",
            "Partly, by analogy and radar-processing literature",
            "A Mars radar paper shows bright reflectors can be explained by multiple-layer interference without liquid water.",
            "Europa radar interpretation could face similar ambiguity, especially with layered ice or briny features.",
            "Add a future stacked-layer echo simulation that compares false interface strength with the ocean return.",
            "The model tests a possible false-positive radar interpretation case.",
            "High",
            "Second best extension",
        ],
        [
            "Briny lens masking the deeper ocean return",
            "Partly",
            "NASA says REASON will search for shallow subsurface water and ice-ocean structure.",
            "The exact strength where a shallow briny lens hides a deeper boundary is a good sensitivity question.",
            "Sweep lens attenuation and depth, then flag when ocean echo margin falls below zero.",
            "The model estimates when a shallow conductive layer could hide a deeper return.",
            "High",
            "Good extension",
        ],
        [
            "Gravity-topography mismatch",
            "Partly",
            "Akiba et al. say gravity-topography admittance can complement radar and that simple Airy models can fail across wavelengths.",
            "Good for later: compare local radar thickness with a smoothed gravity/topography proxy.",
            "Add a local-vs-smoothed shell thickness mismatch graph.",
            "The model checks whether local radar inferences disagree with broad gravity/topography behavior.",
            "Medium",
            "Later",
        ],
        [
            "Plasma / ionosphere delay in radar or magnetic sounding",
            "Yes broadly",
            "NASA says PIMS separates plasma distortions from induced magnetic signal; REASON also has ionosphere/plume context.",
            "Could be important, but harder to model accurately without more physics.",
            "Add a tunable extra delay term only after radar geometry tests are finished.",
            "The model can include a first-order delay stress test, but it is not the cleanest main project.",
            "Medium",
            "Later",
        ],
    ]

    for r, values in enumerate(rows, 8):
        write_row(ws, r, values, None, s["body_font"], s["left"], s["border"])
        decision = values[-1]
        if "Best" in decision:
            fill = PatternFill("solid", fgColor=s["pale_green"])
        elif decision in ("Use carefully", "Second best extension", "Good extension"):
            fill = PatternFill("solid", fgColor="EBF1DE")
        elif decision == "Later":
            fill = PatternFill("solid", fgColor=s["pale_orange"])
        else:
            fill = PatternFill("solid", fgColor=s["gray"])
        row_style(ws, r, 1, 8, fill, s["body_font"], s["left"], s["border"])

    ws.merge_cells("A17:H17")
    ws["A17"] = "Bottom line"
    style(ws["A17"], PatternFill("solid", fgColor=s["blue"]), s["section_font"], s["left"], s["border"])
    bottom = [
        ["Strongest project claim", "A topography + Doppler/look-angle correction error can become a radar interpretation risk for Europa's bottom layer depth."],
        ["What not to claim", "Do not say NASA did not think about topography, roughness, Doppler, or radar clutter. Say your model quantifies a simplified risk threshold."],
        ["Next proof step", "Use the live stress test sheet to show how many along-track points exceed 100 m and 500 m depth error when the look angle is biased."],
    ]
    for r, values in enumerate(bottom, 18):
        ws.cell(r, 1, values[0])
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        ws.cell(r, 2, values[1])
        style(ws.cell(r, 1), PatternFill("solid", fgColor=s["pale_green"]), s["header_font"], s["left"], s["border"])
        for c in range(2, 9):
            style(ws.cell(r, c), PatternFill("solid", fgColor=s["gray"]), s["body_font"], s["left"], s["border"])

    ws.merge_cells("A23:H23")
    ws["A23"] = "Primary sources"
    style(ws["A23"], PatternFill("solid", fgColor=s["blue"]), s["section_font"], s["left"], s["border"])
    write_row(ws, 24, ["ID", "Source", "Use", "Link"], PatternFill("solid", fgColor=s["pale_blue"]), s["header_font"], s["center"], s["border"])
    sources = [
        ["S1", "NASA Europa Clipper instruments page", "REASON, roughness, surface elevation, gravity/radio frequency shifts", "https://science.nasa.gov/mission/europa-clipper/spacecraft-instruments/"],
        ["S2", "NASA Europa Clipper mission page", "Mission goals, instruments operating together, 49 flybys", "https://science.nasa.gov/mission/europa-clipper/"],
        ["S3", "Verma & Margot 2018", "Two-way Doppler and radar altimeter crossover simulations for Europa Clipper gravity", "https://arxiv.org/abs/1801.08946"],
        ["S4", "Akiba, Ermakov & Militzer 2021", "Gravity-topography admittance complements radar for icy shell thickness", "https://arxiv.org/abs/2105.02790"],
        ["S5", "Lalich, Hayes & Poggiali 2021", "Radar-bright reflections can be caused by layer interference without liquid water", "https://arxiv.org/abs/2107.03497"],
    ]
    for r, values in enumerate(sources, 25):
        write_row(ws, r, values, None, s["small_font"], s["left"], s["border"])

    widths = [28, 25, 42, 44, 34, 36, 12, 18]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in range(1, 32):
        ws.row_dimensions[row].height = 40
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 54
    ws.freeze_panes = "A8"


def add_series(chart, ws, x_col, y_col, title, color, width=17000, dash=None):
    x = Reference(ws, min_col=x_col, min_row=DATA_START, max_row=DATA_END)
    y = Reference(ws, min_col=y_col, min_row=DATA_START, max_row=DATA_END)
    series = Series(y, x, title=title)
    series.graphicalProperties.line.solidFill = color
    series.graphicalProperties.line.width = width
    if dash:
        series.graphicalProperties.line.dashStyle = dash
    series.marker.symbol = "none"
    chart.series.append(series)


def chart(title, x_title, y_title):
    c = ScatterChart()
    c.scatterStyle = "line"
    c.title = title
    c.x_axis.title = x_title
    c.y_axis.title = y_title
    c.legend.position = "b"
    c.height = 10.5
    c.width = 25
    c.style = 2
    c.display_blanks = "gap"
    return c


def build_test_sheet(wb, insert_at):
    if TEST in wb.sheetnames:
        del wb[TEST]
    ws = wb.create_sheet(TEST, insert_at)
    ws.sheet_view.showGridLines = False
    s = common_styles()

    ws.merge_cells("A1:N1")
    ws["A1"] = "Live Gap Test - Topography + Doppler Look-Angle Depth Risk"
    style(ws["A1"], PatternFill("solid", fgColor=s["navy"]), s["title_font"], s["left"])

    ws.merge_cells("A2:N4")
    ws["A2"] = (
        "This is the numerical loop for the strongest research topic. It asks: if the look angle "
        "used to convert radar slant depth into actual depth is off by a small amount, how badly can "
        "the inferred bottom layer depth shift? The data are live from the generated topography and "
        "Doppler depth sheets."
    )
    style(ws["A2"], PatternFill("solid", fgColor="244062"), Font(name="Aptos", size=10, color=s["white"]), s["left"])

    ws.merge_cells("A6:D6")
    ws["A6"] = "Inputs for stress test"
    style(ws["A6"], PatternFill("solid", fgColor=s["blue"]), s["section_font"], s["left"], s["border"])
    input_rows = [
        ["Angle bias tested", 3, "deg", "Change this to test 1, 3, 5 degrees"],
        ["Medium-risk threshold", 100, "m", "Depth error above this is worth flagging"],
        ["High-risk threshold", 500, "m", "Depth error above this is a serious interpretation risk"],
    ]
    for r, values in enumerate(input_rows, 7):
        write_row(ws, r, values, PatternFill("solid", fgColor=s["gray"]), s["body_font"], s["left"], s["border"])
        style(ws.cell(r, 2), PatternFill("solid", fgColor=s["pale_blue"]), s["value_font"], s["right"], s["border"])

    ws.merge_cells("F6:N6")
    ws["F6"] = "Live results"
    style(ws["F6"], PatternFill("solid", fgColor=s["blue"]), s["section_font"], s["left"], s["border"])
    result_rows = [
        ["Max local surface slope", f"=MAX(MAX(D{DATA_START}:D{DATA_END}),-MIN(D{DATA_START}:D{DATA_END}))", "deg"],
        ["Max raw slant-depth error", f"=MAX(MAX(H{DATA_START}:H{DATA_END}),-MIN(H{DATA_START}:H{DATA_END}))", "m"],
        ["Max error with angle bias", f"=MAX(MAX(J{DATA_START}:J{DATA_END}),-MIN(J{DATA_START}:J{DATA_END}))", "m"],
        ["Rows above medium threshold", f'=COUNTIF(K{DATA_START}:K{DATA_END},"MEDIUM")+COUNTIF(K{DATA_START}:K{DATA_END},"HIGH")', "points"],
        ["Rows above high threshold", f'=COUNTIF(K{DATA_START}:K{DATA_END},"HIGH")', "points"],
        ["Project-risk verdict", f'=IF(COUNTIF(K{DATA_START}:K{DATA_END},"HIGH")>0,"PROBLEM CASE FOUND",IF(COUNTIF(K{DATA_START}:K{DATA_END},"MEDIUM")>0,"SENSITIVITY RISK","LOW IN THIS RUN"))', ""],
    ]
    for r, values in enumerate(result_rows, 7):
        ws.cell(r, 6, values[0])
        ws.cell(r, 7, values[1])
        ws.cell(r, 8, values[2])
        ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=14)
        if r == 12:
            ws.cell(r, 9, "If this says PROBLEM CASE FOUND, your topic has a clear testable problem case.")
        style(ws.cell(r, 6), PatternFill("solid", fgColor=s["pale_green"]), s["body_font"], s["left"], s["border"])
        style(ws.cell(r, 7), PatternFill("solid", fgColor=s["gray"]), s["value_font"], s["right"], s["border"])
        style(ws.cell(r, 8), PatternFill("solid", fgColor=s["gray"]), s["body_font"], s["left"], s["border"])
        for c in range(9, 15):
            style(ws.cell(r, c), PatternFill("solid", fgColor=s["gray"]), s["small_font"], s["left"], s["border"])

    ws.merge_cells("A15:N15")
    ws["A15"] = "Stress-test data"
    style(ws["A15"], PatternFill("solid", fgColor=s["blue"]), s["section_font"], s["left"], s["border"])
    headers = [
        "x_km",
        "surface_height_m",
        "doppler_angle_deg",
        "local_surface_slope_deg",
        "true_ocean_depth_m",
        "raw_slant_depth_m",
        "corrected_depth_m",
        "raw_error_m",
        "biased_corrected_depth_m",
        "biased_depth_error_m",
        "risk_flag",
        "topography_angle_risk_index",
        "claim_note",
        "source_data",
    ]
    write_row(ws, DATA_START - 1, headers, PatternFill("solid", fgColor=s["navy"]), Font(name="Aptos", size=8, bold=True, color=s["white"]), s["center"], s["border"])

    # Table source rows: Doppler_Depth_Data rows 4:244 and Subsurface_Live_Data rows 2:242.
    for i in range(DATA_ROWS):
        r = DATA_START + i
        drow = 4 + i
        srow = 2 + i
        if i == 0:
            slope_formula = f"=DEGREES(ATAN((Subsurface_Live_Data!B{srow+1}-Subsurface_Live_Data!B{srow})/((Subsurface_Live_Data!A{srow+1}-Subsurface_Live_Data!A{srow})*1000)))"
        elif i == DATA_ROWS - 1:
            slope_formula = f"=DEGREES(ATAN((Subsurface_Live_Data!B{srow}-Subsurface_Live_Data!B{srow-1})/((Subsurface_Live_Data!A{srow}-Subsurface_Live_Data!A{srow-1})*1000)))"
        else:
            slope_formula = f"=DEGREES(ATAN((Subsurface_Live_Data!B{srow+1}-Subsurface_Live_Data!B{srow-1})/((Subsurface_Live_Data!A{srow+1}-Subsurface_Live_Data!A{srow-1})*1000)))"
        formulas = [
            f"={DATA}!A{drow}",
            f"=Subsurface_Live_Data!B{srow}",
            f"={DATA}!D{drow}",
            slope_formula,
            f"={DATA}!H{drow}",
            f"={DATA}!L{drow}",
            f"={DATA}!M{drow}",
            f"={DATA}!N{drow}",
            f"=F{r}*COS(RADIANS(C{r}+$B$7))",
            f"=I{r}-E{r}",
            f'=IF(ABS(J{r})>=$B$9,"HIGH",IF(ABS(J{r})>=$B$8,"MEDIUM","LOW"))',
            f"=ABS(D{r})*ABS(J{r})/100",
            f'=IF(K{r}="HIGH","Depth interpretation could be seriously biased",IF(K{r}="MEDIUM","Worth checking in presentation","Low in this scenario"))',
            "Live: Subsurface_Live_Data + Doppler_Depth_Data",
        ]
        write_row(ws, r, formulas, None, s["small_font"], s["right"], s["border"])
        style(ws.cell(r, 13), None, s["small_font"], s["left"], s["border"])
        style(ws.cell(r, 14), None, s["small_font"], s["left"], s["border"])

    formats = {
        "A": "0.0",
        "B": "0.0",
        "C": "0.000",
        "D": "0.000",
        "E": "0.0",
        "F": "0.0",
        "G": "0.0",
        "H": "0.000",
        "I": "0.0",
        "J": "0.000",
        "L": "0.000",
    }
    for col, fmt in formats.items():
        for row in range(DATA_START, DATA_END + 1):
            ws[f"{col}{row}"].number_format = fmt
    for cell in ["B7", "B8", "B9", "G7", "G8", "G9", "G10", "G11"]:
        ws[cell].number_format = "0.000"

    # Charts
    chart_row = DATA_END + 5

    c1 = chart("Depth Error if Look Angle Is Biased", "Along-track position x (km)", "Depth error (m)")
    add_series(c1, ws, 1, 8, "Raw slant-depth error", "ED7D31", 16000, "dash")
    add_series(c1, ws, 1, 10, "Error after assumed angle bias", "C00000", 18000)
    ws.add_chart(c1, f"A{chart_row}")

    c2 = chart("Generated Topography Slope Used in Risk Test", "Along-track position x (km)", "Local slope (deg)")
    add_series(c2, ws, 1, 4, "Local surface slope", "4472C4", 17000)
    ws.add_chart(c2, f"H{chart_row}")

    widths = [9, 15, 15, 18, 17, 17, 17, 14, 22, 18, 12, 22, 34, 30]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in range(1, DATA_END + 1):
        ws.row_dimensions[row].height = 18
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 54
    ws.freeze_panes = f"A{DATA_START}"
    ws.auto_filter.ref = f"A{DATA_START-1}:N{DATA_END}"


def main():
    wb = load_workbook(IN_FILE)
    # Keep research outputs together right after Research_Gaps.
    base_insert = wb.sheetnames.index("Research_Gaps") + 1
    for sheet in [COVERAGE, TEST]:
        if sheet in wb.sheetnames:
            del wb[sheet]
    build_coverage_sheet(wb, base_insert)
    build_test_sheet(wb, base_insert + 1)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(OUT_FILE)
    print(f"created {OUT_FILE}")


if __name__ == "__main__":
    main()
