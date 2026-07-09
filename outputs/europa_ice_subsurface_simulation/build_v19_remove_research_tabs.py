from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
IN_FILE = BASE_DIR / "v18.xlsx"
OUT_FILE = BASE_DIR / "v19.xlsx"

REMOVE_SHEETS = [
    "Research_Gaps",
    "NASA_Coverage_Matrix",
    "Gap_Test_Topography",
]


def main():
    wb = load_workbook(IN_FILE)
    removed = []
    for name in REMOVE_SHEETS:
        if name in wb.sheetnames:
            del wb[name]
            removed.append(name)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(OUT_FILE)
    print(f"created={OUT_FILE}")
    print(f"removed={removed}")
    print(f"sheets={wb.sheetnames[:12]}")


if __name__ == "__main__":
    main()
