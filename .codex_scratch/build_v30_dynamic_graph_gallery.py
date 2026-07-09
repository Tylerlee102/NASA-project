from copy import deepcopy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path(r"C:\Users\tyboy\OneDrive\Documents\Nasa project")
SRC = ROOT / "outputs" / "europa_ice_subsurface_simulation" / "v29_dynamic_signal_tests.xlsx"
OUT = ROOT / "outputs" / "europa_ice_subsurface_simulation" / "v30_all_dynamic_graphs.xlsx"


SECTIONS = [
    ("Front Story Graphs", "00_MAIN_GRAPHS"),
    ("Signal Tests", "Signal_Tests_Dynamic"),
    ("V3 Confidence Dashboard", "V3_Dashboard"),
    ("V2 Paper-Calibrated Dashboard", "V2_Dashboard"),
    ("Original Subsurface Dashboard", "Dashboard"),
    ("Subsurface Model Dashboard", "Subsurface_Dashboard"),
    ("Doppler Depth Inversion", "Doppler_Depth_Inversion"),
]


def chart_title(chart, fallback):
    try:
        return chart.title.tx.rich.p[0].r[0].t
    except Exception:
        return fallback


def main():
    wb = load_workbook(SRC)

    if "01_GRAPH_GALLERY" in wb.sheetnames:
        old = wb["01_GRAPH_GALLERY"]
        idx = wb.sheetnames.index("01_GRAPH_GALLERY")
        del wb["01_GRAPH_GALLERY"]
    else:
        idx = 1

    ws = wb.create_sheet("01_GRAPH_GALLERY", idx)
    ws.sheet_view.showGridLines = False

    navy = "17365D"
    blue = "1F4E78"
    light = "D9EAF7"
    white = "FFFFFF"
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "Dynamic Graph Gallery"
    ws["A1"].font = Font(name="Arial", bold=True, color=white, size=16)
    ws["A1"].fill = PatternFill("solid", fgColor=navy)
    ws.merge_cells("A1:R1")

    ws["A2"] = (
        "This sheet now uses native Excel chart objects copied from the workbook sections, not pasted pictures. "
        "The charts stay linked to their original sheet ranges."
    )
    ws["A2"].font = Font(name="Arial", size=10)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:R3")

    row = 5
    copied = 0
    for section_title, source_sheet in SECTIONS:
        if source_sheet not in wb.sheetnames:
            continue
        src_ws = wb[source_sheet]
        charts = list(getattr(src_ws, "_charts", []))
        if not charts:
            continue

        ws.cell(row, 1, section_title)
        ws.cell(row, 1).font = Font(name="Arial", bold=True, color=white, size=12)
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=blue)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=18)
        row += 1

        ws.cell(row, 1, f"Native charts copied from {source_sheet}.")
        ws.cell(row, 1).font = Font(name="Arial", italic=True, color="666666", size=9)
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=light)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=18)
        row += 1

        for i in range(0, len(charts), 2):
            left_chart = deepcopy(charts[i])
            left_chart.width = 14
            left_chart.height = 8
            left_title = chart_title(left_chart, f"{section_title} chart {i + 1}")
            ws.cell(row, 1, left_title)
            ws.cell(row, 1).font = Font(name="Arial", bold=True, size=10)
            ws.cell(row, 1).border = border
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            ws.add_chart(left_chart, f"A{row + 1}")
            copied += 1

            if i + 1 < len(charts):
                right_chart = deepcopy(charts[i + 1])
                right_chart.width = 14
                right_chart.height = 8
                right_title = chart_title(right_chart, f"{section_title} chart {i + 2}")
                ws.cell(row, 10, right_title)
                ws.cell(row, 10).font = Font(name="Arial", bold=True, size=10)
                ws.cell(row, 10).border = border
                ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=17)
                ws.add_chart(right_chart, f"J{row + 1}")
                copied += 1

            row += 18

        row += 1

    # If a section had an odd/even row mismatch, keep page readable.
    for col in range(1, 19):
        ws.column_dimensions[chr(64 + col) if col <= 26 else "A"].width = 12
    for col in ["A", "J"]:
        ws.column_dimensions[col].width = 18
    for r in range(1, row + 5):
        ws.row_dimensions[r].height = 18

    # Remove any accidental images and force recalculation on open.
    ws._images = []
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    wb.save(OUT)
    print(OUT)
    print(f"copied_charts={copied}")


if __name__ == "__main__":
    main()
