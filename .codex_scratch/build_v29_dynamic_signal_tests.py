from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(r"C:\Users\tyboy\OneDrive\Documents\Nasa project")
SRC = ROOT / "outputs" / "europa_ice_subsurface_simulation" / "v26_paper_calibrated_dirty_ice_v3_streamlined_plus.xlsx"
OUT = ROOT / "outputs" / "europa_ice_subsurface_simulation" / "v29_dynamic_signal_tests.xlsx"


def style_range(ws, min_row, max_row, min_col, max_col, *, header_row=None):
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if header_row is not None and cell.row == header_row:
                cell.font = Font(name="Arial", bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
            else:
                cell.font = Font(name="Arial", size=10)


def add_table(ws, display_name, ref):
    tab = Table(displayName=display_name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)


def add_line_chart(ws, title, y_title, category_col, data_cols, header_row, first_data_row, last_data_row, anchor):
    chart = LineChart()
    chart.title = title
    chart.y_axis.title = y_title
    chart.x_axis.title = "x position (km)"
    chart.legend.position = "b"
    chart.width = 15
    chart.height = 7.8

    cats = Reference(ws, min_col=category_col, min_row=first_data_row, max_row=last_data_row)
    for col in data_cols:
        data = Reference(ws, min_col=col, min_row=header_row, max_row=last_data_row)
        chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, anchor)


def main():
    wb = load_workbook(SRC)

    for name in [
        "Signal_Test_Dashboard",
        "Signal_Test_Data",
        "Signal_Pulse_Tests",
        "Signal_Test_Assumptions",
        "Signal_Tests_Combined",
        "Signal_Tests_Dynamic",
    ]:
        if name in wb.sheetnames:
            del wb[name]

    ws = wb.create_sheet("Signal_Tests_Dynamic", 3)
    ws.sheet_view.showGridLines = False

    navy = "17365D"
    blue = "1F4E78"
    light_blue = "D9EAF7"
    white = "FFFFFF"

    ws["A1"] = "Dynamic Signal Tests"
    ws["A1"].font = Font(name="Arial", bold=True, color=white, size=16)
    ws["A1"].fill = PatternFill("solid", fgColor=navy)
    ws.merge_cells("A1:W1")

    ws["A2"] = (
        "Formula-backed tests for geometric spreading, pulse-compression gain, coherent Fresnel-zone sums, "
        "and frequency-dependent surface response. The charts below are native Excel charts connected to the table."
    )
    ws["A2"].font = Font(name="Arial", size=10)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:W3")

    # Explanation / test checklist.
    overview = [
        ["Test", "What changes dynamically", "Formula / method"],
        ["Geometric spreading", "Uses Model_Data ranges from the current workbook.", "P_rel=(R_nadir/R_off)^4; dB=10*LOG10(P_rel)"],
        ["Pulse compression", "Change pulse length, bandwidth, or window loss in assumptions.", "10*LOG10(B*T) + window loss"],
        ["Coherent Fresnel sums", "Depends on wavelength, off-nadir range, and along-track spacing.", "N≈2*r_F/spacing; gain=10*LOG10(N)"],
        ["Frequency response", "Change slope or reference frequency in assumptions.", "slope*LOG(freq/ref,2)"],
    ]
    for r, row in enumerate(overview, start=5):
        for c, value in enumerate(row, start=1):
            ws.cell(r, c, value)
    style_range(ws, 5, 9, 1, 3, header_row=5)
    for r in range(6, 10):
        ws.cell(r, 1).fill = PatternFill("solid", fgColor=light_blue)
        ws.cell(r, 1).font = Font(name="Arial", bold=True)

    # Editable assumptions. These are intentionally formulas where possible.
    assumptions = [
        ["Parameter", "Value", "Unit", "Notes"],
        ["Ice refractive index", "=Inputs!$C$12", "", "Used for range resolution."],
        ["Along-track spacing", "=ABS((Model_Data!A3-Model_Data!A2)*1000)", "m", "Distance between modeled points."],
        ["Default pulse length", 20, "us", "Edit this to test pulse length."],
        ["Window loss", "=10*LOG10((0.54^2)/0.3974)", "dB", "Hamming-window processing loss estimate."],
        ["Base surface reflectivity", 0, "dB", "Relative constant reflectivity assumption."],
        ["Frequency slope", -2, "dB/octave", "Sensitivity check for frequency-dependent surface response."],
        ["Reference frequency", 9, "MHz", "HF reference frequency."],
        ["HF bandwidth", 1, "MHz", "Simple HF chirp bandwidth."],
        ["VHF bandwidth", 10, "MHz", "Simple VHF chirp bandwidth."],
        ["HF wavelength", "=Inputs!$C$14", "m", "From workbook input."],
        ["VHF wavelength", "=Inputs!$C$13", "m", "From workbook input."],
    ]
    for r, row in enumerate(assumptions, start=11):
        for c, value in enumerate(row, start=1):
            ws.cell(r, c, value)
    style_range(ws, 11, 22, 1, 4, header_row=11)
    add_table(ws, "DynamicSignalAssumptions", "A11:D22")

    # Pulse helper table for native dynamic pulse chart.
    pulse_headers = ["Pulse length us", "HF gain dB", "VHF gain dB", "HF range res ice m", "VHF range res ice m"]
    for c, value in enumerate(pulse_headers, start=6):
        ws.cell(11, c, value)
    for idx, pulse in enumerate([5, 10, 20, 50, 100], start=12):
        ws.cell(idx, 6, pulse)
        ws.cell(idx, 7, f"=10*LOG10($B$19*F{idx})+$B$15")
        ws.cell(idx, 8, f"=10*LOG10($B$20*F{idx})+$B$15")
        ws.cell(idx, 9, "=299792458/(2*$B$12*$B$19*1000000)")
        ws.cell(idx, 10, "=299792458/(2*$B$12*$B$20*1000000)")
    style_range(ws, 11, 16, 6, 10, header_row=11)
    add_table(ws, "DynamicPulseGain", "F11:J16")

    chart = LineChart()
    chart.title = "Pulse compression gain vs pulse length"
    chart.y_axis.title = "dB"
    chart.x_axis.title = "Pulse length (us)"
    chart.legend.position = "b"
    chart.width = 15
    chart.height = 7.8
    chart.add_data(Reference(ws, min_col=7, max_col=8, min_row=11, max_row=16), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=6, min_row=12, max_row=16))
    ws.add_chart(chart, "L11")

    # Formula-backed along-track table. Rows 25:265 map to Model_Data 2:242.
    headers = [
        "x_km",
        "R_nadir_km",
        "R_off_km",
        "Geom rel power flat",
        "Geom dB flat",
        "R_nadir_topo_km",
        "R_off_topo_km",
        "Geom rel power topo",
        "Geom dB topo",
        "HF Fresnel radius m",
        "HF coherent sums",
        "HF coherent gain dB",
        "VHF Fresnel radius m",
        "VHF coherent sums",
        "VHF coherent gain dB",
        "HF pulse gain dB",
        "VHF pulse gain dB",
        "HF freq response dB",
        "VHF freq response dB",
        "HF total dB",
        "VHF total constant dB",
        "VHF total freq-dependent dB",
        "VHF freq response effect dB",
    ]
    header_row = 25
    first_row = 26
    last_row = first_row + 240
    for c, value in enumerate(headers, start=1):
        ws.cell(header_row, c, value)

    for out_row, model_row in enumerate(range(2, 243), start=first_row):
        formulas = [
            f"=Model_Data!A{model_row}",
            f"=Model_Data!C{model_row}",
            f"=Model_Data!D{model_row}",
            f"=(B{out_row}/C{out_row})^4",
            f"=10*LOG10(D{out_row})",
            f"=Model_Data!Z{model_row}",
            f"=Model_Data!AA{model_row}",
            f"=(F{out_row}/G{out_row})^4",
            f"=10*LOG10(H{out_row})",
            f"=SQRT($B$21*G{out_row}*1000/2)",
            f"=MAX(1,2*J{out_row}/$B$13)",
            f"=10*LOG10(K{out_row})",
            f"=SQRT($B$22*G{out_row}*1000/2)",
            f"=MAX(1,2*M{out_row}/$B$13)",
            f"=10*LOG10(N{out_row})",
            f"=10*LOG10($B$19*$B$14)+$B$15",
            f"=10*LOG10($B$20*$B$14)+$B$15",
            f"=$B$17*LOG(9/$B$18,2)",
            f"=$B$17*LOG(60/$B$18,2)",
            f"=I{out_row}+P{out_row}+L{out_row}+$B$16+R{out_row}",
            f"=I{out_row}+Q{out_row}+O{out_row}+$B$16",
            f"=U{out_row}+S{out_row}",
            f"=S{out_row}",
        ]
        for c, formula in enumerate(formulas, start=1):
            ws.cell(out_row, c, formula)

    style_range(ws, header_row, last_row, 1, len(headers), header_row=header_row)
    for row in ws.iter_rows(min_row=first_row, max_row=last_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.number_format = "0.00"
    add_table(ws, "DynamicAlongTrackSignalTests", f"A{header_row}:W{last_row}")

    # Native charts tied directly to formula cells.
    add_line_chart(ws, "Geometric spreading power dB", "dB", 1, [5, 9], header_row, first_row, last_row, "A270")
    add_line_chart(ws, "Coherent Fresnel-zone gain", "dB", 1, [12, 15], header_row, first_row, last_row, "I270")
    add_line_chart(ws, "Total VHF dB: constant vs frequency-dependent response", "dB", 1, [21, 22], header_row, first_row, last_row, "Q270")

    # Notes for the user/mentor.
    ws["A288"] = "Use note"
    ws["A288"].font = Font(name="Arial", bold=True, color="FFFFFF")
    ws["A288"].fill = PatternFill("solid", fgColor=blue)
    ws["B288"] = (
        "These are native Excel charts connected to formula-backed cells. If you change assumptions in B12:B22 "
        "or upstream Model_Data/Inputs values, the table and charts recalculate in Excel."
    )
    ws["B288"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("B288:W290")

    # Column widths and freeze panes.
    widths = {
        "A": 18,
        "B": 18,
        "C": 18,
        "D": 34,
        "E": 16,
        "F": 18,
        "G": 18,
        "H": 18,
        "I": 16,
        "J": 18,
        "K": 18,
        "L": 18,
        "M": 18,
        "N": 18,
        "O": 18,
        "P": 18,
        "Q": 18,
        "R": 18,
        "S": 18,
        "T": 18,
        "U": 20,
        "V": 22,
        "W": 22,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row_idx in [2, 6, 7, 8, 9, 288]:
        ws.row_dimensions[row_idx].height = 38
    ws.freeze_panes = "A26"

    # Make Excel recalculate formulas and charts on open.
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
