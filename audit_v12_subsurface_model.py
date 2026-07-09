from __future__ import annotations

import math
from pathlib import Path

from openpyxl import load_workbook


XLSX = Path(r"C:\Users\tyboy\OneDrive\Documents\Nasa project\outputs\europa_ice_subsurface_simulation\v12.xlsx")
TOL = 1e-6


def is_number(value):
    return isinstance(value, (int, float)) and not isinstance(value, bool) and math.isfinite(value)


def close(a, b, tol=TOL):
    return is_number(a) and is_number(b) and abs(a - b) <= tol * max(1, abs(a), abs(b))


def col_number(letter):
    n = 0
    for ch in letter:
        n = n * 26 + ord(ch.upper()) - 64
    return n


def values(ws, row):
    return {idx: ws.cell(row, idx).value for idx in range(1, ws.max_column + 1)}


def count_bad(cases):
    return sum(1 for ok in cases if not ok)


def main():
    wb_values = load_workbook(XLSX, data_only=True, read_only=True)
    wb_formulas = load_workbook(XLSX, data_only=False, read_only=True)

    live = wb_values["Subsurface_Live_Data"]
    scen = wb_values["Subsurface_Scenario_Data"]
    radar = wb_values["Subsurface_Radargram_Data"]
    chart = wb_values["Subsurface_Chart_Data"]
    chart_formulas = wb_formulas["Subsurface_Chart_Data"]
    formulas = wb_formulas["Subsurface_Formulas"]
    inputs = wb_values["Subsurface_Inputs"]
    material = wb_values["Subsurface_Materials_Evidence"]

    report = []

    # Workbook/formula health.
    error_cells = []
    for ws in wb_values.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("#"):
                    error_cells.append((ws.title, cell.coordinate, cell.value))
    report.append(("cached_error_values", len(error_cells), error_cells[:10]))

    chart_formula_cells = 0
    chart_static_cells = 0
    for row in chart_formulas.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            if isinstance(cell.value, str) and cell.value.startswith("="):
                chart_formula_cells += 1
            else:
                chart_static_cells += 1
    report.append(("subsurface_chart_data_formula_cells", chart_formula_cells, None))
    report.append(("subsurface_chart_data_static_label_cells", chart_static_cells, None))

    # Logical model relationships in the live data.
    upper_positive = []
    depth_order = []
    elevation_order = []
    delay_order = []
    elevation_math = []
    delay_math = []
    echo_math = []
    margin_math = []
    ratio_math = []
    flag_logic = []

    n_ice = inputs["C34"].value
    speed = inputs["C35"].value
    attenuation = inputs["C36"].value
    upper_base = inputs["C37"].value
    lens_base = inputs["C38"].value
    lens_bonus = inputs["C39"].value
    ocean_base = inputs["C40"].value
    roughness = inputs["C41"].value
    lens_threshold = inputs["C42"].value
    detect_threshold = inputs["C43"].value

    ocean_elev = [live.cell(r, col_number("L")).value for r in range(2, 243)]
    avg_ocean_elev = sum(v for v in ocean_elev if is_number(v)) / len(ocean_elev)
    ocean_depths = [live.cell(r, col_number("K")).value for r in range(2, 243)]
    avg_ocean_depth = sum(v for v in ocean_depths if is_number(v)) / len(ocean_depths)

    for r in range(2, 243):
        surface = live.cell(r, col_number("B")).value
        upper_depth = live.cell(r, col_number("F")).value
        upper_elev = live.cell(r, col_number("G")).value
        lens_strength = live.cell(r, col_number("H")).value
        lens_depth = live.cell(r, col_number("I")).value
        lens_elev = live.cell(r, col_number("J")).value
        ocean_depth = live.cell(r, col_number("K")).value
        ocean_elev_value = live.cell(r, col_number("L")).value
        upper_delay = live.cell(r, col_number("M")).value
        lens_delay = live.cell(r, col_number("N")).value
        ocean_delay = live.cell(r, col_number("O")).value
        upper_echo = live.cell(r, col_number("P")).value
        lens_echo = live.cell(r, col_number("Q")).value
        ocean_echo = live.cell(r, col_number("R")).value
        thickness_variation = live.cell(r, col_number("S")).value
        ratio = live.cell(r, col_number("T")).value
        lens_flag = live.cell(r, col_number("U")).value
        ocean_flag = live.cell(r, col_number("V")).value
        threshold = live.cell(r, col_number("W")).value
        lens_margin = live.cell(r, col_number("X")).value
        ocean_margin = live.cell(r, col_number("Y")).value
        best_margin = live.cell(r, col_number("Z")).value
        zero_margin = live.cell(r, col_number("AA")).value

        upper_positive.append(upper_depth > 0 and lens_depth > 0 and ocean_depth > 0)
        depth_order.append(upper_depth < lens_depth < ocean_depth)
        elevation_order.append(surface > upper_elev > lens_elev > ocean_elev_value)
        delay_order.append(0 < upper_delay < lens_delay < ocean_delay)
        elevation_math.extend(
            [
                close(upper_elev, surface - upper_depth),
                close(lens_elev, surface - lens_depth),
                close(ocean_elev_value, surface - ocean_depth),
            ]
        )
        delay_math.extend(
            [
                close(upper_delay, 2 * n_ice * upper_depth / speed * 1_000_000),
                close(lens_delay, 2 * n_ice * lens_depth / speed * 1_000_000),
                close(ocean_delay, 2 * n_ice * ocean_depth / speed * 1_000_000),
            ]
        )
        echo_math.extend(
            [
                close(upper_echo, upper_base - 2 * attenuation * (upper_depth / 1000)),
                close(lens_echo, lens_base + lens_bonus * lens_strength - 2 * attenuation * (lens_depth / 1000)),
                close(
                    ocean_echo,
                    ocean_base
                    - 2 * attenuation * (ocean_depth / 1000)
                    - roughness * abs(ocean_elev_value - avg_ocean_elev),
                ),
            ]
        )
        margin_math.extend(
            [
                close(threshold, detect_threshold),
                close(lens_margin, lens_echo - threshold),
                close(ocean_margin, ocean_echo - threshold),
                close(best_margin, max(lens_echo, ocean_echo) - threshold),
                close(zero_margin, 0),
            ]
        )
        ratio_math.extend(
            [
                close(thickness_variation, ocean_depth - avg_ocean_depth),
                close(ratio, ocean_depth / upper_depth),
            ]
        )
        flag_logic.extend(
            [
                (lens_flag == "Lens visible") == (lens_strength >= lens_threshold),
                (ocean_flag == "Likely visible") == (ocean_echo > -45),
            ]
        )

    live_checks = [
        ("positive_layer_depths", count_bad(upper_positive)),
        ("depth_order_upper_lens_ocean", count_bad(depth_order)),
        ("elevation_order_surface_upper_lens_ocean", count_bad(elevation_order)),
        ("delay_order_upper_lens_ocean", count_bad(delay_order)),
        ("elevation_formulas", count_bad(elevation_math)),
        ("delay_formulas", count_bad(delay_math)),
        ("echo_formulas", count_bad(echo_math)),
        ("margin_formulas", count_bad(margin_math)),
        ("variation_ratio_formulas", count_bad(ratio_math)),
        ("visibility_flag_logic", count_bad(flag_logic)),
    ]
    report.extend((name, bad, None) for name, bad in live_checks)

    # Scenario/radargram formulas and graph-source logic.
    uncertainty = inputs["C44"].value
    thin_mult = inputs["C45"].value
    thick_mult = inputs["C46"].value
    no_ocean_echo = inputs["C47"].value
    jitter = inputs["C56"].value
    clutter_band = inputs["C57"].value

    scenario_checks = []
    radar_checks = []
    radar_expected_na = 0
    radar_actual_na = 0
    for r in range(2, 243):
        x = scen.cell(r, 1).value
        mean = scen.cell(r, 2).value
        low = scen.cell(r, 3).value
        high = scen.cell(r, 4).value
        thin = scen.cell(r, 5).value
        medium = scen.cell(r, 6).value
        thick = scen.cell(r, 7).value
        ocean_margin = scen.cell(r, 8).value
        no_ocean_margin = scen.cell(r, 9).value
        zero = scen.cell(r, 10).value
        live_ocean_depth = live.cell(r, col_number("K")).value
        live_ocean_margin = live.cell(r, col_number("Y")).value
        scenario_checks.extend(
            [
                close(mean, live_ocean_depth),
                close(low, max(0, mean - uncertainty)),
                close(high, mean + uncertainty),
                close(thin, mean * thin_mult),
                close(medium, mean),
                close(thick, mean * thick_mult),
                thin < medium < thick,
                low <= mean <= high,
                close(ocean_margin, live_ocean_margin),
                close(no_ocean_margin, no_ocean_echo - detect_threshold + 2 * math.sin(2 * math.pi * x / 42)),
                close(zero, 0),
            ]
        )

        r_x = radar.cell(r, 1).value
        clutter_lower = radar.cell(r, 2).value
        clutter_upper = radar.cell(r, 3).value
        shallow_return = radar.cell(r, 4).value
        lens_return = radar.cell(r, 5).value
        ocean_return = radar.cell(r, 6).value
        noise_floor = radar.cell(r, 7).value
        expected_lens_visible = live.cell(r, col_number("H")).value >= lens_threshold
        if not expected_lens_visible:
            radar_expected_na += 1
        if isinstance(lens_return, str) and lens_return.startswith("#"):
            radar_actual_na += 1
        radar_checks.extend(
            [
                close(r_x, x),
                close(clutter_lower, 0),
                close(clutter_upper, clutter_band + 1.5 * math.sin(2 * math.pi * x / 35)),
                close(shallow_return, live.cell(r, col_number("M")).value + jitter * math.sin(2 * math.pi * x / 31)),
                (
                    isinstance(lens_return, str)
                    and lens_return.startswith("#")
                    if not expected_lens_visible
                    else close(lens_return, live.cell(r, col_number("N")).value + jitter * math.sin(2 * math.pi * x / 27))
                ),
                close(ocean_return, live.cell(r, col_number("O")).value + jitter * math.sin(2 * math.pi * x / 47)),
                close(noise_floor, -55 + 3 * math.sin(2 * math.pi * x / 19)),
            ]
        )

    report.append(("scenario_formulas_and_order", count_bad(scenario_checks), None))
    report.append(("radargram_formulas", count_bad(radar_checks), None))
    report.append(("radargram_expected_na_lens_points", radar_expected_na, None))
    report.append(("radargram_actual_na_lens_points", radar_actual_na, None))

    # Chart data links should point to formulas, not pasted values, in plotted numeric cells.
    chart_blocks = [
        ("Icy layers", "A3:H243"),
        ("Scenarios", "A248:F488"),
        ("Uncertainty", "A493:F733"),
        ("Ocean control", "A738:F978"),
        ("Radargram", "A983:H1223"),
        ("Detectability", "A1228:F1468"),
        ("Materials", "A1473:B1477"),
        ("Evidence", "A1492:B1495"),
    ]
    for label, range_text in chart_blocks:
        formula_count = 0
        bad_values = 0
        cells = chart_formulas[range_text]
        for row in cells:
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
                else:
                    bad_values += 1
        report.append((f"chart_block_{label}_formula_cells", formula_count, None))
        report.append((f"chart_block_{label}_nonformula_cells", bad_values, None))

    # Formula sheet has graph documentation.
    graph_doc_rows = []
    for r in range(1, formulas.max_row + 1):
        if formulas.cell(r, 1).value in [
            "Subsurface Truth Model: Icy Layers",
            "Scenario Comparison: Thin / Medium / Thick Ice",
            "Boundary Uncertainty Band",
            "Ocean Model vs No-Ocean Control",
            "Radargram-Style Return Timing With Clutter",
            "Detectability Margin vs Threshold",
            "Reflection Strength by Material / Interface",
            "Cross-Instrument Evidence Score",
        ]:
            graph_doc_rows.append(r)
    report.append(("documented_graph_formula_rows", len(graph_doc_rows), graph_doc_rows))

    # Useful ranges for summary.
    def minmax(ws, col, rows=range(2, 243)):
        vals = [ws.cell(r, col_number(col)).value for r in rows]
        vals = [v for v in vals if is_number(v)]
        return min(vals), max(vals), sum(vals) / len(vals)

    summary = {
        "upper_depth_m_min_max_avg": minmax(live, "F"),
        "lens_depth_m_min_max_avg": minmax(live, "I"),
        "ocean_depth_m_min_max_avg": minmax(live, "K"),
        "ocean_margin_db_min_max_avg": minmax(live, "Y"),
        "lens_margin_db_min_max_avg": minmax(live, "X"),
        "evidence_scores": [material.cell(r, 2).value for r in range(10, 14)],
    }

    print("AUDIT_REPORT")
    for name, value, detail in report:
        print(f"{name}: {value}" + (f" detail={detail}" if detail else ""))
    print("SUMMARY")
    for key, value in summary.items():
        print(f"{key}: {value}")


if __name__ == "__main__":
    main()
