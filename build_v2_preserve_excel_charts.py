from pathlib import Path
import csv
import json

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


SOURCE_XLSX = Path(
    r"C:\Users\tyboy\Downloads\parabolic-motion-radar-model-baseline-and-runs-dashboard-native-excel-charts-fixed (1).xlsx"
)
SIM_DIR = Path(r"C:\Users\tyboy\OneDrive\Documents\Nasa project\outputs\europa_ice_subsurface_simulation")
OUTPUT_XLSX = SIM_DIR / "v2.xlsx"


def chart_title(chart):
    try:
        rich = chart.title.tx.rich
        if rich and rich.p and rich.p[0].r:
            return rich.p[0].r[0].t or ""
    except Exception:
        return ""
    return ""


def delete_sheet_if_exists(workbook, sheet_name):
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]


def add_table_style(ws, max_row, max_col):
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D4DAE3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def maybe_number(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return value


def add_subsurface_model_sheet(wb):
    csv_path = SIM_DIR / "europa_ice_subsurface_model.csv"
    ws = wb.create_sheet("Subsurface_Model")
    with csv_path.open(newline="", encoding="utf-8") as fh:
        reader = csv.reader(fh)
        for row in reader:
            ws.append([maybe_number(v) for v in row])
    add_table_style(ws, ws.max_row, ws.max_column)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {
        "A": 12,
        "B": 12,
        "C": 18,
        "D": 22,
        "E": 26,
        "F": 25,
        "G": 25,
        "H": 27,
        "I": 24,
        "J": 26,
        "K": 25,
        "L": 23,
        "M": 26,
        "N": 27,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.000"


def add_assumptions_sheet(wb):
    assumptions = json.loads((SIM_DIR / "europa_ice_subsurface_assumptions.json").read_text(encoding="utf-8"))
    ws = wb.create_sheet("Subsurface_Assumptions")
    ws.merge_cells("A1:C1")
    ws["A1"] = "Europa Ice Subsurface Simulation Assumptions"
    ws["A1"].fill = PatternFill("solid", fgColor="1F2937")
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws.append([])
    ws.append(["Item", "Value", "Why it matters"])
    rows = [
        ["Surface material", assumptions["surface"], "Keeps the generated topography as icy terrain, not rock."],
        [
            "Upper internal ice layer",
            f"{assumptions['upper_ice_layer_mean_depth_m']} m mean depth",
            "Represents shallow ice structure, fractures, impurities, or layer boundaries.",
        ],
        [
            "Possible briny/warm-ice lens",
            assumptions["possible_briny_warm_lens_depth_m"],
            "A localized material contrast inside ice; treated as possible, not proven.",
        ],
        [
            "Possible ice-ocean boundary",
            f"{assumptions['nominal_ice_ocean_boundary_depth_m']} m nominal depth",
            "Deep reflector where ice may meet liquid water or warmer basal ice.",
        ],
        ["Ice refractive index", assumptions["ice_refractive_index"], "Converts depth into radar two-way delay."],
        [
            "One-way ice attenuation",
            f"{assumptions['attenuation_db_per_km_one_way']} dB/km",
            "Weakens deeper radar returns as the signal travels through ice.",
        ],
        ["Model warning", assumptions["note"], "Prevents this from being mistaken for measured mission data."],
    ]
    for row in rows:
        ws.append(row)
    ws["A3"].fill = ws["B3"].fill = ws["C3"].fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[3]:
        cell.font = Font(bold=True, color="1F2937")
    thin = Side(style="thin", color="D4DAE3")
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=3):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 62
    ws.column_dimensions["C"].width = 70
    for i in range(4, ws.max_row + 1):
        ws.row_dimensions[i].height = 34


def add_graphs_sheet(wb):
    ws = wb.create_sheet("Subsurface_Graphs")
    ws["A1"] = "Europa Ice Subsurface Simulation"
    ws["A1"].font = Font(bold=True, size=14, color="1F2937")
    ws["A2"] = "Python-generated previews added after checking that the outputs are coherent. Existing dashboard graphs are unchanged."
    ws["A2"].alignment = Alignment(wrap_text=True)
    rows = [
        ["Graph", "What it shows", "Interpretation", "Source"],
        [
            "Subsurface truth model",
            "Synthetic icy layers below the generated surface topography.",
            "Topography is ice; deeper features are internal ice, possible warm/briny ice, and possible ocean boundary.",
            "Subsurface_Model",
        ],
        [
            "Simulated radargram",
            "Radar two-way delay after surface return.",
            "Brighter bands are stronger reflectors; deeper returns are weaker because of ice attenuation.",
            "Python preview",
        ],
        [
            "Delay and echo strength",
            "Layer return timing and relative signal strength.",
            "Ocean boundary appears late in time; lens strength is localized and hypothetical.",
            "Subsurface_Model",
        ],
        [],
        ["Embedded preview image", "File used", "Workbook section", "Purpose"],
        ["Subsurface truth model", "01_europa_ice_subsurface_truth.png", "Left", "Shows the hidden icy layers."],
        [
            "Radargram + response charts",
            "02_simulated_ice_radargram.png and 03_layer_delays_and_echo_strength.png",
            "Right / below",
            "Shows simulated radar return timing and strength.",
        ],
    ]
    for row in rows:
        ws.append(row)
    for row_number in (4, 9):
        for cell in ws[row_number]:
            cell.fill = PatternFill("solid", fgColor="1F4E79" if row_number == 4 else "D9EAF7")
            cell.font = Font(bold=True, color="FFFFFF" if row_number == 4 else "1F2937")
    thin = Side(style="thin", color="D4DAE3")
    for row in ws.iter_rows(min_row=4, max_row=11, max_col=4):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 54
    ws.column_dimensions["C"].width = 62
    ws.column_dimensions["D"].width = 24

    image_specs = [
        ("01_europa_ice_subsurface_truth.png", "A14", 900, 552),
        ("02_simulated_ice_radargram.png", "H14", 900, 540),
        ("03_layer_delays_and_echo_strength.png", "A47", 900, 672),
    ]
    for file_name, anchor, width, height in image_specs:
        image = ExcelImage(str(SIM_DIR / file_name))
        image.width = width
        image.height = height
        ws.add_image(image, anchor)


def main():
    wb = load_workbook(SOURCE_XLSX)

    dashboard = wb["Dashboard"]
    fixed = False
    for chart in dashboard._charts:
        if chart_title(chart) == "Terrain Baseline: Total Radar Elevation Error":
            chart.legend = None
            fixed = True
    if not fixed:
        raise RuntimeError("Could not find Terrain Baseline chart to hide the broken legend.")

    for sheet_name in ["Subsurface_Model", "Subsurface_Assumptions", "Subsurface_Graphs"]:
        delete_sheet_if_exists(wb, sheet_name)

    add_subsurface_model_sheet(wb)
    add_assumptions_sheet(wb)
    add_graphs_sheet(wb)

    wb.save(OUTPUT_XLSX)
    print(OUTPUT_XLSX)


if __name__ == "__main__":
    main()
