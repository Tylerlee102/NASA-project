from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import BarChart, ScatterChart, Series, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


BASE_XLSX = Path(r"C:\Users\tyboy\OneDrive\Documents\Nasa project\outputs\europa_ice_subsurface_simulation\v6.xlsx")
OUT_XLSX = Path(r"C:\Users\tyboy\OneDrive\Documents\Nasa project\outputs\europa_ice_subsurface_simulation\v7.xlsx")


COLORS = {
    "navy": "1F4E79",
    "dark": "1F2937",
    "muted": "526174",
    "blue": "2F6DB3",
    "green": "219653",
    "orange": "E76F24",
    "gold": "D49A22",
    "purple": "7048B8",
    "red": "C43A3A",
    "gray": "808080",
    "light_blue": "D9EAF7",
    "grid": "D4DAE3",
}


def chart_title(chart):
    try:
        rich = chart.title.tx.rich
        if rich and rich.p and rich.p[0].r:
            return rich.p[0].r[0].t or ""
    except Exception:
        return ""
    return ""


def keep_existing_chart_fix(wb):
    for chart in wb["Dashboard"]._charts:
        if chart_title(chart) == "Terrain Baseline: Total Radar Elevation Error":
            chart.legend = None


def set_calc_mode(wb):
    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass


def style_header(ws, row, min_col, max_col):
    for col in range(min_col, max_col + 1):
        cell = ws.cell(row, col)
        cell.fill = PatternFill("solid", fgColor=COLORS["navy"])
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def add_borders(ws, min_row, max_row, min_col, max_col):
    thin = Side(style="thin", color=COLORS["grid"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def set_dashboard_sizing(ws):
    widths = {
        "A": 31,
        "B": 13,
        "C": 9,
        "D": 34,
        "E": 14,
        "F": 18,
        "G": 14,
        "H": 13,
        "I": 28,
        "J": 16,
        "K": 13,
        "L": 13,
        "M": 13,
        "N": 13,
        "O": 13,
        "P": 13,
        "Q": 13,
        "R": 13,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    heights = {
        1: 21.0,
        2: 28.0,
        4: 15.75,
        6: 26.25,
        7: 26.25,
        8: 26.25,
        9: 26.25,
        10: 26.25,
        11: 26.25,
        13: 26.25,
        14: 26.25,
    }
    for row, height in heights.items():
        ws.row_dimensions[row].height = height


def add_line_series(chart, ws, x_col, y_col, label, color, width=22000):
    xvalues = Reference(ws, min_col=x_col, min_row=2, max_row=242)
    yvalues = Reference(ws, min_col=y_col, min_row=2, max_row=242)
    series = Series(yvalues, xvalues, title=label)
    series.graphicalProperties.line.solidFill = color
    series.graphicalProperties.line.width = width
    chart.series.append(series)


def scatter_chart(title, y_title, ws, specs, legend=True):
    chart = ScatterChart()
    chart.title = title
    chart.scatterStyle = "line"
    chart.width = 15
    chart.height = 7.5
    chart.x_axis.title = "Along-track position x (km)"
    chart.y_axis.title = y_title
    chart.x_axis.majorGridlines = None
    if legend:
        chart.legend.position = "b"
    else:
        chart.legend = None
    for col, label, color in specs:
        add_line_series(chart, ws, 1, col, label, color)
    return chart


def bar_chart(title, y_title, ws, cats_col, values_col, min_row, max_row, color):
    chart = BarChart()
    chart.title = title
    chart.type = "bar"
    chart.width = 15
    chart.height = 7.5
    chart.y_axis.title = y_title
    chart.legend = None
    data = Reference(ws, min_col=values_col, min_row=min_row - 1, max_row=max_row)
    cats = Reference(ws, min_col=cats_col, min_row=min_row, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    if chart.series:
        chart.series[0].graphicalProperties.solidFill = color
    return chart


def build_dashboard(wb):
    if "Subsurface_Dashboard" in wb.sheetnames:
        del wb["Subsurface_Dashboard"]
    ws = wb.create_sheet("Subsurface_Dashboard", 1)
    live = wb["Subsurface_Live_Data"]
    scen = wb["Subsurface_Scenario_Data"]
    material = wb["Subsurface_Materials_Evidence"]
    radargram = wb["Subsurface_Radargram_Data"]

    set_dashboard_sizing(ws)
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Europa Ice Subsurface Interface"
    ws["A1"].font = Font(bold=True, size=16, color=COLORS["dark"])
    ws["A2"] = "Live subsurface outputs from Subsurface_Inputs. The model is synthetic: it helps test what radar might see below icy topography, but it does not prove real Europa layers by itself."
    ws["A2"].font = Font(color=COLORS["muted"])
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")

    ws["A4"] = "Key Subsurface Results"
    ws["A4"].font = Font(bold=True, size=12, color=COLORS["dark"])
    left_rows = [
        ["Output", "Value", "Unit", "What it means"],
        ["Average bottom depth", "=AVERAGE(Subsurface_Live_Data!K2:K242)", "m", "Mean proposed ice-ocean boundary depth."],
        ["Boundary uncertainty band", "=Subsurface_Inputs!$C$44", "m", "Plus/minus range around bottom reflector."],
        ["Best ocean echo margin", "=MAX(Subsurface_Live_Data!Y2:Y242)", "dB", "How far deep reflector rises above/below threshold."],
        ["Best lens echo margin", "=MAX(Subsurface_Live_Data!X2:X242)", "dB", "How far internal lens return rises above/below threshold."],
        ["Likely visible ocean samples", '=COUNTIF(Subsurface_Live_Data!V2:V242,"Likely visible")', "count", "Pass samples clearing the detection threshold."],
        ["Total evidence support", "=Subsurface_Materials_Evidence!D14", "%", "Radar + thermal + composition + magnetic/plasma support."],
    ]
    for r_idx, row in enumerate(left_rows, start=5):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(r_idx, c_idx).value = value
    style_header(ws, 5, 1, 4)
    add_borders(ws, 5, 11, 1, 4)
    for row in range(6, 12):
        ws.cell(row, 2).number_format = "0.000"

    ws["F4"] = "Realism / Proof Strength"
    ws["F4"].font = Font(bold=True, size=12, color=COLORS["dark"])
    right_rows = [
        ["Score", "Value", "Meaning", "Improve by"],
        ["Concept realism", 0.70, "Good for showing layered-ice radar behavior.", "Compare more scenarios."],
        ["Europa prediction realism", "=MIN(0.75,0.25+Subsurface_Materials_Evidence!D14/100*0.45)", "Plausible, but synthetic and not measured.", "Use published constraints or real mission data."],
        ["Proof strength", "=MIN(0.60,0.15+Subsurface_Materials_Evidence!D14/100*0.35)", "Supports a hypothesis; does not prove a real layer.", "Validate with radar, thermal, composition, and magnetic/plasma data."],
    ]
    for r_idx, row in enumerate(right_rows, start=5):
        for c_idx, value in enumerate(row, start=6):
            ws.cell(r_idx, c_idx).value = value
    style_header(ws, 5, 6, 9)
    add_borders(ws, 5, 8, 6, 9)
    for row in range(6, 9):
        ws.cell(row, 7).number_format = "0%"

    ws["F10"] = "Read this first"
    ws["F10"].font = Font(bold=True, color=COLORS["dark"])
    ws["G10"] = "Positive detectability margin means the simulated return is above the chosen threshold. Change assumptions on Subsurface_Inputs."
    ws["G10"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("G10:I11")
    add_borders(ws, 10, 11, 6, 9)

    chart_specs = [
        (
            "A17",
            scatter_chart(
                "Subsurface Truth Model: Icy Layers",
                "Elevation relative to reference (m)",
                live,
                [(2, "Icy top surface", COLORS["orange"]), (7, "Shallow ice layer", COLORS["green"]), (10, "Warm/briny lens", COLORS["gold"]), (12, "Ice-ocean boundary", COLORS["purple"])],
            ),
        ),
        (
            "J17",
            scatter_chart(
                "Scenario Comparison: Thin / Medium / Thick Ice",
                "Depth to possible boundary (m)",
                scen,
                [(5, "Thin shell", COLORS["green"]), (6, "Medium shell", COLORS["purple"]), (7, "Thick shell", COLORS["red"])],
            ),
        ),
        (
            "A38",
            scatter_chart(
                "Boundary Uncertainty Band",
                "Depth to possible boundary (m)",
                scen,
                [(3, "Lower bound", COLORS["green"]), (2, "Mean boundary", COLORS["purple"]), (4, "Upper bound", COLORS["red"])],
            ),
        ),
        (
            "J38",
            scatter_chart(
                "Ocean Model vs No-Ocean Control",
                "Margin above threshold (dB)",
                scen,
                [(8, "Ocean model margin", COLORS["purple"]), (9, "No-ocean control margin", COLORS["gray"]), (10, "Zero threshold", COLORS["red"])],
            ),
        ),
        (
            "A59",
            scatter_chart(
                "Radargram-Style Return Timing With Clutter",
                "Two-way delay after surface return (us)",
                radargram,
                [(3, "Surface clutter upper", COLORS["gray"]), (4, "Shallow ice return", COLORS["green"]), (5, "Warm/briny lens return", COLORS["gold"]), (6, "Ocean boundary return", COLORS["purple"])],
            ),
        ),
        (
            "J59",
            scatter_chart(
                "Detectability Margin vs Threshold",
                "Margin above threshold (dB)",
                live,
                [(24, "Lens echo margin", COLORS["gold"]), (25, "Ocean echo margin", COLORS["purple"]), (27, "Zero margin threshold", COLORS["red"])],
            ),
        ),
        (
            "A80",
            bar_chart(
                "Reflection Strength by Material / Interface",
                "Pre-depth-loss reflection strength (dB)",
                material,
                1,
                2,
                2,
                6,
                COLORS["blue"],
            ),
        ),
        (
            "J80",
            bar_chart(
                "Cross-Instrument Evidence Score",
                "Support score (%)",
                material,
                1,
                2,
                10,
                13,
                COLORS["green"],
            ),
        ),
    ]
    for anchor, chart in chart_specs:
        ws.add_chart(chart, anchor)

    ws.freeze_panes = "A5"


def main():
    wb = load_workbook(BASE_XLSX)
    set_calc_mode(wb)
    keep_existing_chart_fix(wb)
    build_dashboard(wb)
    wb.save(OUT_XLSX)
    print(OUT_XLSX)


if __name__ == "__main__":
    main()
